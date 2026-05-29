"""
format_rack_validation.py
-------------------------
Automates the rack validation workbook formatting we built up:

  1. Renames file to the rack number (e.g. 7417.xlsx).
  2. Splits the 'LLDP Mismatch + Link Down' sheet into 'Downlink' (LLDP
     Status = DOWN) and 'Mismatch' (LLDP Status = MISMATCH).
  3. Cleans columns on Downlink, Mismatch, and Optic Errors.
  4. Matches device+port against a CFAB cutsheet to populate the cable
     path (long path = patch panels; short path = direct cable).
  5. Splits Downlink / Mismatch / Optic Errors into T3-T2 (long) and
     T2-T1-T0 (short) sub-tabs.
  6. Adds Device B side match for both Mismatch tabs, prefixed with
     "Current" and highlighted pink.
  7. Greys-out Optic Errors rows whose port also appears in Downlink.
  8. Adds a 'Note' column, applies thin borders + autofilter, bold
     yellow frozen headers (Rx Power column also frozen on Optic tabs).
  9. Rebuilds the Summary tab with live COUNTA formulas and a Total.

Usage:
    python format_rack_validation.py

The script prompts for one cutsheet path and one or more input files.
"""

from __future__ import annotations
from copy import copy
from pathlib import Path
import sys

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
PINK   = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
THIN   = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LIGHT_GREY = "A6A6A6"


# ---------------------------------------------------------------------------
# Cutsheet lookup
# ---------------------------------------------------------------------------
def build_cutsheet_lookup(path: Path) -> tuple[dict, dict]:
    """
    Load the CFAB cutsheet and build a lookup keyed by 'device port'.

    Row formats (all long-path rows have col4 populated):
      3-PP long: col0=A dev+port, col1=A rack, col2=PP1, col3=PP2, col4=PP3,
                 col5=B dev+port, col6=B rack, col7=NaN
      4-PP long: col0=A dev+port, col1=A rack, col2=PP1, col3=PP2, col4=PP3,
                 col5=PP4,        col6=B dev+port, col7=B rack
      Short:     col0=A dev+port, col1=A rack, col2=B dev+port, col3=B rack,
                 col4..col7=NaN

    We register BOTH endpoints so a match works from either side.
    Also returns a device_rack_lookup: device_name -> rack for fallback use.
    """
    df = pd.read_excel(path, sheet_name=0, header=None)
    lookup: dict = {}
    device_rack_lookup: dict = {}

    def add(key, device_rack, pp_list, peer_device, peer_rack):
        if not key or key in lookup:
            return
        lookup[key] = {
            "device_rack": device_rack,
            "pp1": pp_list[0] if len(pp_list) > 0 else None,
            "pp2": pp_list[1] if len(pp_list) > 1 else None,
            "pp3": pp_list[2] if len(pp_list) > 2 else None,
            "pp4": pp_list[3] if len(pp_list) > 3 else None,
            "peer_device": peer_device,
            "peer_rack": peer_rack,
        }
        # Register device name -> rack for fallback
        dev_name = key.split(" ", 1)[0] if key else None
        if dev_name and dev_name not in device_rack_lookup and pd.notna(device_rack):
            device_rack_lookup[dev_name] = device_rack

    for _, row in df.iterrows():
        is_long = pd.notna(row[4])
        if is_long:
            # Detect 3-PP vs 4-PP: if col5 starts with "PP" it is the 4th patch panel
            has_4pp = pd.notna(row[5]) and isinstance(row[5], str) and row[5].strip().startswith("PP")
            if has_4pp:
                # 4-PP path: col2-col5 are PPs, col6=B dev+port, col7=B rack
                a_key  = str(row[0]) if pd.notna(row[0]) else None
                b_key  = str(row[6]) if pd.notna(row[6]) else None
                add(a_key, row[1], [row[2], row[3], row[4], row[5]], row[6], row[7])
                add(b_key, row[7], [row[5], row[4], row[3], row[2]], row[0], row[1])
            else:
                # 3-PP path: col2-col4 are PPs, col5=B dev+port, col6=B rack
                a_key  = str(row[0]) if pd.notna(row[0]) else None
                b_key  = str(row[5]) if pd.notna(row[5]) else None
                add(a_key, row[1], [row[2], row[3], row[4]], row[5], row[6])
                add(b_key, row[6], [row[4], row[3], row[2]], row[0], row[1])
        else:
            # Short / direct cable: col0=A dev+port, col1=A rack, col2=B dev+port, col3=B rack
            a_key = str(row[0]) if pd.notna(row[0]) else None
            b_key = str(row[2]) if pd.notna(row[2]) else None
            add(a_key, row[1], [], row[2], row[3])
            add(b_key, row[3], [], row[0], row[1])

    return lookup, device_rack_lookup


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------
def split_device_port(s):
    """Split a combined 'device_name port' string into (name, port).
    Device names never contain spaces, so we split at the first space."""
    if not s or not isinstance(s, str):
        return None, None
    parts = s.strip().split(" ", 1)
    return (parts[0], parts[1]) if len(parts) == 2 else (parts[0], None)


def style_cell(cell, ref):
    cell.font = copy(ref.font)
    cell.fill = copy(ref.fill)
    cell.border = copy(ref.border)
    cell.alignment = copy(ref.alignment)
    cell.number_format = ref.number_format


def header_indices(ws) -> dict:
    return {c.value: i + 1 for i, c in enumerate(ws[1])}


def delete_columns_by_name(ws, names: list[str]):
    for name in names:
        hdrs = [c.value for c in ws[1]]
        if name in hdrs:
            ws.delete_cols(hdrs.index(name) + 1)


def rack_number_from(ws) -> str:
    """Pull the first number before ':' in the Device A Rack column."""
    hdrs = header_indices(ws)
    if "Device A Rack" not in hdrs:
        return "output"
    col = hdrs["Device A Rack"]
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v:
            return str(v).split(":")[0].strip()
    return "output"


def autofit_columns(ws, min_w=12, max_w=40):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, min_w), max_w)


# ---------------------------------------------------------------------------
# Step 1 — Split LLDP sheet by status
# ---------------------------------------------------------------------------
def split_lldp_sheet(wb):
    src_name = "LLDP Mismatch + Link Down"
    if src_name not in wb.sheetnames:
        return
    src = wb[src_name]
    hdrs = [c.value for c in src[1]]
    status_idx = hdrs.index("LLDP Status")

    header_styles = [{
        "font": copy(c.font), "fill": copy(c.fill),
        "border": copy(c.border), "alignment": copy(c.alignment),
        "number_format": c.number_format,
    } for c in src[1]]
    col_widths = {k: v.width for k, v in src.column_dimensions.items()}

    down, mismatch = [], []
    for row in src.iter_rows(min_row=2, values_only=True):
        if row[status_idx] == "DOWN":
            down.append(row)
        elif row[status_idx] == "MISMATCH":
            mismatch.append(row)

    def build(name, rows):
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        ws.append(hdrs)
        for i, c in enumerate(ws[1], start=1):
            s = header_styles[i - 1]
            c.font = s["font"]; c.fill = s["fill"]; c.border = s["border"]
            c.alignment = s["alignment"]; c.number_format = s["number_format"]
        for r in rows:
            ws.append(r)
        for k, w in col_widths.items():
            if w:
                ws.column_dimensions[k].width = w
        ws.freeze_panes = "A2"

    build("Downlink", down)
    build("Mismatch", mismatch)
    del wb[src_name]


# ---------------------------------------------------------------------------
# Step 2 — Column cleanup on the three data sheets
# ---------------------------------------------------------------------------
def clean_columns(wb):
    # Downlink: keep only Device A Name + Device A Port
    if "Downlink" in wb.sheetnames:
        delete_columns_by_name(wb["Downlink"], [
            "Device A Rack", "Expected Device B Rack", "Expected Device B Name",
            "Expected Device B Port", "Device B Rack", "Device B Name",
            "Device B Port", "LLDP Status", "Patch Panel Matrix",
        ])

    # Mismatch: keep Device A Name/Port + Device B Name/Port
    if "Mismatch" in wb.sheetnames:
        delete_columns_by_name(wb["Mismatch"], [
            "Device A Rack", "Expected Device B Rack", "Expected Device B Name",
            "Expected Device B Port", "Device B Rack",
            "LLDP Status", "Patch Panel Matrix",
        ])

    # Optic Errors: drop, then move Rx Power to first column
    if "Optic Errors" in wb.sheetnames:
        ws = wb["Optic Errors"]
        delete_columns_by_name(ws, [
            "Remote Device Name", "Remote Device Port", "Patch Panel Matrix",
        ])

        # Move Rx Power to the first column on Optic Errors
        hdrs = [c.value for c in ws[1]]
        if "Rx Power" in hdrs and hdrs.index("Rx Power") != 0:
            rx_idx = hdrs.index("Rx Power")
            new_order = [rx_idx] + [i for i in range(len(hdrs)) if i != rx_idx]

            data = [[c.value for c in row] for row in ws.iter_rows()]
            styles = [[{
                "font": copy(c.font), "fill": copy(c.fill),
                "border": copy(c.border), "alignment": copy(c.alignment),
                "number_format": c.number_format,
            } for c in row] for row in ws.iter_rows()]
            old_widths = [ws.column_dimensions[get_column_letter(i + 1)].width
                          for i in range(len(hdrs))]

            ws.delete_rows(1, ws.max_row)
            for r_i, (row_vals, row_styles) in enumerate(zip(data, styles), start=1):
                for new_c, old_c in enumerate(new_order, start=1):
                    cell = ws.cell(row=r_i, column=new_c, value=row_vals[old_c])
                    st = row_styles[old_c]
                    cell.font = st["font"]; cell.fill = st["fill"]
                    cell.border = st["border"]; cell.alignment = st["alignment"]
                    cell.number_format = st["number_format"]
            for new_c, old_c in enumerate(new_order, start=1):
                w = old_widths[old_c]
                if w:
                    ws.column_dimensions[get_column_letter(new_c)].width = w

    # Interface Down Errors: keep only Source Device Name + Source Device Port
    if "Interface Down Errors" in wb.sheetnames:
        delete_columns_by_name(wb["Interface Down Errors"], [
            "Source Device Location", "Remote Device Name", "Remote Device Port",
            "Issue", "Patch Panel Matrix",
        ])


# ---------------------------------------------------------------------------
# Step 3 — Enrich a sheet by matching device+port against cutsheet
# ---------------------------------------------------------------------------
HOP_HEADERS = ["Device Rack U", "PP 1", "PP 2", "PP 3", "PP 4", "Peer Device", "Peer Port", "Peer Rack"]


def enrich(ws, name_col_idx, port_col_idx, insert_after_idx, lookup):
    """
    Adds 8 columns after `insert_after_idx` and fills them by matching
    'name port' against the cutsheet.

    Column layout: Device Rack U | PP 1 | PP 2 | PP 3 | PP 4 | Peer Device | Peer Port | Peer Rack
    Long-path rows: rack + up to 4 PPs + peer device name + peer port + peer rack
    Short-path rows: rack + (PP cols empty) + peer device name + peer port + peer rack
    """
    header_ref = ws.cell(row=1, column=1)
    data_ref = ws.cell(row=2, column=1) if ws.max_row >= 2 else header_ref

    new_data = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=name_col_idx).value
        port = ws.cell(row=r, column=port_col_idx).value
        key = f"{name} {port}" if name and port else None
        info = lookup.get(key) if key else None
        if info:
            peer_name, peer_port = split_device_port(info["peer_device"])
            if info["pp1"] is not None:
                new_data.append([
                    info["device_rack"],
                    info["pp1"], info["pp2"], info["pp3"], info["pp4"],
                    peer_name, peer_port, info["peer_rack"],
                ])
            else:
                new_data.append([
                    info["device_rack"],
                    None, None, None, None,
                    peer_name, peer_port, info["peer_rack"],
                ])
        else:
            new_data.append([None] * 8)

    for i in range(8):
        ws.insert_cols(insert_after_idx + 1 + i)
    for i, h in enumerate(HOP_HEADERS):
        c = ws.cell(row=1, column=insert_after_idx + 1 + i, value=h)
        style_cell(c, header_ref)
    for r, vals in enumerate(new_data, start=2):
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=insert_after_idx + 1 + i)
            c.value = v
            style_cell(c, data_ref)


def enrich_all(wb, lookup):
    if "Downlink" in wb.sheetnames:
        # Device A Name = col 1, Device A Port = col 2
        enrich(wb["Downlink"], 1, 2, 2, lookup)
    if "Optic Errors" in wb.sheetnames:
        # Rx Power = col 1, Source Device Name = col 2, Source Device Port = col 3
        enrich(wb["Optic Errors"], 2, 3, 3, lookup)
    if "Mismatch" in wb.sheetnames:
        # Device A Name = col 1, Device A Port = col 2
        enrich(wb["Mismatch"], 1, 2, 2, lookup)
    if "Interface Down Errors" in wb.sheetnames:
        # Source Device Name = col 1, Source Device Port = col 2
        enrich(wb["Interface Down Errors"], 1, 2, 2, lookup)


# ---------------------------------------------------------------------------
# Step 4 — Split each of the 3 tabs into T3-T2 (long) and T2-T1-T0 (short)
# ---------------------------------------------------------------------------
# Which (device-name, port) columns identify each sheet's row in the cutsheet.
# Header-name based, not positional, so column re-ordering elsewhere won't break it.
SPLIT_KEY_HEADERS = {
    "Downlink":              ("Device A Name",      "Device A Port"),
    "Mismatch":              ("Device A Name",      "Device A Port"),
    "Optic Errors":          ("Source Device Name", "Source Device Port"),
    "Interface Down Errors": ("Source Device Name", "Source Device Port"),
}


def split_long_short(wb, lookup):
    """
    Route each row to its long-path (T3-T2) or short-path (T2-T1-T0) tab.

    Classification asks the cutsheet directly: a cable is long-path iff its
    lookup entry has a non-empty pp1. We never inspect the hop strings —
    that decouples the split from how patch panels are named (PP:, PP., pp-,
    whatever the next cutsheet convention happens to be).
    """
    for src_name in ["Downlink", "Mismatch", "Optic Errors", "Interface Down Errors"]:
        if src_name not in wb.sheetnames:
            continue
        src = wb[src_name]
        headers = [c.value for c in src[1]]
        name_hdr, port_hdr = SPLIT_KEY_HEADERS[src_name]
        name_col = headers.index(name_hdr) + 1
        port_col = headers.index(port_hdr) + 1
        header_styles = [{
            "font": copy(c.font), "fill": copy(c.fill),
            "border": copy(c.border), "alignment": copy(c.alignment),
            "number_format": c.number_format,
        } for c in src[1]]
        data_ref = src.cell(row=2, column=1) if src.max_row >= 2 else src.cell(row=1, column=1)
        col_widths = {k: v.width for k, v in src.column_dimensions.items()}

        long_rows, short_rows = [], []
        for r in range(2, src.max_row + 1):
            row_vals = [src.cell(row=r, column=c).value for c in range(1, src.max_column + 1)]
            name = row_vals[name_col - 1]
            port = row_vals[port_col - 1]
            key = f"{name} {port}" if name and port else None
            info = lookup.get(key) if key else None
            # Long-path iff the cutsheet says so. Unmatched rows fall through
            # to short — same as the prior behaviour.
            if info and info.get("pp1") is not None:
                long_rows.append(row_vals)
            else:
                short_rows.append(row_vals)

        def build(name, rows):
            if name in wb.sheetnames:
                del wb[name]
            ws = wb.create_sheet(name)
            ws.append(headers)
            for i, c in enumerate(ws[1], start=1):
                s = header_styles[i - 1]
                c.font = s["font"]; c.fill = s["fill"]; c.border = s["border"]
                c.alignment = s["alignment"]; c.number_format = s["number_format"]
            for row_vals in rows:
                ws.append(row_vals)
                for c in ws[ws.max_row]:
                    style_cell(c, data_ref)
            for k, w in col_widths.items():
                if w:
                    ws.column_dimensions[k].width = w
            ws.freeze_panes = "A2"

        build(f"T3-T2 {src_name}", long_rows)
        build(f"T2-T1-T0 {src_name}", short_rows)
        del wb[src_name]


# ---------------------------------------------------------------------------
# Step 5 — Drop unused Hops from T2-T1-T0 tabs
# ---------------------------------------------------------------------------
def trim_short_tabs(wb):
    for sn in ["T2-T1-T0 Downlink", "T2-T1-T0 Mismatch", "T2-T1-T0 Optic Errors",
               "T2-T1-T0 Interface Down Errors"]:
        if sn in wb.sheetnames:
            delete_columns_by_name(wb[sn], ["PP 1", "PP 2", "PP 3", "PP 4"])


# ---------------------------------------------------------------------------
# Step 6 — Enrich Device B side on both Mismatch tabs, rename + pink
# ---------------------------------------------------------------------------
def enrich_mismatch_b_side(wb, lookup, device_rack_lookup):
    for sn, b_headers, expect_long in [
        ("T3-T2 Mismatch",
         ["Act. Rack U",
          "Cut. PP 1", "Cut. PP 2", "Cut. PP 3", "Cut. PP 4",
          "Cut. Other End", "Cut. Other End Port", "Cut. Other End Rack"],
         True),
        ("T2-T1-T0 Mismatch",
         ["Act. Rack U",
          "Cut. Other End", "Cut. Other End Port", "Cut. Other End Rack"],
         False),
    ]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        header_ref = ws.cell(row=1, column=1)
        data_ref = ws.cell(row=2, column=1) if ws.max_row >= 2 else header_ref

        # Rename A-side columns: rack/PPs get "A " prefix; peer becomes "Exp."
        rename_map = {
            "Device Rack U": "A Rack U",
            "PP 1": "A PP 1", "PP 2": "A PP 2",
            "PP 3": "A PP 3", "PP 4": "A PP 4",
            "Peer Device": "Exp. Device",
            "Peer Port":   "Exp. Port",
            "Peer Rack":   "Exp. Rack",
        }
        for c in ws[1]:
            if c.value in rename_map:
                c.value = rename_map[c.value]

        # Append B-side header columns (Act./Cut. names)
        start_col = ws.max_column + 1
        for i, h in enumerate(b_headers):
            c = ws.cell(row=1, column=start_col + i, value=h)
            style_cell(c, header_ref)

        hdrs = [c.value for c in ws[1]]
        bname_col = hdrs.index("Device B Name") + 1
        bport_col = hdrs.index("Device B Port") + 1
        slots = len(b_headers)

        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=bname_col).value
            port = ws.cell(row=r, column=bport_col).value
            key = f"{name} {port}" if name and port else None
            info = lookup.get(key) if key else None

            if info:
                peer_name, peer_port = split_device_port(info["peer_device"])
                if expect_long:
                    vals = [
                        info["device_rack"],
                        info["pp1"], info["pp2"], info["pp3"], info["pp4"],
                        peer_name, peer_port, info["peer_rack"],
                    ]
                else:
                    vals = [info["device_rack"], peer_name, peer_port, info["peer_rack"]]
            else:
                # Fallback: at least try to populate rack from device name
                fallback_rack = device_rack_lookup.get(str(name).strip()) if name else None
                if expect_long:
                    vals = [fallback_rack, None, None, None, None, None, None, None]
                else:
                    vals = [fallback_rack, None, None, None]

            for i, v in enumerate(vals):
                c = ws.cell(row=r, column=start_col + i)
                c.value = v
                style_cell(c, data_ref)

        # Rename "Device B Name/Port" → "Act. Device/Port" then pink the whole
        # active + cutsheet block
        hdrs = [c.value for c in ws[1]]
        start_pink = hdrs.index("Device B Name") + 1
        rename_b = {"Device B Name": "Act. Device", "Device B Port": "Act. Port"}
        for col in range(start_pink, ws.max_column + 1):
            cur_h = ws.cell(row=1, column=col).value
            if cur_h in rename_b:
                ws.cell(row=1, column=col).value = rename_b[cur_h]
            for r in range(1, ws.max_row + 1):
                ws.cell(row=r, column=col).fill = PINK


# ---------------------------------------------------------------------------
# Step 7 — Grey-out Optic rows that also appear in Downlink
# ---------------------------------------------------------------------------
def grey_optics_matching_downlink(wb):
    keys = set()
    for sn in ["T3-T2 Downlink", "T2-T1-T0 Downlink"]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        hdrs = header_indices(ws)
        n_col, p_col, r_col = hdrs.get("Device A Name"), hdrs.get("Device A Port"), hdrs.get("Device Rack U")
        if not (n_col and p_col):
            continue
        for r in range(2, ws.max_row + 1):
            n = ws.cell(row=r, column=n_col).value
            p = ws.cell(row=r, column=p_col).value
            ru = ws.cell(row=r, column=r_col).value if r_col else None
            if n and p:
                keys.add((str(n).strip(), str(p).strip(),
                          str(ru).strip() if ru else None))

    for sn in ["T3-T2 Optic Errors", "T2-T1-T0 Optic Errors"]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        hdrs = header_indices(ws)
        n_col = hdrs.get("Source Device Name")
        p_col = hdrs.get("Source Device Port")
        r_col = hdrs.get("Device Rack U")
        if not (n_col and p_col):
            continue
        for r in range(2, ws.max_row + 1):
            n = ws.cell(row=r, column=n_col).value
            p = ws.cell(row=r, column=p_col).value
            ru = ws.cell(row=r, column=r_col).value if r_col else None
            key = (str(n).strip() if n else None,
                   str(p).strip() if p else None,
                   str(ru).strip() if ru else None)
            if n and p and key in keys:
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    f = copy(cell.font)
                    f.color = LIGHT_GREY
                    cell.font = f


# ---------------------------------------------------------------------------
# Step 8b — Rename generic "Peer" columns to context-aware names
# ---------------------------------------------------------------------------
def finalise_column_names(wb):
    """
    Rename the generic 'Peer Device/Port/Rack' columns produced by enrich()
    to QFAB-style names that make the semantic meaning explicit:

      Downlink / Interface Down tabs  → Exp. Device / Exp. Port / Exp. Rack
      Optic Error tabs                → Cut. Other End / Cut. Other End Port /
                                        Cut. Other End Rack
      Mismatch tabs                   → already renamed in enrich_mismatch_b_side
      Summary                         → unchanged
    """
    DOWNLINK_MAP = {
        "Peer Device": "Exp. Device",
        "Peer Port":   "Exp. Port",
        "Peer Rack":   "Exp. Rack",
    }
    OPTIC_MAP = {
        "Peer Device": "Cut. Other End",
        "Peer Port":   "Cut. Other End Port",
        "Peer Rack":   "Cut. Other End Rack",
    }
    for sn in wb.sheetnames:
        if sn == "Summary" or "Mismatch" in sn:
            continue
        rename = OPTIC_MAP if "Optic" in sn else DOWNLINK_MAP
        for c in wb[sn][1]:
            if c.value in rename:
                c.value = rename[c.value]


# ---------------------------------------------------------------------------
# Step 8c — Fill empty PP slots with "<=>" (QFAB convention)
# ---------------------------------------------------------------------------
def fill_empty_pps(wb):
    """
    Replace None / empty values in patch-panel columns with '<=>' so readers
    can instantly see a PP slot was expected but absent.  Applies to any column
    whose header contains 'PP' (e.g. 'PP 1', 'A PP 2', 'Cut. PP 3').
    """
    for sn in wb.sheetnames:
        if sn == "Summary":
            continue
        ws = wb[sn]
        pp_cols = [
            i + 1 for i, c in enumerate(ws[1])
            if c.value and "PP" in str(c.value)
        ]
        for r in range(2, ws.max_row + 1):
            for col in pp_cols:
                cell = ws.cell(row=r, column=col)
                if cell.value is None or cell.value == "":
                    cell.value = "<=>"


# ---------------------------------------------------------------------------
# Step 8 — Note column on every non-Summary tab
# ---------------------------------------------------------------------------
def add_note_column(wb):
    for sn in wb.sheetnames:
        if sn == "Summary":
            continue
        ws = wb[sn]
        ref = ws.cell(row=1, column=1)
        col = ws.max_column + 1
        c = ws.cell(row=1, column=col, value="Note")
        f = copy(ref.font); f.bold = True; c.font = f
        c.alignment = copy(ref.alignment)
        c.border = copy(ref.border)
        c.fill = YELLOW
        ws.column_dimensions[get_column_letter(col)].width = 30


# ---------------------------------------------------------------------------
# Step 9 — Header styling, borders, autofilter, freezes
# ---------------------------------------------------------------------------
def finalize_styling(wb):
    for sn in wb.sheetnames:
        ws = wb[sn]
        # bold yellow headers (font size left at default)
        for c in ws[1]:
            f = copy(c.font); f.bold = True; f.size = 11; c.font = f
            c.fill = YELLOW
        # freezes
        if sn in ("T3-T2 Optic Errors", "T2-T1-T0 Optic Errors"):
            ws.freeze_panes = "B2"   # also freeze Rx Power column
        else:
            ws.freeze_panes = "A2"
        # all-around borders on every used cell
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).border = BORDER
        # autofilter on data tabs
        if sn != "Summary":
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


# ---------------------------------------------------------------------------
# Step 10 — Rebuild Summary
# ---------------------------------------------------------------------------
def rebuild_summary(wb):
    if "Summary" not in wb.sheetnames:
        s = wb.create_sheet("Summary", 0)
        s.cell(row=1, column=1, value="Error Category")
        s.cell(row=1, column=2, value="Error Count")
    s = wb["Summary"]

    # clear data rows
    for r in range(s.max_row, 1, -1):
        s.delete_rows(r)

    rows = [
        ("T3-T2 Downlink",                 "=COUNTA('T3-T2 Downlink'!A:A)-1"),
        ("T2-T1-T0 Downlink",              "=COUNTA('T2-T1-T0 Downlink'!A:A)-1"),
        ("T3-T2 Mismatch",                 "=COUNTA('T3-T2 Mismatch'!A:A)-1"),
        ("T2-T1-T0 Mismatch",              "=COUNTA('T2-T1-T0 Mismatch'!A:A)-1"),
        ("T3-T2 Optic Errors",             "=COUNTA('T3-T2 Optic Errors'!A:A)-1"),
        ("T2-T1-T0 Optic Errors",          "=COUNTA('T2-T1-T0 Optic Errors'!A:A)-1"),
        ("T3-T2 Interface Down Errors",    "=COUNTA('T3-T2 Interface Down Errors'!A:A)-1"),
        ("T2-T1-T0 Interface Down Errors", "=COUNTA('T2-T1-T0 Interface Down Errors'!A:A)-1"),
    ]
    # Only include rows for tabs that actually exist
    rows = [(label, formula) for (label, formula) in rows
            if label in wb.sheetnames]

    data_ref_cell = s.cell(row=1, column=1)  # any cell as starting style ref
    no_fill = PatternFill(fill_type=None)

    for i, (label, formula) in enumerate(rows, start=2):
        c1 = s.cell(row=i, column=1, value=label)
        c2 = s.cell(row=i, column=2, value=formula)
        style_cell(c1, data_ref_cell)
        style_cell(c2, data_ref_cell)
        # de-bold the data rows AND strip any inherited fill
        for c in (c1, c2):
            f = copy(c.font); f.bold = False; c.font = f
            c.fill = no_fill

    total_row = len(rows) + 2
    c1 = s.cell(row=total_row, column=1, value="Total")
    c2 = s.cell(row=total_row, column=2, value=f"=SUM(B2:B{len(rows) + 1})")
    style_cell(c1, data_ref_cell); style_cell(c2, data_ref_cell)
    for c in (c1, c2):
        f = copy(c.font); f.bold = True; c.font = f
        c.fill = no_fill

    s.column_dimensions["A"].width = 24
    s.column_dimensions["B"].width = 12


# ---------------------------------------------------------------------------
# Step 11 — Final tab order
# ---------------------------------------------------------------------------
def reorder_tabs(wb):
    desired = [
        "Summary",
        "T3-T2 Downlink", "T2-T1-T0 Downlink",
        "T3-T2 Mismatch", "T2-T1-T0 Mismatch",
        "T3-T2 Optic Errors", "T2-T1-T0 Optic Errors",
        "T3-T2 Interface Down Errors", "T2-T1-T0 Interface Down Errors",
    ]
    wb._sheets = [wb[n] for n in desired if n in wb.sheetnames] + \
                 [wb[n] for n in wb.sheetnames if n not in desired]


# ---------------------------------------------------------------------------
# Per-file processing
# ---------------------------------------------------------------------------
def process_file(input_path: Path, lookup: dict, device_rack_lookup: dict, output_dir: Path) -> Path:
    wb = openpyxl.load_workbook(input_path)

    # Derive rack number for output filename (before any sheet manipulation)
    rack = "output"
    if "LLDP Mismatch + Link Down" in wb.sheetnames:
        rack = rack_number_from(wb["LLDP Mismatch + Link Down"])

    split_lldp_sheet(wb)            # Step 1
    clean_columns(wb)               # Step 2
    enrich_all(wb, lookup)          # Step 3
    split_long_short(wb, lookup)    # Step 4
    trim_short_tabs(wb)             # Step 5
    enrich_mismatch_b_side(wb, lookup, device_rack_lookup)  # Step 6
    grey_optics_matching_downlink(wb)   # Step 7
    finalise_column_names(wb)           # Step 8b — Exp./Act./Cut. naming
    fill_empty_pps(wb)                  # Step 8c — empty PPs → <=>
    add_note_column(wb)                 # Step 8
    finalize_styling(wb)            # Step 9
    rebuild_summary(wb)             # Step 10
    reorder_tabs(wb)                # Step 11

    # Auto-fit each tab one last time so the new columns get readable widths
    for sn in wb.sheetnames:
        if sn != "Summary":
            autofit_columns(wb[sn])

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"{rack}.xlsx"
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# CLI (uses native file-picker dialogs via tkinter)
# ---------------------------------------------------------------------------
def main():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()             # hide the empty root window
    root.attributes("-topmost", True)

    print("Rack validation workbook formatter")
    print("-" * 40)

    # 1) Pick the CFAB cutsheet
    print("Select the CFAB cutsheet ...")
    cutsheet_str = filedialog.askopenfilename(
        title="Select CFAB cutsheet",
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
    )
    if not cutsheet_str:
        messagebox.showwarning("Cancelled", "No cutsheet selected. Exiting.")
        sys.exit(1)
    cutsheet = Path(cutsheet_str)
    print(f"  Cutsheet: {cutsheet}")
    lookup, device_rack_lookup = build_cutsheet_lookup(cutsheet)
    print(f"  {len(lookup)} cutsheet lookup keys built.\n")

    # 2) Pick one or more input files (multi-select)
    print("Select the rack validation file(s) to format ...")
    files_tuple = filedialog.askopenfilenames(
        title="Select rack validation file(s) (you can pick multiple)",
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
    )
    files = [Path(p) for p in files_tuple]
    if not files:
        messagebox.showwarning("Cancelled", "No input files selected. Exiting.")
        sys.exit(1)
    print(f"  {len(files)} file(s) selected.\n")

    # 3) Process — output goes into the same folder as each input file
    results, errors = [], []
    for f in files:
        try:
            out = process_file(f, lookup, device_rack_lookup, f.parent)
            print(f"  OK  {f.name}  ->  {out}")
            results.append(out)
        except Exception as exc:
            print(f"  ERR {f.name}: {exc}")
            errors.append(f"{f.name}: {exc}")

    print("\nDone.")
    summary = f"Processed {len(results)} of {len(files)} file(s).\n\n"
    summary += "\n".join(f"OK  {p}" for p in results)
    if errors:
        summary += "\n\nErrors:\n" + "\n".join(errors)
    messagebox.showinfo("Finished", summary)


if __name__ == "__main__":
    main()
