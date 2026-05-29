#!/usr/bin/env python3
"""
Excel Formatter Tool — Mac
Steps performed on each target file:
  1.  Split lldp_sp → Downlinks (interface down) + Mismatches (swp)
  2.  Rename optics_rx_tx_threshold → Optics, remove unwanted columns,
      move Measured (dBm) to first column
  3.  Remove interfaces_sp tab
  4.  Reorder tabs: Downlinks, Mismatches, Optics, combined_fec
  5.  Insert L/R column after every Interface column (Z L/R after Z Interface)
  6.  Populate Source_port, DMARC1, DMARC2, Destination_port from Cutsheets
  6b. Mismatches: add Possible columns (matched via cutsheet Z-side) +
      Active Z columns (from Act. data), both with pink background
  7.  Add Summary tab (first) with row counts
  8.  Remove all fills (except pink Possible/Z cols); apply borders everywhere
  9.  Rename output file to top-2 Rack numbers
"""

import sys
import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from collections import Counter

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── File-picker helpers ──────────────────────────────────────────────────────

def _root():
    r = tk.Tk()
    r.withdraw()
    r.attributes("-topmost", True)
    return r


def pick_file(title):
    r = _root()
    path = filedialog.askopenfilename(
        title=title, filetypes=[("Excel files", "*.xlsx *.xls")])
    r.destroy()
    return path or None


def pick_files(title):
    r = _root()
    paths = filedialog.askopenfilenames(
        title=title, filetypes=[("Excel files", "*.xlsx *.xls")])
    r.destroy()
    return list(paths) if paths else []


def identify_refs(paths):
    """
    Given two reference file paths, identify which is the Formula L&R and
    which is the Cutsheet. The cutsheet is the one that contains an
    'Installation Sheet' tab; the other is the L&R.
    Returns (formula_path, cutsheet_path).
    """
    formula = cutsheet = None
    for p in paths:
        try:
            sheets = load_workbook(p, read_only=True).sheetnames
        except Exception as e:
            raise ValueError(f"Could not open {os.path.basename(p)}: {e}")
        if 'Installation Sheet' in sheets:
            if cutsheet:
                raise ValueError(
                    "Both selected files contain an 'Installation Sheet' tab. "
                    "Could not tell which one is the cutsheet."
                )
            cutsheet = p
        else:
            if formula:
                raise ValueError(
                    "Neither selected file contains an 'Installation Sheet' tab. "
                    "Could not identify the cutsheet."
                )
            formula = p
    if not cutsheet:
        raise ValueError(
            "Neither selected file contains an 'Installation Sheet' tab; "
            "expected one of them to be the cutsheet."
        )
    if not formula:
        raise ValueError("Could not identify the Formula L&R file.")
    return formula, cutsheet


# ── Style helpers ────────────────────────────────────────────────────────────

PINK   = 'FFB6C1'
YELLOW = 'FFFF00'

def thin_border():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)


def clear_and_border(ws, pink_cols=None):
    """Remove all fills (preserve pink cols); yellow-highlight row 1; apply black border everywhere."""
    bd        = thin_border()
    no_fill   = PatternFill(fill_type=None)
    pink_fill = PatternFill('solid', start_color=PINK)
    yellow_fill = PatternFill('solid', start_color=YELLOW)
    pink_cols = set(pink_cols or [])
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                             min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row == 1:
                cell.fill = yellow_fill
            elif cell.column in pink_cols:
                cell.fill = pink_fill
            else:
                cell.fill = no_fill
            cell.border = bd
            if cell.font:
                cell.font = Font(
                    bold=cell.font.bold,
                    name=cell.font.name or 'Arial',
                    size=cell.font.size or 10,
                    color='FF000000'
                )


def header_cell(cell, value, fill=None):
    cell.value     = value
    cell.font      = Font(bold=True, name='Arial', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    cell.border    = thin_border()
    cell.fill      = fill if fill else PatternFill('solid', start_color=YELLOW)


def data_cell(cell, value, fill=None):
    cell.value  = value
    cell.border = thin_border()
    if fill:
        cell.fill = fill


def autofit_sheet(ws, header_row_height=24, data_row_height=20, max_col_width=80):
    """Expand columns to fit content and give rows a comfortable height.

    - Column width = longest cell content in that column (+ padding), capped
      at `max_col_width` so a single huge value can't blow up the layout.
    - Cells inside a merged range are ignored when measuring column width.
    - Row 1 gets `header_row_height`; remaining rows get `data_row_height`.
    """
    # Cells that are part of a merge — skip them when measuring widths
    merged = set()
    for mrange in ws.merged_cells.ranges:
        for r in range(mrange.min_row, mrange.max_row + 1):
            for c in range(mrange.min_col, mrange.max_col + 1):
                merged.add((r, c))

    col_max = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                             min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue
            if (cell.row, cell.column) in merged:
                continue
            # Handle multi-line values: width = longest line
            longest_line = max(
                (len(line) for line in str(cell.value).splitlines()),
                default=0
            )
            letter = get_column_letter(cell.column)
            if longest_line > col_max.get(letter, 0):
                col_max[letter] = longest_line

    for letter, length in col_max.items():
        ws.column_dimensions[letter].width = min(length + 4, max_col_width)

    # Row heights
    if ws.max_row >= 1:
        ws.row_dimensions[1].height = header_row_height
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = data_row_height


def write_sheet(wb, name, df):
    ws = wb.create_sheet(name)
    bd = thin_border()
    for c, col in enumerate(df.columns, 1):
        header_cell(ws.cell(row=1, column=c), col)
    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, col in enumerate(df.columns, 1):
            val = row[col]
            cell = ws.cell(row=r, column=c, value=None if pd.isna(val) else val)
            cell.border = bd
    for c, col in enumerate(df.columns, 1):
        mx = max([len(str(col))] + [len(str(v)) for v in df[col].dropna()])
        ws.column_dimensions[get_column_letter(c)].width = min(mx + 2, 40)
    ws.freeze_panes = 'A2'
    return ws


# ── Reference-file loaders ───────────────────────────────────────────────────

def load_lr_lookup(path):
    df = pd.read_excel(path, header=0)
    df.columns = ['key', 'value']
    return dict(zip(df['key'].astype(str).str.strip(),
                    df['value'].astype(str).str.strip()))


def load_cutsheet(path):
    return pd.read_excel(path, sheet_name='Installation Sheet')


def build_cutsheet_lookup(cut_df):
    """Key: (L/R, Rack, Elevation) → row dict for Source_port etc."""
    fill_cols = ['Source_port', 'DMARC1', 'DMARC2', 'Destination_port']
    lookup = {}
    for _, row in cut_df.iterrows():
        key = (
            str(row['L/R']).strip(),
            str(row['Rack']).strip().split('.')[0],
            str(row['Elevation']).strip().split('.')[0],
        )
        lookup[key] = {c: row[c] for c in fill_cols}
    return lookup


def build_z_lookup(cut_df):
    """Key: (Z Hostname, Z Interface, Z Rack, Z Elevation) → full row."""
    lookup = {}
    for _, row in cut_df.iterrows():
        key = (
            str(row['Z Hostname']).strip(),
            str(row['Z Interface']).strip(),
            str(int(float(str(row['Z Rack'])))),
            str(int(float(str(row['Z Elevation'])))),
        )
        lookup[key] = row
    return lookup


def paired_subport(iface):
    """Return the paired sub-port interface name.

    s0 ↔ s1 are a pair, s2 ↔ s3 are a pair.
    e.g. 'swp4s0' → 'swp4s1', 'swp15s3' → 'swp15s2'.
    Returns None if iface doesn't end in s0/s1/s2/s3.
    """
    pairs = {'s0': 's1', 's1': 's0', 's2': 's3', 's3': 's2'}
    for suffix, mate in pairs.items():
        if iface.endswith(suffix):
            return iface[:-len(suffix)] + mate
    return None


# ── Core processor ───────────────────────────────────────────────────────────

def process_file(input_path, output_path, lr_lookup, cut_df, log):
    shutil.copy2(input_path, output_path)
    wb = load_workbook(output_path)

    cutsheet_lookup = build_cutsheet_lookup(cut_df)
    z_lookup        = build_z_lookup(cut_df)

    # ── 1. Split lldp_sp ────────────────────────────────────────────────────
    mis_orig_df = None   # keep for step 6b
    if 'lldp_sp' in wb.sheetnames:
        log("  · Splitting lldp_sp → Downlinks / Mismatches")
        df = pd.read_excel(input_path, sheet_name='lldp_sp')
        down_df     = df[df['Act. Interface'] == 'interface down'].copy()
        mis_orig_df = df[df['Act. Interface'].str.startswith('swp', na=False)].copy()
        drop = ['Active Host', 'Act. Interface', 'Act. Rack', 'Act. Elevation',
                'Expected Hostname', 'Exp. Interface', 'Exp. Rack', 'Exp. Elevation']
        down_df.drop(columns=[c for c in drop if c in down_df.columns], inplace=True)
        del wb['lldp_sp']
        write_sheet(wb, 'Downlinks', down_df)
        write_sheet(wb, 'Mismatches', mis_orig_df.drop(
            columns=[c for c in [] if c in mis_orig_df.columns]))  # keep all cols for now

    # ── 2. Optics tab ───────────────────────────────────────────────────────
    optics_src = 'optics_rx_tx_threshold'
    if optics_src in wb.sheetnames:
        log("  · Processing Optics tab")
        drop_cols = {'Transceiver', 'Channel',
                     'Min Threshold (dBm)', 'Max Threshold (dBm)'}
        optics_df = pd.read_excel(input_path, sheet_name=optics_src)
        optics_df.drop(columns=[c for c in drop_cols if c in optics_df.columns], inplace=True)
        if 'Measured (dBm)' in optics_df.columns:
            cols = optics_df.columns.tolist()
            cols.insert(0, cols.pop(cols.index('Measured (dBm)')))
            optics_df = optics_df[cols]
        del wb[optics_src]
        write_sheet(wb, 'Optics', optics_df)
        # Freeze row 1 + column A so "Measured (dBm)" stays visible while scrolling
        wb['Optics'].freeze_panes = 'B2'

    # ── 3. Remove interfaces_sp ─────────────────────────────────────────────
    if 'interfaces_sp' in wb.sheetnames:
        log("  · Removing interfaces_sp")
        del wb['interfaces_sp']

    # ── 3b. combined_fec: move Lock Status + Pre-FEC BER to the front ───────
    if 'combined_fec' in wb.sheetnames:
        log("  · Reordering combined_fec (Lock Status, Pre-FEC BER first)")
        fec_df = pd.read_excel(input_path, sheet_name='combined_fec')

        def _norm(s):
            # Normalize hyphen variants and whitespace for tolerant matching
            return (str(s)
                    .replace('\u2011', '-')   # non-breaking hyphen
                    .replace('\u2013', '-')   # en dash
                    .replace('\u2014', '-')   # em dash
                    .strip()
                    .lower())

        wanted = ['lock status', 'pre-fec ber']
        front = []
        for target in wanted:
            for col in fec_df.columns:
                if _norm(col) == target and col not in front:
                    front.append(col)
                    break

        if front:
            rest = [c for c in fec_df.columns if c not in front]
            fec_df = fec_df[front + rest]
            del wb['combined_fec']
            write_sheet(wb, 'combined_fec', fec_df)
        else:
            log("    ⚠ Lock Status / Pre-FEC BER not found — leaving combined_fec as-is")

    # ── 4. Reorder tabs ─────────────────────────────────────────────────────
    desired  = ['Downlinks', 'Mismatches', 'Optics', 'combined_fec']
    existing = [s for s in desired if s in wb.sheetnames]
    others   = [s for s in wb.sheetnames if s not in desired]
    for i, name in enumerate(existing + others):
        wb.move_sheet(name, offset=wb.sheetnames.index(name) - i)

    # ── 5. Insert L/R columns ───────────────────────────────────────────────
    log("  · Adding L/R mapped columns")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        targets = [(i+1, h) for i, h in enumerate(header)
                   if h in ('Interface', 'Z Interface')]
        for col_idx, col_name in sorted(targets, reverse=True):
            new_name = 'L/R' if col_name == 'Interface' else 'Z L/R'
            ws.insert_cols(col_idx + 1)
            header_cell(ws.cell(row=1, column=col_idx + 1), new_name)
            for r in range(2, ws.max_row + 1):
                val = str(ws.cell(row=r, column=col_idx).value or '').strip()
                ws.cell(row=r, column=col_idx + 1, value=lr_lookup.get(val, ''))
                ws.cell(row=r, column=col_idx + 1).border = thin_border()
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = 10

    # ── 6. Populate Source_port / DMARC1 / DMARC2 / Destination_port ────────
    log("  · Filling Source_port / DMARC1 / DMARC2 / Destination_port")
    fill_cols = ['Source_port', 'DMARC1', 'DMARC2', 'Destination_port']
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if not all(c in header for c in ['L/R', 'Rack', 'Elevation'] + fill_cols):
            continue
        lr_c, rack_c, elev_c = header.index('L/R')+1, header.index('Rack')+1, header.index('Elevation')+1
        fill_idx = {c: header.index(c)+1 for c in fill_cols}
        for r in range(2, ws.max_row + 1):
            lr   = str(ws.cell(row=r, column=lr_c).value or '').strip()
            rack = str(ws.cell(row=r, column=rack_c).value or '').strip().split('.')[0]
            elev = str(ws.cell(row=r, column=elev_c).value or '').strip().split('.')[0]
            match = cutsheet_lookup.get((lr, rack, elev))
            if match:
                for col_name, col_idx in fill_idx.items():
                    ws.cell(row=r, column=col_idx, value=match[col_name])

    # ── 6b. Mismatches: Possible columns + Active Z columns (pink) ───────────
    if 'Mismatches' in wb.sheetnames:
        log("  · Building Possible + Active Z columns in Mismatches")
        pink_fill   = PatternFill('solid', start_color=PINK)
        yellow_fill = PatternFill('solid', start_color=YELLOW)
        bd          = thin_border()

        # Build act_lookup from original lldp_sp
        act_lookup = {}
        src_sheets = pd.ExcelFile(input_path).sheet_names
        if 'lldp_sp' in src_sheets:
            orig_df  = pd.read_excel(input_path, sheet_name='lldp_sp')
            mis_rows = orig_df[orig_df['Act. Interface'].str.startswith('swp', na=False)]
            for _, row in mis_rows.iterrows():
                key       = (str(row['Hostname']).strip(), str(row['Interface']).strip())
                act_iface = str(row['Act. Interface']).strip()
                act_lookup[key] = {
                    'Z Hostname' : str(row['Active Host']).strip(),
                    'Z Interface': act_iface,
                    'Z L/R'      : lr_lookup.get(act_iface, ''),
                    'Z Rack'     : int(float(str(row['Act. Rack']))),
                    'Z Elevation': int(float(str(row['Act. Elevation']))),
                }

        ws_m   = wb['Mismatches']
        header = [ws_m.cell(row=1, column=c).value for c in range(1, ws_m.max_column + 1)]

        # Remove leftover Act/Expected columns
        act_drop = {'Active Host', 'Act. Interface', 'Act. Rack', 'Act. Elevation',
                    'Expected Hostname', 'Exp. Interface', 'Exp. Rack', 'Exp. Elevation'}
        for idx in sorted([i+1 for i, h in enumerate(header) if h in act_drop], reverse=True):
            ws_m.delete_cols(idx)
        header = [ws_m.cell(row=1, column=c).value for c in range(1, ws_m.max_column + 1)]

        h_idx = header.index('Hostname') + 1
        i_idx = header.index('Interface') + 1

        # Collect act data per row
        act_rows = []
        for r in range(2, ws_m.max_row + 1):
            hn    = str(ws_m.cell(row=r, column=h_idx).value or '').strip()
            iface = str(ws_m.cell(row=r, column=i_idx).value or '').strip()
            act_rows.append(act_lookup.get((hn, iface), {}))

        # Possible columns: match act Z key against cutsheet Z side
        possible_cols = [
            ('Possible Hostname',         'Hostname'),
            ('Possible Interface',        'Interface'),
            ('Possible L/R',              '__lr__'),
            ('Possible Rack',             'Rack'),
            ('Possible Elevation',        'Elevation'),
            ('Possible Source_port',      'Source_port'),
            ('Possible DMARC1',           'DMARC1'),
            ('Possible DMARC2',           'DMARC2'),
            ('Possible Destination_port', 'Destination_port'),
        ]

        possible_data = {col: [] for col, _ in possible_cols}
        for act in act_rows:
            zh   = act.get('Z Hostname', '')
            zi   = act.get('Z Interface', '')
            zr   = str(act.get('Z Rack', '')).split('.')[0]
            ze   = str(act.get('Z Elevation', '')).split('.')[0]
            match = z_lookup.get((zh, zi, zr, ze)) if zh else None
            # Fallback: if exact sub-port not in cutsheet, try its pair.
            # s0↔s1 are a pair, s2↔s3 are a pair.
            if match is None and zh and zi:
                mate = paired_subport(zi)
                if mate:
                    match = z_lookup.get((zh, mate, zr, ze))
            for col, src in possible_cols:
                if match is not None:
                    val = lr_lookup.get(str(match.get('Interface', '')).strip(), '') \
                          if src == '__lr__' else match.get(src, '')
                else:
                    val = ''
                possible_data[col].append(val)

        # Write Possible columns
        pink_col_indices = []
        start = ws_m.max_column + 1
        for c_off, (col_name, _) in enumerate(possible_cols):
            col_idx = start + c_off
            pink_col_indices.append(col_idx)
            hdr = ws_m.cell(row=1, column=col_idx, value=col_name)
            hdr.font = Font(bold=True, name='Arial', size=10)
            hdr.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            hdr.fill  = yellow_fill
            hdr.border = bd
            ws_m.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name)+3, 14)
            for r_off, val in enumerate(possible_data[col_name]):
                cell = ws_m.cell(row=r_off+2, column=col_idx,
                                 value=val if val != '' else None)
                cell.fill   = pink_fill
                cell.border = bd

        # Write Active Z columns
        act_z_cols = ['Z Hostname', 'Z Interface', 'Z L/R', 'Z Rack', 'Z Elevation']
        start2 = ws_m.max_column + 1
        for c_off, col_name in enumerate(act_z_cols):
            col_idx = start2 + c_off
            pink_col_indices.append(col_idx)
            hdr = ws_m.cell(row=1, column=col_idx, value=col_name)
            hdr.font = Font(bold=True, name='Arial', size=10)
            hdr.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            hdr.fill  = yellow_fill
            hdr.border = bd
            ws_m.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name)+3, 14)
            for r_off, act in enumerate(act_rows):
                val = act.get(col_name, '')
                cell = ws_m.cell(row=r_off+2, column=col_idx,
                                 value=val if val != '' else None)
                cell.fill   = pink_fill
                cell.border = bd

    # ── 7. Summary tab (per Rack breakdown) ─────────────────────────────────
    log("  · Creating Summary tab")

    # Gather rack counts per sheet from the workbook data
    tab_rack  = {}
    all_racks = set()
    no_fill_s   = PatternFill(fill_type=None)
    yellow_fill_s = PatternFill('solid', start_color=YELLOW)
    center_s  = Alignment(horizontal='center', vertical='center', wrap_text=False)
    bd_s      = thin_border()

    def _s(cell, value, bold=False, header=False):
        cell.value     = value
        cell.font      = Font(bold=bold, name='Arial', size=10)
        cell.alignment = center_s
        cell.border    = bd_s
        cell.fill      = yellow_fill_s if header else no_fill_s

    for sname in wb.sheetnames:
        ws_tmp = wb[sname]
        hdr = [ws_tmp.cell(row=1, column=c).value for c in range(1, ws_tmp.max_column+1)]
        if 'Rack' not in hdr:
            tab_rack[sname] = {}
            continue
        rack_col = hdr.index('Rack') + 1
        counts = {}
        for r in range(2, ws_tmp.max_row + 1):
            val = ws_tmp.cell(row=r, column=rack_col).value
            if val is not None:
                try:
                    k = int(float(str(val)))
                    counts[k] = counts.get(k, 0) + 1
                    all_racks.add(k)
                except ValueError:
                    pass
        tab_rack[sname] = counts

    racks      = sorted(all_racks)
    tabs_order = [n for n in wb.sheetnames]
    total_cols = 1 + len(racks) + 1  # Tab Name + per rack + Total

    if 'Summary' in wb.sheetnames:
        del wb['Summary']
    wb.create_sheet('Summary', 0)
    ws_s = wb['Summary']

    # Title
    ws_s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_c = ws_s.cell(row=1, column=1, value='Tab Summary by Rack')
    title_c.font = Font(bold=True, name='Arial', size=13)
    title_c.alignment = center_s
    title_c.border    = bd_s
    title_c.fill      = yellow_fill_s
    ws_s.row_dimensions[1].height = 28

    # Header row
    _s(ws_s.cell(row=2, column=1), 'Tab Name', bold=True, header=True)
    for c, rack in enumerate(racks, start=2):
        _s(ws_s.cell(row=2, column=c), str(rack), bold=True, header=True)
    _s(ws_s.cell(row=2, column=total_cols), 'Total', bold=True, header=True)

    # Data rows (exclude Summary itself)
    rack_totals = {r: 0 for r in racks}
    data_tabs   = [n for n in tabs_order if n != 'Summary']
    for i, tab_name in enumerate(data_tabs, start=3):
        _s(ws_s.cell(row=i, column=1), tab_name)
        row_total = 0
        for c, rack in enumerate(racks, start=2):
            count = tab_rack.get(tab_name, {}).get(rack, 0)
            _s(ws_s.cell(row=i, column=c), count if count > 0 else '')
            rack_totals[rack] += count
            row_total += count
        _s(ws_s.cell(row=i, column=total_cols), row_total, bold=True)

    # Grand total row
    tot_r = 3 + len(data_tabs)
    _s(ws_s.cell(row=tot_r, column=1), 'TOTAL', bold=True)
    grand = 0
    for c, rack in enumerate(racks, start=2):
        _s(ws_s.cell(row=tot_r, column=c), rack_totals[rack], bold=True)
        grand += rack_totals[rack]
    _s(ws_s.cell(row=tot_r, column=total_cols), grand, bold=True)

    # Column widths
    ws_s.column_dimensions['A'].width = 20
    for c in range(2, total_cols + 1):
        ws_s.column_dimensions[get_column_letter(c)].width = 14

    # ── 8. No fill + borders (preserve pink in Mismatches) ──────────────────
    log("  · Removing fills and applying borders")
    for sheet_name in wb.sheetnames:
        pcols = pink_col_indices if sheet_name == 'Mismatches' else []
        clear_and_border(wb[sheet_name], pink_cols=pcols)

    # ── 8c. Centre-align all cells across all tabs ──────────────────────────
    log("  · Aligning all cells to middle-centre")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                cell.alignment = center_align

    # ── 8b. Add NOTE column + autofilter to all tabs ───────────────────────
    log("  · Adding NOTE column and filters to all tabs")
    no_fill     = PatternFill(fill_type=None)
    yellow_fill = PatternFill('solid', start_color=YELLOW)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col_name in ['NOTE']:
            col_idx = ws.max_column + 1
            hdr = ws.cell(row=1, column=col_idx, value=col_name)
            hdr.font      = Font(bold=True, name='Arial', size=10)
            hdr.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            hdr.fill      = yellow_fill
            hdr.border    = thin_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.fill   = no_fill
                cell.border = thin_border()
        if ws.max_row > 1 and ws.max_column > 0:
            ws.auto_filter.ref = ws.dimensions

    # ── 8d. Grey-out Optics rows that are matched in Downlinks ──────────────
    if 'Optics' in wb.sheetnames and 'Downlinks' in wb.sheetnames:
        log("  · Greying out matched Optics rows")

        MATCH_COLS = [
            'Hostname', 'Interface', 'L/R', 'Rack', 'Elevation',
            'Source_port', 'DMARC1', 'DMARC2', 'Destination_port',
            'Z Hostname', 'Z Interface', 'Z L/R', 'Z Rack', 'Z Elevation',
        ]
        GREY_FONT_COLOR = 'FFD3D3D3'  # light grey

        ws_dl = wb['Downlinks']
        dl_header = [ws_dl.cell(row=1, column=c).value
                     for c in range(1, ws_dl.max_column + 1)]

        # Build a set of tuples from Downlinks for all match columns present
        dl_match_cols = [c for c in MATCH_COLS if c in dl_header]
        dl_col_idx    = {c: dl_header.index(c) + 1 for c in dl_match_cols}

        dl_keys = set()
        for r in range(2, ws_dl.max_row + 1):
            key = tuple(
                str(ws_dl.cell(row=r, column=dl_col_idx[c]).value or '').strip()
                for c in dl_match_cols
            )
            dl_keys.add(key)

        ws_op = wb['Optics']
        op_header = [ws_op.cell(row=1, column=c).value
                     for c in range(1, ws_op.max_column + 1)]

        # Only match on columns present in both sheets
        common_cols  = [c for c in dl_match_cols if c in op_header]
        op_col_idx   = {c: op_header.index(c) + 1 for c in common_cols}
        dl_col_idx_c = {c: dl_header.index(c) + 1 for c in common_cols}

        # Rebuild dl_keys using only common columns
        dl_keys_common = set()
        for r in range(2, ws_dl.max_row + 1):
            key = tuple(
                str(ws_dl.cell(row=r, column=dl_col_idx_c[c]).value or '').strip()
                for c in common_cols
            )
            dl_keys_common.add(key)

        for r in range(2, ws_op.max_row + 1):
            op_key = tuple(
                str(ws_op.cell(row=r, column=op_col_idx[c]).value or '').strip()
                for c in common_cols
            )
            if op_key in dl_keys_common:
                for c in range(1, ws_op.max_column + 1):
                    cell = ws_op.cell(row=r, column=c)
                    cell.font = Font(
                        bold=cell.font.bold if cell.font else False,
                        name=(cell.font.name if cell.font else None) or 'Arial',
                        size=(cell.font.size if cell.font else None) or 10,
                        color=GREY_FONT_COLOR,
                    )

    # ── 8e. Expand all columns and rows on every sheet ──────────────────────
    log("  · Expanding all columns and rows")
    for sheet_name in wb.sheetnames:
        autofit_sheet(wb[sheet_name])

    wb.save(output_path)

    # ── 9. Rename by top-2 Rack numbers ─────────────────────────────────────
    try:
        all_racks = []
        for sheet_name in wb.sheetnames:
            ws     = wb[sheet_name]
            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
            if 'Rack' in header:
                rc = header.index('Rack') + 1
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(row=r, column=rc).value
                    if val is not None:
                        try: all_racks.append(int(float(str(val))))
                        except ValueError: pass
        if all_racks:
            top2     = [str(r) for r, _ in Counter(all_racks).most_common(2)]
            new_name = '+'.join(top2) + '.xlsx'
            new_path = os.path.join(os.path.dirname(output_path), new_name)
            load_workbook(output_path).save(new_path)
            if new_path != output_path:
                os.remove(output_path)
            log(f"  ✓ Saved → {new_name}")
            return new_path
    except Exception as e:
        log(f"  ⚠ Could not rename by Rack: {e}")

    log(f"  ✓ Saved → {os.path.basename(output_path)}")


# ── Progress window ──────────────────────────────────────────────────────────

class ProgressWindow:
    def __init__(self, total):
        self.root = tk.Tk()
        self.root.title("Excel Formatter")
        self.root.geometry("620x420")
        self.root.resizable(False, False)
        self.root.attributes("-topmost", True)

        tk.Label(self.root, text="Excel Formatter",
                 font=("Arial", 14, "bold")).pack(pady=(14, 2))
        tk.Label(self.root, text="Processing files — please wait…",
                 font=("Arial", 11)).pack(pady=(0, 8))

        frame = tk.Frame(self.root)
        frame.pack(padx=12, pady=4, fill='both', expand=True)
        sb = tk.Scrollbar(frame)
        sb.pack(side='right', fill='y')
        self.text = tk.Text(frame, wrap='word', height=14,
                            font=("Courier", 10), yscrollcommand=sb.set)
        self.text.pack(side='left', fill='both', expand=True)
        sb.config(command=self.text.yview)

        self.bar = ttk.Progressbar(self.root, mode='determinate', maximum=total)
        self.bar.pack(fill='x', padx=12, pady=6)

        self.btn = tk.Button(self.root, text="Close", state='disabled',
                             command=self.root.destroy,
                             font=("Arial", 11), width=12)
        self.btn.pack(pady=8)

    def log(self, msg):
        self.text.insert('end', msg + '\n')
        self.text.see('end')
        self.root.update()

    def step(self):
        self.bar['value'] += 1
        self.root.update()

    def done(self):
        self.btn.config(state='normal')
        self.root.mainloop()


# ── Entry point ──────────────────────────────────────────────────────────────

def main():
    _dummy = tk.Tk()
    _dummy.withdraw()

    # ── Step 1: pick L&R + Cutsheet in one shot ─────────────────────────────
    refs = pick_files("Select Formula L&R + Cutsheets (pick both)")
    if len(refs) != 2:
        messagebox.showerror(
            "Cancelled",
            f"Need exactly 2 reference files (L&R + Cutsheets). Got {len(refs)}."
        )
        sys.exit(1)
    try:
        formula_path, cutsheet_path = identify_refs(refs)
    except ValueError as e:
        messagebox.showerror("Could not identify reference files", str(e))
        sys.exit(1)

    # ── Step 2: pick rack validation file(s) to format ──────────────────────
    input_files = pick_files("Select Excel files to format")
    if not input_files:
        messagebox.showerror("Cancelled", "No files selected.")
        sys.exit(1)

    _dummy.destroy()

    try:
        lr_lookup = load_lr_lookup(formula_path)
        cut_df    = load_cutsheet(cutsheet_path)
    except Exception as e:
        messagebox.showerror("Error loading reference files", str(e))
        sys.exit(1)

    # Always save to ~/Downloads
    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    os.makedirs(downloads, exist_ok=True)

    pw = ProgressWindow(total=len(input_files))
    pw.log(f"Formula L&R : {os.path.basename(formula_path)}")
    pw.log(f"Cutsheets   : {os.path.basename(cutsheet_path)}")
    pw.log(f"Files       : {len(input_files)}")
    pw.log(f"Output dir  : {downloads}\n")

    errors = []
    for i, input_path in enumerate(input_files, 1):
        base     = os.path.splitext(os.path.basename(input_path))[0]
        out_path = os.path.join(downloads, f"{base}_formatted.xlsx")
        pw.log(f"[{i}/{len(input_files)}]  {os.path.basename(input_path)}")
        try:
            process_file(input_path, out_path, lr_lookup, cut_df, pw.log)
        except Exception as e:
            pw.log(f"  ✗ ERROR: {e}")
            errors.append((os.path.basename(input_path), str(e)))
        pw.step()
        pw.log("")

    if errors:
        pw.log(f"⚠  {len(errors)} file(s) had errors:")
        for name, err in errors:
            pw.log(f"   • {name}: {err}")
    else:
        pw.log(f"✅  All done! Files saved to {downloads}")

    pw.done()


if __name__ == '__main__':
    main()
