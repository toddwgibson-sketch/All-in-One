"""
lv_portal_formatter.py
─────────────────────────────────────────────────────────────────────────────
Formats the NEW LV Portal validation export into the standard JBP15 report
layout (matches DG19_25-05.xlsx / highlight_slack_report_v2.py output).

INPUT
  - LV Portal export with sheets:
      Summary
      Optic Errors             (Source/Remote device+port, Rx Power, Patch Panel Matrix)
      FEC_BER Errors           (Device, Remote, PRE_FEC_BER, Optical RawBer, Lock Status, Issue, Matrix)
      Interface Down Errors    (Source location, Source/Remote device+port, Issue, Matrix)
  - Master cutsheet (JBP15_mastercutsheet.xlsx style):
      Cols: PhysA | DeviceA | RackA | Source_port | DMARC1 | DMARC2 |
            Destination_port | DeviceB | RackB | EasyMark+ | PhysB | Cable Color
      DeviceA is "<hostname> <interface>" (single space separated)

OUTPUT — five tabs matching DG19 format:
  Summary       (navy)    — total counts + per-rack breakdown
  Mispatches    (red)     — empty stub (new LV format has no LLDP sheet)
  Downlinks     (orange)  — 13 cols
  Optics        (brown)   — 16 cols (one row per failed channel)
  FEC Errors    (purple)  — 17 cols

Run:
    python lv_portal_formatter.py
"""

from __future__ import annotations

import json
import os
import re
import sys
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    HAS_TK = True
except Exception:
    HAS_TK = False


# ── Config (cutsheet path persisted between runs) ─────────────────────────────
CONFIG_FILE = Path.home() / ".lv_portal_formatter.json"


def load_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text())
        except Exception:
            return {}
    return {}


def save_config(cfg: dict) -> None:
    try:
        CONFIG_FILE.write_text(json.dumps(cfg, indent=2))
    except Exception:
        pass


# ── Style constants (match DG19 / highlight_slack_report_v2.py) ───────────────
TAB_SUMM = "1F4E79"
TAB_MISS = "C00000"
TAB_DOWN = "ED7D31"
TAB_OPT  = "833C00"
TAB_FEC  = "7030A0"

HDR_NAVY    = "1F4E79"
HDR_SRC     = "C0504D"   # Source_port column header
HDR_DM1     = "7F6000"   # DMARC1
HDR_DM2     = "375623"   # DMARC2
HDR_DEST    = "17375E"   # Destination_port / Z columns
HDR_POSS_A  = "833C00"   # Possible A side (Mispatches)
HDR_POSS_Z  = "375623"   # Possible Z side (Mispatches)
HDR_ACT     = "9C0006"   # Actual (Mispatches)
HDR_EXP     = "375623"   # Expected (Mispatches)
HDR_GREY    = "595959"   # History / DL Flag

POSS_A_BG   = "FDDCB5"   # peach
POSS_DM1_BG = "FFF2CC"   # light yellow
POSS_DM2_BG = "FCE4D6"   # light peach
POSS_Z_BG   = "D5F5E3"   # light green

# Optics "also downlink" greyed-out row colours
DL_GREY_BG    = "C8C8C8"
DL_GREY_LR_BG = "A8A8A8"
DL_GREY_FG    = "888888"
DL_FLAG_FG    = "666666"

WHITE = "FFFFFF"
MISS_BG = "FFF2CC"   # pale yellow — row not in cutsheet, manually verify

# Border styles (match DG19 — thin grey internal, medium edges, no header borders)
_THIN  = Side(style="thin",   color="AAAAAA")
_MED   = Side(style="medium", color="555555")

BORDER_LEFTMOST = Border(left=_MED,  right=_THIN, top=_MED, bottom=_MED)
BORDER_MIDDLE   = Border(left=_THIN, right=_THIN, top=_MED, bottom=_MED)
BORDER_RIGHTMOST = Border(left=_THIN, right=_MED, top=_MED, bottom=_MED)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def border_for(col: int, max_col: int) -> Border:
    """First column gets medium-left, last gets medium-right, all get medium top/bottom."""
    if col == 1:
        return BORDER_LEFTMOST
    if col == max_col:
        return BORDER_RIGHTMOST
    return BORDER_MIDDLE


def draw_pair_borders(ws, iface_col: int = 1, rack_col: int = 3, u_col: int = 4) -> None:
    """
    Overlay thick (medium-weight) borders around each physical+logical pair group
    so it's easy to see which rows belong to the same physical cable.

    Grouping key: (rack, U, base_port, cable_group) where
      base_port    = "swp22" from "swp22s0"
      cable_group  = 0 for s0,        1 for s1+s2 (the paired cable),
                     3 for s3         (each is its own physical cable)

    Same physical switch (same rack+U) + same QSFP port + same cable group
    = same physical cable. Rows in a group get a medium box drawn around them.
    Single-row groups get no special border (the standard thin-grey is kept).
    """
    if ws.max_row < 2:
        return

    thin  = Side(style="thin",   color="AAAAAA")
    thick = Side(style="medium", color="555555")

    def group_key(row: int) -> tuple:
        iface = str(ws.cell(row, iface_col).value or "")
        rack  = str(ws.cell(row, rack_col).value  or "")
        u     = str(ws.cell(row, u_col).value     or "")
        m = IFACE_RE.match(iface)
        if not m:
            return (rack, u, iface, -1)
        base_port = f"swp{m.group(1)}"
        lane = int(m.group(2))
        # Cable groups: s0=0, s1/s2=1 (paired), s3=3. Each is one physical cable.
        cable_group = 1 if lane in (1, 2) else lane
        return (rack, u, base_port, cable_group)

    max_col = ws.max_column
    dr = 2
    while dr <= ws.max_row:
        key = group_key(dr)
        if not key or key[3] < 0:
            dr += 1
            continue
        grp_end = dr
        while grp_end + 1 <= ws.max_row and group_key(grp_end + 1) == key:
            grp_end += 1

        # Only draw borders if it's an actual group (2+ rows)
        if grp_end > dr:
            for rr in range(dr, grp_end + 1):
                is_top = (rr == dr)
                is_bot = (rr == grp_end)
                for cc in range(1, max_col + 1):
                    is_left  = (cc == 1)
                    is_right = (cc == max_col)
                    ws.cell(rr, cc).border = Border(
                        top    = thick if is_top   else thin,
                        bottom = thick if is_bot   else thin,
                        left   = thick if is_left  else thin,
                        right  = thick if is_right else thin,
                    )
        dr = grp_end + 1


# ── Native pickers ────────────────────────────────────────────────────────────
def pick_file(title: str) -> str | None:
    if HAS_TK:
        root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title=title, filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All", "*.*")]
        )
        root.destroy()
        return path or None
    p = input(f"{title}\nPath: ").strip().strip('"')
    return p or None


def pick_multiple_files(title: str) -> list[str]:
    if HAS_TK:
        root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
        paths = filedialog.askopenfilenames(
            title=title, filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All", "*.*")]
        )
        root.destroy()
        return list(paths) if paths else []
    p = input(f"{title}\nPath(s) (semicolon-separated): ").strip().strip('"')
    return [x.strip() for x in p.split(";") if x.strip()] if p else []


def show_msg(title: str, msg: str, error: bool = False) -> None:
    if HAS_TK:
        root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
        (messagebox.showerror if error else messagebox.showinfo)(title, msg)
        root.destroy()
    else:
        prefix = "ERROR: " if error else ""
        print(f"{prefix}{title}: {msg}")


# ── Parse helpers ─────────────────────────────────────────────────────────────
RACK_RE = re.compile(r"Rack\s+(\S+)\s+U\s*(\S+)", re.IGNORECASE)
IFACE_RE = re.compile(r"swp(\d+)s(\d+)", re.IGNORECASE)


def parse_rack_u(s: str) -> tuple[str, str]:
    """`Rack 9108 U1` → ('9108', '1')."""
    if not s:
        return "", ""
    m = RACK_RE.match(str(s).strip())
    return (m.group(1), m.group(2)) if m else ("", "")


def iface_to_lr(iface: str) -> str:
    """
    `swp14s3` → '14L' or '14R'. The L/R side depends on BOTH the port number
    (even/odd) AND the sub-lane (s0-s3), per the standard JBP15 cutsheet:

        Even port + s0/s1 → L      Odd port + s0/s1 → R
        Even port + s2/s3 → R      Odd port + s2/s3 → L

    This is a fallback only — cutsheet PhysA values are always preferred when
    available.
    """
    if not iface:
        return ""
    m = IFACE_RE.match(str(iface).strip())
    if not m:
        return ""
    port = int(m.group(1))
    lane = int(m.group(2))
    even_port = (port % 2 == 0)
    low_lane  = (lane <= 1)
    # Truth table: (even_port, low_lane) → side
    #   (T, T) L  | (T, F) R | (F, T) R | (F, F) L
    side = "L" if even_port == low_lane else "R"
    return f"{port}{side}"


# ── Build cutsheet lookup ─────────────────────────────────────────────────────
def build_lookup(cutsheet_paths: list[str]) -> tuple[dict, dict]:
    """
    Build two lookups from the cutsheet:
      forward: (hostname, interface) → full row dict (T0 side as DeviceA)
      reverse: (t1_hostname, t1_interface) → {"rack": <rackB>, "elev": <elevB>}
               (T1 side as DeviceB, used to find Z-side rack/U for logical-pair
               lookups where the destination differs from the cutsheet pair entry)
    """
    forward: dict = {}
    reverse: dict = {}

    for path in cutsheet_paths:
        try:
            wb = load_workbook(path, data_only=True)
        except Exception as e:
            print(f"  WARN: could not load {os.path.basename(path)}: {e}")
            continue

        for ws in wb.worksheets:
            if ws.max_row < 2:
                continue

            # Find column indices by header (defensive against variant cutsheets)
            headers = {}
            for col in range(1, ws.max_column + 1):
                h = ws.cell(1, col).value
                if isinstance(h, str):
                    headers[h.strip().lower()] = col

            def col_of(*names: str) -> int | None:
                for n in names:
                    c = headers.get(n.lower())
                    if c:
                        return c
                return None

            # Column resolution supports TWO schemas:
            #  A) Combined: DeviceA="<host> <iface>", RackA="Rack X U Y", DeviceB="<t1> <iface>"
            #  B) Separated: Hostname, Interface, Rack (numeric), Elevation (numeric),
            #     Z Hostname, Z Interface, Z Rack, Z Elevation, with optional L/R column

            # T0 (A) side
            c_phys_a   = col_of("DeviceA Physical Port", "physa", "phys a",
                                "L/R", "L&R", "lr")
            c_dev_a    = col_of("DeviceA", "device a", "a device")    # combined "host iface"
            c_host_a   = col_of("Hostname", "host", "device", "device name")  # separate host
            c_iface_a  = col_of("Interface", "port", "iface")                  # separate iface
            c_rack_a   = col_of("RackA", "rack a", "a rack", "Rack")           # combined or numeric
            c_elev_a   = col_of("Elevation", "U", "elevation a")               # separate numeric U
            c_src_pp   = col_of("Source_port", "source port", "source_port")
            c_dm1      = col_of("DMARC1", "dmarc 1")
            c_dm2      = col_of("DMARC2", "dmarc 2")
            c_dst_pp   = col_of("Destination_port", "destination port", "dest_port")

            # T1 (B/Z) side
            c_dev_b    = col_of("DeviceB", "device b", "b device")    # combined
            c_host_b   = col_of("Z Hostname", "z host", "remote hostname", "remote device")
            c_iface_b  = col_of("Z Interface", "z port", "z iface", "remote interface", "remote port")
            c_rack_b   = col_of("RackB", "rack b", "b rack", "Z Rack")
            c_elev_b   = col_of("Z Elevation", "z u", "elevation b", "remote elevation")
            c_phys_b   = col_of("DeviceB Physical Port", "physb", "phys b",
                                "Z L/R", "Z L&R", "z lr")

            # Require at minimum a way to identify A-side hostname+interface AND src_pp
            have_a_combined = bool(c_dev_a)
            have_a_separate = bool(c_host_a and c_iface_a)
            if not (have_a_combined or have_a_separate) or not c_src_pp:
                continue  # Not a cutsheet sheet — skip

            def read_str(ws, row, col):
                if not col:
                    return ""
                v = ws.cell(row, col).value
                return str(v).strip() if v is not None else ""

            for row in range(2, ws.max_row + 1):
                # Get A-side hostname + interface, handling both schemas
                if have_a_separate:
                    hostname = read_str(ws, row, c_host_a)
                    iface    = read_str(ws, row, c_iface_a)
                else:
                    dev_a = read_str(ws, row, c_dev_a)
                    parts = dev_a.split()
                    if len(parts) < 2:
                        continue
                    hostname, iface = parts[0], parts[1]
                if not hostname or not iface:
                    continue

                # Resolve rack/U for A side — either combined "Rack X U Y" or separate
                rack_raw = read_str(ws, row, c_rack_a)
                rack_a, elev_a = parse_rack_u(rack_raw)
                if not rack_a and rack_raw:
                    # Looks like a separate numeric Rack column
                    rack_a = rack_raw
                    elev_a = read_str(ws, row, c_elev_a)
                rack_a_full = f"Rack {rack_a} U{elev_a}" if rack_a else ""

                # Z (T1) side — same dual-schema handling
                if c_host_b and c_iface_b:
                    dev_b   = read_str(ws, row, c_host_b)
                    iface_b = read_str(ws, row, c_iface_b)
                elif c_dev_b:
                    dev_b_full = read_str(ws, row, c_dev_b)
                    db_parts = dev_b_full.split()
                    dev_b   = db_parts[0] if db_parts else ""
                    iface_b = db_parts[1] if len(db_parts) > 1 else ""
                else:
                    dev_b, iface_b = "", ""

                rack_b_raw = read_str(ws, row, c_rack_b)
                rack_b, elev_b = parse_rack_u(rack_b_raw)
                if not rack_b and rack_b_raw:
                    rack_b = rack_b_raw
                    elev_b = read_str(ws, row, c_elev_b)
                rack_b_full = f"Rack {rack_b} U{elev_b}" if rack_b else ""

                forward[(hostname.lower(), iface.lower())] = {
                    "phys_a":      read_str(ws, row, c_phys_a),
                    "rack_a":      rack_a,
                    "elev_a":      elev_a,
                    "rack_a_full": rack_a_full,
                    "src_pp":      read_str(ws, row, c_src_pp),
                    "dmarc1":      read_str(ws, row, c_dm1),
                    "dmarc2":      read_str(ws, row, c_dm2),
                    "dest_pp":     read_str(ws, row, c_dst_pp),
                    "dev_b":       dev_b,
                    "iface_b":     iface_b,
                    "rack_b":      rack_b,
                    "elev_b":      elev_b,
                    "rack_b_full": rack_b_full,
                    "phys_b":      read_str(ws, row, c_phys_b),
                }

                # Reverse index the T1 side
                if dev_b and iface_b:
                    reverse[(dev_b.lower(), iface_b.lower())] = {
                        "rack": rack_b,
                        "elev": elev_b,
                    }
    return forward, reverse


# ── Matrix fallback parser (used only when cutsheet has no entry) ────────────
def parse_matrix(matrix: str) -> dict:
    """Extract rack/U and PP fields from LV portal's multi-line Patch Panel Matrix."""
    out = {"rack_a": "", "elev_a": "", "rack_b": "", "elev_b": "",
           "src_pp": "", "dmarc1": "", "dmarc2": "", "dest_pp": ""}
    if not matrix or str(matrix).strip().lower() == "missing":
        return out

    lines = [ln.strip() for ln in str(matrix).splitlines() if ln.strip()]
    rack_idxs = [i for i, ln in enumerate(lines) if RACK_RE.match(ln)]
    if rack_idxs:
        m = RACK_RE.match(lines[rack_idxs[0]])
        if m:
            out["rack_a"], out["elev_a"] = m.group(1), m.group(2)
    if len(rack_idxs) >= 2:
        m = RACK_RE.match(lines[rack_idxs[-1]])
        if m:
            out["rack_b"], out["elev_b"] = m.group(1), m.group(2)

    for ln in lines:
        u = ln.upper()
        if not u.startswith("PP."):
            continue
        if ".DH" in u and "MPO" in u:
            # DMARC line — DH1 vs DH2
            if ".DH1." in u or "DH10" in u:  # DH10 also appears in some sites
                if not out["dmarc1"]:
                    out["dmarc1"] = ln
            elif ".DH2." in u:
                if not out["dmarc2"]:
                    out["dmarc2"] = ln
        else:
            # Internal PP line — first = source, last = dest
            if not out["src_pp"]:
                out["src_pp"] = ln
            else:
                out["dest_pp"] = ln
    return out


def logical_pair_iface(iface: str) -> str | None:
    """
    Deprecated single-pair helper kept for callers that want a default pair.
    Returns the s1↔s2 pair (the "inner" sub-lane pair). Prefer
    `pair_search_order` below for full lookup-with-fallback.
    """
    if not iface:
        return None
    m = IFACE_RE.match(str(iface).strip())
    if not m:
        return None
    lane = int(m.group(2))
    port = int(m.group(1))
    if lane == 1:
        return f"swp{port}s2"
    if lane == 2:
        return f"swp{port}s1"
    return None


def pair_search_order(iface: str) -> list[str]:
    """
    Ordered list of sub-lane interfaces to try when the direct cutsheet
    lookup misses. The cutsheet stores ONE entry per physical cable, but
    different switches store different sub-lanes (some keep s0/s2, others
    keep s1/s3) depending on cable orientation. Confirmed by inspecting the
    LV portal's own Matrix output across multiple validation exports.

    Strategy: pick the closest sub-lane by distance, biased toward the
    inner lanes (s1, s2) since they're most often what the Matrix resolves
    to:

        s0 → [s0, s1]
        s1 → [s1, s2, s0]
        s2 → [s2, s1, s3]
        s3 → [s3, s2]

    First entry is the direct lookup; subsequent entries are pair candidates.
    """
    if not iface:
        return []
    m = IFACE_RE.match(str(iface).strip())
    if not m:
        return [iface]
    port = int(m.group(1))
    lane = int(m.group(2))
    base = f"swp{port}"

    # Order each lane by closeness, biased toward inner lanes (s1, s2)
    orderings = {
        0: [0, 1],         # s0 pairs to s1
        1: [1, 2, 0],      # s1 pairs to s2 first, then s0
        2: [2, 1, 3],      # s2 pairs to s1 first, then s3
        3: [3, 2],         # s3 pairs to s2
    }
    return [f"{base}s{n}" for n in orderings.get(lane, [lane])]


def lookup_or_matrix(lookup, hostname: str, iface: str, matrix: str,
                     remote_dev: str = "", remote_port: str = "") -> dict:
    """
    Cutsheet lookup with logical-pair fallback.

    `lookup` is a tuple (forward, reverse):
      forward: (hostname, interface) → cutsheet row dict (T0 side)
      reverse: (t1_dev, t1_iface)    → {"rack": ..., "elev": ...}

    Lookup order:
      1. Direct hit on (hostname, iface)
      2. Logical-pair fallback (s0↔s1, s2↔s3) — copies ONLY the four
         shared-MPO patch panel fields. T1 destination differs per sub-lane
         so Z-side info is sourced from LV portal's reported remote +
         reverse index for rack/U.
      3. Patch Panel Matrix from LV portal export
      4. Blank fields, flagged for manual review
    """
    forward, reverse = lookup
    host_lower = str(hostname).strip().lower()
    iface_lower = str(iface).strip().lower()

    # 1. Direct hit
    if (host_lower, iface_lower) in forward:
        info = dict(forward[(host_lower, iface_lower)])
        info["cutsheet_miss"] = False
        return info

    # 2. Logical-pair fallback — try sub-lanes in order of closeness. The
    # cutsheet typically stores only ONE entry per physical cable, but which
    # sub-lane (s0/s1/s2/s3) depends on the switch's cable orientation.
    # IMPORTANT: pair fallback copies ONLY the four patch panel fields
    # (Source_port, DMARC1, DMARC2, Destination_port). Everything else
    # (A-side rack/U, Z-side dev/port/rack/U) is sourced separately:
    #   A-side rack/U: any entry on the SAME source switch (same host)
    #                  — every cutsheet row for that host shows the same rack/U
    #   Z-side iface/host: from LV portal's Remote Device/Port
    #   Z-side rack/U: reverse-lookup the LV-reported remote against the cutsheet
    candidates = pair_search_order(iface_lower)
    pair_entry = None
    for cand in candidates[1:]:   # skip [0] — that was already tried above
        if (host_lower, cand) in forward:
            pair_entry = forward[(host_lower, cand)]
            break

    if pair_entry is not None:
        # A-side rack/U: take from ANY entry on the same source switch. Use
        # the pair entry itself as a convenient source (it's on the same host).
        # This is NOT inheriting the pair's connection — same host = same switch
        # = same physical rack/U regardless of which sub-lane we ask about.
        a_rack = pair_entry["rack_a"]
        a_elev = pair_entry["elev_a"]
        a_rack_full = pair_entry["rack_a_full"]

        # Z-side: LV portal tells us which T1 switch is involved. Reverse-lookup
        # that switch's rack/U; pair fallback is safe in the reverse direction
        # because a Z host sits at one rack/U regardless of sub-lane.
        z_dev   = str(remote_dev).strip()
        z_iface = str(remote_port).strip()
        z_rack, z_elev = "", ""
        if z_dev and z_iface:
            for z_cand in pair_search_order(z_iface.lower()):
                r = reverse.get((z_dev.lower(), z_cand))
                if r:
                    z_rack, z_elev = r["rack"], r["elev"]
                    break

        return {
            "phys_a":        iface_to_lr(iface),
            "rack_a":        a_rack,                       # same-host lookup
            "elev_a":        a_elev,                       # same-host lookup
            "rack_a_full":   a_rack_full,
            "src_pp":        pair_entry["src_pp"],         # ✓ patch panel from pair
            "dmarc1":        pair_entry["dmarc1"],         # ✓ DMARC from pair
            "dmarc2":        pair_entry["dmarc2"],         # ✓ DMARC from pair
            "dest_pp":       pair_entry["dest_pp"],        # ✓ patch panel from pair
            "dev_b":         z_dev,                        # from LV portal
            "iface_b":       z_iface,                      # from LV portal
            "rack_b":        z_rack,                       # reverse-lookup of LV remote
            "elev_b":        z_elev,                       # reverse-lookup of LV remote
            "rack_b_full":   f"Rack {z_rack} U{z_elev}" if z_rack else "",
            "phys_b":        iface_to_lr(z_iface),
            "cutsheet_miss": False,
        }

    # 3. Patch Panel Matrix fallback (may still be 'missing')
    parsed = parse_matrix(matrix)
    has_matrix = any(parsed.values())
    return {
        "phys_a":        iface_to_lr(iface),
        "rack_a":        parsed["rack_a"],
        "elev_a":        parsed["elev_a"],
        "rack_a_full":   f"Rack {parsed['rack_a']} U{parsed['elev_a']}" if parsed["rack_a"] else "",
        "src_pp":        parsed["src_pp"],
        "dmarc1":        parsed["dmarc1"],
        "dmarc2":        parsed["dmarc2"],
        "dest_pp":       parsed["dest_pp"],
        "dev_b":         "",
        "iface_b":       "",
        "rack_b":        parsed["rack_b"],
        "elev_b":        parsed["elev_b"],
        "rack_b_full":   f"Rack {parsed['rack_b']} U{parsed['elev_b']}" if parsed["rack_b"] else "",
        "phys_b":        "",
        "cutsheet_miss": True,
        "matrix_only":   has_matrix,
    }


# ── Read LV portal sheets ─────────────────────────────────────────────────────
def find_sheet(wb, *patterns: str):
    for name in wb.sheetnames:
        low = name.lower()
        for p in patterns:
            if p.lower() in low:
                return wb[name]
    return None


def header_map(ws) -> dict:
    m = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if isinstance(h, str):
            m[h.strip().lower()] = c
    return m


def gv(ws, row: int, hmap: dict, *names: str) -> str:
    for n in names:
        c = hmap.get(n.lower())
        if c:
            v = ws.cell(row, c).value
            if v is not None:
                return str(v)
    return ""


# ── Optics parser: split per-channel multi-line cell into rows ────────────────
CHAN_LINE_RE = re.compile(r"channel_?(\d+)\s*:\s*([-+]?\d+\.?\d*)\s*(\(failed\))?", re.IGNORECASE)


def parse_rx_power(s: str) -> list[tuple[str, str, bool]]:
    """`channel_7: 0.37\nchannel_8: -7.88 (failed)` → [('7', '0.37', False), ('8', '-7.88', True)]"""
    out = []
    if not s:
        return out
    for line in str(s).splitlines():
        m = CHAN_LINE_RE.search(line)
        if m:
            out.append((m.group(1), m.group(2), bool(m.group(3))))
    return out


# ── Sheet builders ────────────────────────────────────────────────────────────
def style_header(cell, fill_hex: str, font_size: int = 10) -> None:
    cell.fill = fill(fill_hex)
    cell.font = Font(bold=True, color=WHITE, name="Arial", size=font_size)
    cell.alignment = center()


def write_headers(ws, headers: list[tuple[str, str]]) -> None:
    """headers = [(text, fill_hex), ...]"""
    for col, (text, fhex) in enumerate(headers, start=1):
        style_header(ws.cell(1, col, text), fhex)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def write_data_cell(ws, row: int, col: int, val, fg: str = "000000",
                    bold: bool = False, bg: str = WHITE,
                    max_col: int | None = None) -> None:
    c = ws.cell(row, col, val if val != "" else None)
    c.fill = fill(bg)
    c.font = Font(color=fg, name="Arial", size=9, bold=bold)
    c.alignment = center()
    if max_col is not None:
        c.border = border_for(col, max_col)


def set_widths(ws, widths: list[int]) -> None:
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Build Downlinks (13 cols) ────────────────────────────────────────────────
def build_downlinks(wb_out, src_ws, lookup: dict) -> tuple[int, dict]:
    """
    Returns (count, downlink_set) where downlink_set = {(host, iface), ...}
    for the Optics "Also Downlink" cross-reference.
    """
    ws = wb_out.create_sheet("Downlinks")
    ws.sheet_properties.tabColor = TAB_DOWN

    headers = [
        ("Interface",        HDR_NAVY),
        ("L&R",              HDR_NAVY),
        ("Rack",             HDR_NAVY),
        ("Elevation",        HDR_NAVY),
        ("Source_port",      HDR_SRC),
        ("DMARC1",           HDR_DM1),
        ("DMARC2",           HDR_DM2),
        ("Destination_port", HDR_DEST),
        ("Z Interface",      HDR_DEST),
        ("L&R",              HDR_DEST),
        ("Z Rack",           HDR_DEST),
        ("Z Elevation",      HDR_DEST),
        ("History",          HDR_GREY),
    ]
    write_headers(ws, headers)
    set_widths(ws, [12, 6, 8, 8, 32, 32, 32, 32, 12, 6, 8, 10, 12])

    downlink_set: set = set()
    if src_ws is None or src_ws.max_row < 2:
        return 0, downlink_set

    hmap = header_map(src_ws)
    out_row = 2
    count = 0
    for r in range(2, src_ws.max_row + 1):
        host = gv(src_ws, r, hmap, "Source Device Name")
        port = gv(src_ws, r, hmap, "Source Device Port")
        if not host or not port:
            continue

        downlink_set.add((host.lower(), port.lower()))

        rem_dev  = gv(src_ws, r, hmap, "Remote Device Name")
        rem_port = gv(src_ws, r, hmap, "Remote Device Port")
        matrix   = gv(src_ws, r, hmap, "Patch Panel Matrix")

        info = lookup_or_matrix(lookup, host, port, matrix,
                                remote_dev=rem_dev, remote_port=rem_port)

        z_iface = info["iface_b"] or rem_port
        z_lr    = info["phys_b"] or iface_to_lr(z_iface)
        z_rack  = info["rack_b"]
        z_elev  = info["elev_b"]

        # Cutsheet miss → tint row pale yellow + note in History
        miss = info.get("cutsheet_miss", False)
        row_bg = MISS_BG if miss else WHITE
        history_note = "⚠ Not in cutsheet" if miss else ""

        values = [
            port,                     # Interface
            info["phys_a"] or iface_to_lr(port),
            info["rack_a"],
            info["elev_a"],
            info["src_pp"],
            info["dmarc1"],
            info["dmarc2"],
            info["dest_pp"],
            z_iface,
            z_lr,
            z_rack,
            z_elev,
            history_note,             # History
        ]
        ws.row_dimensions[out_row].height = 15
        for col, v in enumerate(values, start=1):
            bold = (col == 2)  # L&R bold
            write_data_cell(ws, out_row, col, v, bold=bold, bg=row_bg,
                            max_col=len(headers))
        out_row += 1
        count += 1
    draw_pair_borders(ws)
    return count, downlink_set


# ── Build Optics (16 cols, per-channel rows) ─────────────────────────────────
def build_optics(wb_out, src_ws, lookup: dict, downlink_set: set) -> int:
    ws = wb_out.create_sheet("Optics")
    ws.sheet_properties.tabColor = TAB_OPT

    headers = [
        ("Interface",        HDR_NAVY),
        ("L&R",              HDR_NAVY),
        ("Rack",             HDR_NAVY),
        ("Elevation",        HDR_NAVY),
        ("Channel",          HDR_NAVY),
        ("Measured (dBm)",   HDR_NAVY),
        ("Source_port",      HDR_NAVY),
        ("DMARC1",           HDR_NAVY),
        ("DMARC2",           HDR_NAVY),
        ("Destination_port", HDR_NAVY),
        ("Z Interface",      HDR_NAVY),
        ("Z L&R",            HDR_NAVY),
        ("Z Rack",           HDR_NAVY),
        ("Z Elevation",      HDR_NAVY),
        ("DL Flag",          HDR_GREY),
        ("History",          HDR_GREY),
    ]
    write_headers(ws, headers)
    set_widths(ws, [12, 6, 8, 8, 8, 12, 32, 32, 32, 32, 12, 6, 8, 10, 22, 12])

    if src_ws is None or src_ws.max_row < 2:
        return 0

    hmap = header_map(src_ws)
    out_row = 2
    count = 0
    for r in range(2, src_ws.max_row + 1):
        host = gv(src_ws, r, hmap, "Source Device Name")
        port = gv(src_ws, r, hmap, "Source Device Port")
        if not host or not port:
            continue

        rx_raw   = gv(src_ws, r, hmap, "Rx Power")
        rem_dev  = gv(src_ws, r, hmap, "Remote Device Name")
        rem_port = gv(src_ws, r, hmap, "Remote Device Port")
        matrix   = gv(src_ws, r, hmap, "Patch Panel Matrix")
        info     = lookup_or_matrix(lookup, host, port, matrix,
                                    remote_dev=rem_dev, remote_port=rem_port)

        is_dl = (host.lower(), port.lower()) in downlink_set
        miss = info.get("cutsheet_miss", False)
        # DL greying takes priority; otherwise yellow for cutsheet miss
        if is_dl:
            row_bg    = DL_GREY_BG
            lr_bg     = DL_GREY_LR_BG
            text_fg   = DL_GREY_FG
            flag_text = "⬇️ Also Downlink — skip"
            history_note = ""
        elif miss:
            row_bg    = MISS_BG
            lr_bg     = MISS_BG
            text_fg   = "000000"
            flag_text = ""
            history_note = "⚠ Not in cutsheet"
        else:
            row_bg    = WHITE
            lr_bg     = WHITE
            text_fg   = "000000"
            flag_text = ""
            history_note = ""

        z_iface = info["iface_b"] or gv(src_ws, r, hmap, "Remote Device Port")
        z_lr    = info["phys_b"] or iface_to_lr(z_iface)

        # One row per failed channel
        channels = [(ch, val) for ch, val, failed in parse_rx_power(rx_raw) if failed]
        if not channels:
            # Fallback: emit a single row with raw rx text
            channels = [("", rx_raw or "")]

        for ch, val in channels:
            row_values = [
                (port,                                       row_bg,  False),
                (info["phys_a"] or iface_to_lr(port),        lr_bg,   True),
                (info["rack_a"],                             row_bg,  False),
                (info["elev_a"],                             row_bg,  False),
                (ch,                                         row_bg,  False),
                (val,                                        row_bg,  False),
                (info["src_pp"],                             row_bg,  False),
                (info["dmarc1"],                             row_bg,  False),
                (info["dmarc2"],                             row_bg,  False),
                (info["dest_pp"],                            row_bg,  False),
                (z_iface,                                    row_bg,  False),
                (z_lr,                                       lr_bg,   True),
                (info["rack_b"],                             row_bg,  False),
                (info["elev_b"],                             row_bg,  False),
                (flag_text,                                  row_bg,  is_dl),
                (history_note,                               row_bg,  False),
            ]
            ws.row_dimensions[out_row].height = 15
            for col, (v, bg, bold) in enumerate(row_values, start=1):
                fg = DL_FLAG_FG if (col == 15 and is_dl) else text_fg
                write_data_cell(ws, out_row, col, v, fg=fg, bold=bold, bg=bg,
                                max_col=len(headers))
            out_row += 1
            count += 1
    draw_pair_borders(ws)
    return count


# ── Build FEC Errors (17 cols) ────────────────────────────────────────────────
RAWBER_LANE_RE = re.compile(r"lane_?(\d+)\s*:\s*([-+]?\d+\.?\d*[eE]?[-+]?\d*)\s*(\(failed\))?")


def classify_ber_severity(ber: str) -> tuple[str, str, str]:
    """
    Classify a Pre-FEC BER value into a severity band.

    Bands (Pre-FEC BER, threshold is 1e-07):
        Marginal   1e-07 ≤ ber < 1e-06    pale yellow
        Warning    1e-06 ≤ ber < 1e-05    orange
        Severe     ber ≥ 1e-05            red

    Returns (label, background_hex, font_color_hex).
    Empty/unparseable BER → ("", WHITE, "000000").
    """
    if not ber:
        return ("", WHITE, "000000")
    try:
        val = float(ber)
    except (TypeError, ValueError):
        return ("", WHITE, "000000")
    if val < 1e-7:
        return ("", WHITE, "000000")
    if val < 1e-6:
        return ("Marginal", "FFF2CC", "7F6000")     # pale yellow / dark amber
    if val < 1e-5:
        return ("Warning",  "FCE4D6", "9C5700")     # light peach / dark orange
    return     ("Severe",   "FFC7CE", "9C0006")     # pink / dark red


def extract_max_failed_ber(raw_ber: str, pre_fec: str) -> tuple[str, str]:
    """
    Return (lock_status_summary, ber_value) from the LV portal cells.

    If PRE_FEC_BER is present and not 'missing', use it directly.
    Otherwise scan Optical RawBer for failed lanes and pick the worst one,
    formatting lock_status as `RAW_BER_MAX=<val> (channel(s)=<n>) > 1e-07`
    to match the DG19 format.
    """
    if pre_fec and pre_fec.lower() != "missing":
        return ("", pre_fec)

    failed = []
    for line in (raw_ber or "").splitlines():
        m = RAWBER_LANE_RE.search(line)
        if m and m.group(3):
            try:
                failed.append((int(m.group(1)), float(m.group(2)), m.group(2)))
            except ValueError:
                pass
    if not failed:
        return ("", "")
    failed.sort(key=lambda x: -x[1])  # worst (highest BER) first
    worst_lane, _, worst_val = failed[0]
    lock = f"RAW_BER_MAX={worst_val} (channel(s)={worst_lane}) > 1e-07"
    return (lock, worst_val)


def build_fec(wb_out, src_ws, lookup: dict, downlink_set: set) -> int:
    ws = wb_out.create_sheet("FEC Errors")
    ws.sheet_properties.tabColor = TAB_FEC

    headers = [
        ("Interface",        HDR_NAVY),
        ("L&R",              HDR_NAVY),
        ("Rack",             HDR_NAVY),
        ("Elevation",        HDR_NAVY),
        ("Lock Status",      HDR_NAVY),
        ("Pre-FEC BER",      HDR_NAVY),
        ("Severity",         HDR_NAVY),
        ("Source_port",      HDR_NAVY),
        ("DMARC1",           HDR_NAVY),
        ("DMARC2",           HDR_NAVY),
        ("Destination_port", HDR_NAVY),
        ("Z Interface",      HDR_NAVY),
        ("Z L&R",            HDR_NAVY),
        ("Z Rack",           HDR_NAVY),
        ("Z Elevation",      HDR_NAVY),
        ("Remote Interface", HDR_NAVY),
        ("DL Flag",          HDR_GREY),
        ("History",          HDR_GREY),
    ]
    write_headers(ws, headers)
    set_widths(ws, [12, 6, 8, 8, 40, 12, 11, 32, 32, 32, 32, 12, 6, 8, 10, 14, 22, 12])

    if src_ws is None or src_ws.max_row < 2:
        return 0

    hmap = header_map(src_ws)
    out_row = 2
    count = 0
    for r in range(2, src_ws.max_row + 1):
        host = gv(src_ws, r, hmap, "Device Name", "Source Device Name")
        port = gv(src_ws, r, hmap, "Device Port", "Source Device Port")
        if not host or not port:
            continue

        rem_dev  = gv(src_ws, r, hmap, "Remote Device Name")
        rem_port = gv(src_ws, r, hmap, "Remote Device Port")
        pre_fec  = gv(src_ws, r, hmap, "PRE_FEC_BER")
        raw_ber  = gv(src_ws, r, hmap, "Optical RawBer")
        lock_in  = gv(src_ws, r, hmap, "Lock Status")
        matrix   = gv(src_ws, r, hmap, "Patch Panel Matrix")

        info = lookup_or_matrix(lookup, host, port, matrix,
                                remote_dev=rem_dev, remote_port=rem_port)

        lock_derived, ber = extract_max_failed_ber(raw_ber, pre_fec)
        lock_status = lock_in if (lock_in and lock_in.lower() != "missing") else lock_derived

        is_dl = (host.lower(), port.lower()) in downlink_set
        miss  = info.get("cutsheet_miss", False)
        if is_dl:
            row_bg  = DL_GREY_BG
            lr_bg   = DL_GREY_LR_BG
            text_fg = DL_GREY_FG
            flag    = "⬇️ Also Downlink — skip"
            history_note = ""
        elif miss:
            row_bg  = MISS_BG
            lr_bg   = MISS_BG
            text_fg = "000000"
            flag    = ""
            history_note = "⚠ Not in cutsheet"
        else:
            row_bg  = WHITE
            lr_bg   = WHITE
            text_fg = "000000"
            flag    = ""
            history_note = ""

        z_iface = info["iface_b"] or rem_port
        z_lr    = info["phys_b"] or iface_to_lr(z_iface)

        sev_label, sev_bg, sev_fg = classify_ber_severity(ber)
        # If the row is greyed-out (Also Downlink), severity styling is overridden
        # to keep the row visually muted.
        if is_dl:
            sev_bg, sev_fg = row_bg, text_fg

        row_values = [
            (port,                                       row_bg,  False),
            (info["phys_a"] or iface_to_lr(port),        lr_bg,   True),
            (info["rack_a"],                             row_bg,  False),
            (info["elev_a"],                             row_bg,  False),
            (lock_status,                                row_bg,  False),
            (ber,                                        row_bg,  False),
            (sev_label,                                  sev_bg,  True),   # ← Severity
            (info["src_pp"],                             row_bg,  False),
            (info["dmarc1"],                             row_bg,  False),
            (info["dmarc2"],                             row_bg,  False),
            (info["dest_pp"],                            row_bg,  False),
            (z_iface,                                    row_bg,  False),
            (z_lr,                                       lr_bg,   True),
            (info["rack_b"],                             row_bg,  False),
            (info["elev_b"],                             row_bg,  False),
            (rem_port,                                   row_bg,  False),
            (flag,                                       row_bg,  is_dl),
            (history_note,                               row_bg,  False),
        ]
        ws.row_dimensions[out_row].height = 15
        for col, (v, bg, bold) in enumerate(row_values, start=1):
            # Column 7 = Severity → use its own font colour for label
            # Column 17 (DL Flag) keeps its existing DL-grey handling
            if col == 7:
                fg = sev_fg
            elif col == 17 and is_dl:
                fg = DL_FLAG_FG
            else:
                fg = text_fg
            write_data_cell(ws, out_row, col, v, fg=fg, bold=bold, bg=bg,
                            max_col=len(headers))
        out_row += 1
        count += 1
    draw_pair_borders(ws)
    return count


# ── Build Mispatches (27-col headers, empty since new LV format has no LLDP) ─
def build_mispatches(wb_out) -> int:
    ws = wb_out.create_sheet("Mispatches")
    ws.sheet_properties.tabColor = TAB_MISS

    headers = [
        ("Interface",            HDR_NAVY),
        ("L&R",                  HDR_NAVY),
        ("Rack",                 HDR_NAVY),
        ("Elevation",            HDR_NAVY),
        ("Source_port",          HDR_SRC),
        ("DMARC1",               HDR_DM1),
        ("DMARC2",               HDR_DM2),
        ("Destination_port",     HDR_DEST),
        ("Z Interface",          HDR_DEST),
        ("L&R",                  HDR_DEST),
        ("Z Rack",               HDR_DEST),
        ("Z Elevation",          HDR_DEST),
        ("Possible Device A",    HDR_POSS_A),
        ("Possible Rack / U",    HDR_POSS_A),
        ("Possible Source Port", HDR_POSS_A),
        ("Possible DMARC1",      HDR_DM1),
        ("Possible DMARC2",      HDR_SRC),
        ("Possible Dest Port",   HDR_POSS_Z),
        ("Possible T1 Rack / U", HDR_POSS_Z),
        ("Possible T1 Port",     HDR_POSS_Z),
        ("Act. Interface",       HDR_ACT),
        ("Act. Rack",            HDR_ACT),
        ("Act. Elevation",       HDR_ACT),
        ("Exp. Interface",       HDR_EXP),
        ("Exp. Rack",            HDR_EXP),
        ("Exp. Elevation",       HDR_EXP),
        ("History",              HDR_GREY),
    ]
    write_headers(ws, headers)
    set_widths(ws, [12, 6, 8, 8, 32, 32, 32, 32, 12, 6, 8, 10,
                    14, 14, 32, 32, 32, 32, 14, 12,
                    12, 10, 10, 12, 10, 10, 12])

    note = ws.cell(2, 1,
        "No LLDP/mispatch sheet in source LV portal export — "
        "if a future export includes one (sheet name contains 'lldp' or 'mispatch'), "
        "this tab will populate automatically.")
    note.font = Font(italic=True, color="808080", name="Arial", size=10)
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=12)
    ws.row_dimensions[2].height = 30
    return 0


# ── Build Summary tab (matches DG19 layout) ──────────────────────────────────
def build_summary(wb_out, report_name: str, n_miss: int, n_down: int,
                  n_opt: int, n_fec: int, racks: dict) -> None:
    """Brief summary: report name, total issue count, and per-type breakdown."""
    ws = wb_out.create_sheet("Summary", 0)  # first
    ws.sheet_properties.tabColor = TAB_SUMM

    # Title
    title = ws.cell(1, 2, "VALIDATION REPORT — SUMMARY")
    style_header(title, "1F4E79", font_size=14)
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, end_row=1, start_column=2, end_column=4)

    # Report name
    rpt = ws.cell(2, 2, f"Report: {report_name}")
    rpt.fill = fill("0D7377")
    rpt.font = Font(color=WHITE, name="Arial", size=10, italic=True)
    rpt.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=2, end_row=2, start_column=2, end_column=4)

    # Total issues
    total = n_miss + n_down + n_opt + n_fec
    style_header(ws.cell(4, 2, "TOTAL ISSUES"), "1F4E79")
    ws.merge_cells(start_row=4, end_row=4, start_column=2, end_column=3)
    tot_cell = ws.cell(4, 4, total)
    tot_cell.fill = fill("1F4E79")
    tot_cell.font = Font(color=WHITE, name="Arial", size=14, bold=True)
    tot_cell.alignment = center()
    ws.row_dimensions[4].height = 24

    # Error-type breakdown
    style_header(ws.cell(6, 2, "ERROR TYPE"),    "1F4E79")
    style_header(ws.cell(6, 3, "COUNT"),         "1F4E79")
    ws.merge_cells(start_row=6, end_row=6, start_column=3, end_column=4)

    breakdown = [
        ("Mispatches", n_miss, "C00000"),  # red
        ("Downlinks",  n_down, "ED7D31"),  # orange
        ("Optics",     n_opt,  "833C00"),  # brown
        ("FEC Errors", n_fec,  "7030A0"),  # purple
    ]
    for i, (label, n, hex_color) in enumerate(breakdown):
        r = 7 + i
        ws.row_dimensions[r].height = 20
        # Label cell
        lab = ws.cell(r, 2, label)
        lab.fill = fill(hex_color)
        lab.font = Font(color=WHITE, name="Arial", size=10, bold=True)
        lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # Count cell (merged across 3-4)
        cnt = ws.cell(r, 3, n)
        cnt.fill = fill("F2F2F2")
        cnt.font = Font(color="000000", name="Arial", size=11, bold=True)
        cnt.alignment = center()
        ws.merge_cells(start_row=r, end_row=r, start_column=3, end_column=4)

    set_widths(ws, [3, 24, 12, 12])


# ── Main ──────────────────────────────────────────────────────────────────────
def main() -> None:
    cfg = load_config()

    print("=" * 60)
    print("  LV Portal Validation Formatter")
    print("=" * 60)

    # Step 1: cutsheet(s)
    saved_paths = cfg.get("cutsheet_paths", [])
    saved_paths = [p for p in saved_paths if os.path.isfile(p)]

    cutsheet_paths: list[str] = []
    if saved_paths:
        names = "\n".join(f"  • {os.path.basename(p)}" for p in saved_paths)
        if HAS_TK:
            root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
            use_saved = messagebox.askyesno(
                "Cutsheets",
                f"Use saved cutsheet(s)?\n\n{names}\n\nNo = pick different ones.",
            )
            root.destroy()
        else:
            print(f"Saved cutsheets:\n{names}")
            use_saved = input("Use these? (y/n): ").strip().lower() != "n"
        if use_saved:
            cutsheet_paths = saved_paths

    if not cutsheet_paths:
        show_msg("Select Cutsheets", "Select cutsheet(s). Hold Ctrl for multiple.")
        cutsheet_paths = pick_multiple_files("Select Cutsheet(s) — Ctrl for multiple")
        if not cutsheet_paths:
            show_msg("Cancelled", "No cutsheets selected.", error=True); sys.exit(0)
        cfg["cutsheet_paths"] = cutsheet_paths
        save_config(cfg)

    print(f"Loading {len(cutsheet_paths)} cutsheet(s)...")
    lookup = build_lookup(cutsheet_paths)
    print(f"  {len(lookup[0])} forward + {len(lookup[1])} reverse entries indexed")

    # Step 2: LV portal report
    time.sleep(0.3)
    report_path = pick_file("Select LV Portal Validation Export (.xlsx)")
    if not report_path:
        show_msg("Cancelled", "No file selected.", error=True); sys.exit(0)

    print(f"Processing: {os.path.basename(report_path)}")
    wb_src = load_workbook(report_path, data_only=True)

    ws_lldp  = find_sheet(wb_src, "lldp", "mismatch", "mispatch")
    ws_optics = find_sheet(wb_src, "optic")
    ws_fec    = find_sheet(wb_src, "fec")
    ws_iface  = find_sheet(wb_src, "interface down", "interface_down", "downlink")

    if not any([ws_optics, ws_fec, ws_iface]):
        show_msg("Error", "Could not find Optic / FEC / Interface Down sheets.", error=True)
        sys.exit(1)

    # Build output workbook
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    n_miss = build_mispatches(wb_out)
    n_down, downlink_set = build_downlinks(wb_out, ws_iface, lookup)
    n_opt = build_optics(wb_out, ws_optics, lookup, downlink_set)
    n_fec = build_fec(wb_out, ws_fec, lookup, downlink_set)

    report_name = Path(report_path).stem
    build_summary(wb_out, report_name, n_miss, n_down, n_opt, n_fec, {})

    # Save
    src_path = Path(report_path)
    out_path = src_path.with_name(src_path.stem + "_formatted.xlsx")
    wb_out.save(out_path)

    print()
    print(f"Mispatches: {n_miss}")
    print(f"Downlinks:  {n_down}")
    print(f"Optics:     {n_opt} channel-rows")
    print(f"FEC:        {n_fec}")
    print()
    print(f"Saved: {out_path}")
    show_msg(
        "Done",
        f"Saved to:\n{out_path}\n\n"
        f"Mispatches: {n_miss}\nDownlinks: {n_down}\n"
        f"Optics: {n_opt}\nFEC: {n_fec}",
    )


if __name__ == "__main__":
    main()
