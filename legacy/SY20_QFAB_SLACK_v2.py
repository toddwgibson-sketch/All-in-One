"""
Combine multiple per-building audit xlsx files into a single formatted workbook,
enriched with cutsheet data.

Usage:
    python SY20_QFAB_SLACK_v2.py

Workflow:
    1. Dialog: pick the cutsheet xlsx
    2. Dialog: pick one or more per-building input xlsx files
    3. Script does the rest and saves to ~/Downloads/<labels>.xlsx

Input report formats supported (auto-detected per file):
  OLD  sheet: full_path_lldp_with_int_down  — columns include PP_A / PP_B, no index
  NEW  sheet: lldp_with_int_down            — has index column, no PP_A / PP_B
  Both formats are normalised to the same column shape before processing,
  so old and new files can be mixed freely in a single run.

Cutsheet formats supported (auto-detected per row):
  EC6 all-in-one  (8 cols): endpoint | rack | PP_A | PP_B | PP_C | other_endpoint | other_rack | spare
  QFAB            (6 cols): endpoint | rack | PP_A | PP_B | other_endpoint | other_rack
  Direct          (4 cols): endpoint | rack | other_endpoint | other_rack   (no PP columns)
"""

import re
import sys
from pathlib import Path
from collections import Counter

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────── Styles ──
YELLOW_FILL = PatternFill("solid", start_color="FFFF00")
PINK_FILL   = PatternFill("solid", start_color="FFC0CB")
HEADER_FONT = Font(name="Arial", bold=True, color="000000")
BODY_FONT   = Font(name="Arial")
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=False)
THIN        = Side(style="thin", color="000000")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PINK_COLS = [
    "Active Host", "Act. Interface", "Act. Rack", "Act. Elevation",
    "Cutsheet PP_A", "Cutsheet PP_B", "Cutsheet PP_C",
    "Cutsheet Other End", "Cutsheet Other End Rack",
]

# Columns appended by the cutsheet enrichment step.
# PP_C is populated for EC6-style (3-PP) cutsheets; None / "<=>" for QFAB (2-PP).
CUT_COLS = [
    "Cutsheet PP_A",
    "Cutsheet PP_B",
    "Cutsheet PP_C",
    "Cutsheet Other End",
    "Cutsheet Other End Rack",
]

# Normalised LLDP column order expected by the rest of the pipeline.
# Both old and new input formats are massaged into this shape.
LLDP_COLS = [
    "Source",
    "Hostname", "Interface", "Building", "Rack", "Elevation",
    "PP_A", "PP_B",
    "Active Host", "Act. Interface", "Act. Building", "Act. Rack", "Act. Elevation",
    "Expected Hostname", "Exp. Interface", "Exp. Building", "Exp. Rack", "Exp. Elevation",
]

# Per-tab column-drop specs
DROP_SPECS = {
    "T2-T1 Downlink": ["Building", "Exp. Building",
                       "Active Host", "Act. Interface", "Act. Building", "Act. Rack", "Act. Elevation"],
    "T1-T0 Downlink": ["Building", "Exp. Building",
                       "PP_A", "PP_B",
                       "Active Host", "Act. Interface", "Act. Building", "Act. Rack", "Act. Elevation"],
    "T2-T1 Mismatch": ["Building", "Act. Building", "Exp. Building"],
    "T1-T0 Mismatch": ["Building", "PP_A", "PP_B", "Act. Building", "Exp. Building"],
    "optics":         ["index", "Building", "Rack", "Elevation"],
    "fec_ber":        ["index", "BER", "Lock", "Rack", "Elevation",
                       "Remote Host", "Remote Interface", "Reason"],
}

# ─────────────────────────────────────────────────────── File selection ──────
def select_files():
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    cutsheet = filedialog.askopenfilename(
        title="Select the CUTSHEET xlsx",
        filetypes=[("Excel files", "*.xlsx *.xlsm")],
    )
    if not cutsheet:
        sys.exit("No cutsheet selected.")
    inputs = filedialog.askopenfilenames(
        title="Select INPUT xlsx files to combine",
        filetypes=[("Excel files", "*.xlsx *.xlsm")],
    )
    if not inputs:
        sys.exit("No input files selected.")
    root.destroy()
    return cutsheet, list(inputs)


def extract_label(path):
    """Pull a 'B<digits>' label from a filename; fall back to stem."""
    m = re.search(r"b(\d+)", Path(path).name, re.IGNORECASE)
    return f"B{m.group(1)}" if m else Path(path).stem


def _coerce_int(val):
    """Convert a value to int where safe (handles str rack/elevation in new format)."""
    if val is None:
        return val
    try:
        return int(str(val).strip())
    except (ValueError, AttributeError):
        return val


# ─────────────────────────────────────────────────────────── Cutsheet ────────
def _is_pp(val):
    return bool(val and str(val).strip().upper().startswith("PP:"))


def build_cutsheet_lookup(path):
    """
    Returns dict: endpoint_string -> (pp_a, pp_b, pp_c, other_end, other_rack).

    Layout auto-detection per row:
      EC6  (col 5 is PP): c1 | c2(rack) | c3(PP_A) | c4(PP_B) | c5(PP_C) | c6 | c7(rack) | c8
      QFAB (col 3 is PP, col 5 not PP): c1 | c2(rack) | c3(PP_A) | c4(PP_B) | c5 | c6(rack)
      Direct (no PPs): c1 | c2(rack) | c3 | c4(rack)

    All sheets in the workbook are scanned (handles multi-tab EC6 cutsheets).
    Reverse mappings are always inserted.
    """
    wb = load_workbook(path, data_only=True)
    lookup = {}

    for ws in wb.worksheets:
        for r in ws.iter_rows(values_only=True):
            if not r:
                continue
            row = list(r) + [None] * max(0, 8 - len(r))
            c1, c2, c3, c4, c5, c6, c7 = (row[i] for i in range(7))

            if len(r) >= 7 and _is_pp(c5):
                # EC6: 3 patch panels
                if c1: lookup[str(c1).strip()] = (c3, c4, c5, c6, c7)
                if c6: lookup[str(c6).strip()] = (c5, c4, c3, c1, c2)

            elif len(r) >= 6 and _is_pp(c3) and not _is_pp(c5):
                # QFAB: 2 patch panels
                if c1: lookup[str(c1).strip()] = (c3, c4, None, c5, c6)
                if c5: lookup[str(c5).strip()] = (c4, c3, None, c1, c2)

            elif len(r) >= 4 and c1 and not _is_pp(c3):
                # Direct: no patch panels
                if c1: lookup[str(c1).strip()] = (None, None, None, c3, c4)
                if c3: lookup[str(c3).strip()] = (None, None, None, c1, c2)

    return lookup


# ───────────────────────────────────────────── Format detection & loading ────
# Names of the LLDP sheet in each report generation:
_OLD_LLDP = "full_path_lldp_with_int_down"
_NEW_LLDP = "lldp_with_int_down"


def _detect_lldp_format(wb):
    """Return ('old', sheet) or ('new', sheet) or (None, None)."""
    if _OLD_LLDP in wb.sheetnames:
        return "old", wb[_OLD_LLDP]
    if _NEW_LLDP in wb.sheetnames:
        return "new", wb[_NEW_LLDP]
    return None, None


def _load_lldp_sheet(ws, tag, fmt, lookup):
    """
    Load an LLDP sheet and return (headers, rows) normalised to LLDP_COLS shape.

    OLD format  — already has PP_A / PP_B; no index column.
    NEW format  — has an index column; lacks PP_A / PP_B.
                  PP_A / PP_B are injected from the cutsheet lookup.
                  Rows with no cutsheet match get PP_info_not_found so the
                  existing split_by_pp() logic continues to work correctly.
                  Rack and Elevation strings are coerced to int where possible.
    """
    raw_headers = [c.value for c in ws[1]]

    if fmt == "old":
        # ── straightforward: prefix Source, done ────────────────────────────
        norm_headers = ["Source"] + list(raw_headers)
        norm_rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or v == "" for v in r):
                continue
            norm_rows.append([tag] + list(r))
        return norm_headers, norm_rows

    # ── NEW format ───────────────────────────────────────────────────────────
    # Strip the leading 'index' column, then inject PP_A / PP_B after Elevation.
    drop_pos = raw_headers.index("index") if "index" in raw_headers else None
    base_headers = [h for i, h in enumerate(raw_headers) if i != drop_pos]

    # Build normalised header list
    elev_pos = base_headers.index("Elevation")
    norm_headers = (
        ["Source"]
        + base_headers[: elev_pos + 1]
        + ["PP_A", "PP_B"]
        + base_headers[elev_pos + 1 :]
    )

    # Pre-compute indices in base_headers (before PP injection)
    h_col = base_headers.index("Hostname")
    i_col = base_headers.index("Interface")
    rack_col = base_headers.index("Rack") if "Rack" in base_headers else None
    elev_col = base_headers.index("Elevation") if "Elevation" in base_headers else None

    norm_rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or v == "" for v in r):
            continue

        # Drop the index column
        base_row = [v for i, v in enumerate(r) if i != drop_pos]

        # Coerce Rack / Elevation strings to int
        if rack_col is not None:
            base_row[rack_col] = _coerce_int(base_row[rack_col])
        if elev_col is not None:
            base_row[elev_col] = _coerce_int(base_row[elev_col])

        # Lookup PP values from cutsheet
        hostname  = base_row[h_col]
        interface = base_row[i_col]
        key = f"{hostname} {interface}".strip() if hostname else ""
        hit = lookup.get(key) if key else None

        # hit[0] (pp_a) can be None when the cutsheet has a Direct (no-PP) entry
        # for this endpoint — e.g. a T1→T0 link present in the cutsheet without
        # patch-panel info.  A non-empty tuple is always truthy, so we must
        # explicitly check that pp_a is populated before using the hit values.
        if hit and hit[0] is not None:
            pp_a, pp_b = hit[0], hit[1]
        else:
            pp_a = pp_b = "PP_info_not_found"

        # Assemble normalised row: Source | ... | Elevation | PP_A | PP_B | Active Host | ...
        norm_row = (
            [tag]
            + base_row[: elev_pos + 1]
            + [pp_a, pp_b]
            + base_row[elev_pos + 1 :]
        )
        norm_rows.append(norm_row)

    return norm_headers, norm_rows


def load_combined(input_paths, lookup):
    """
    Stack data from all input files into a single dict keyed by logical sheet name.
    Both old-format (full_path_lldp_with_int_down) and new-format (lldp_with_int_down)
    files are accepted and can be freely mixed — they are normalised to the same
    column shape before being combined.
    """
    combined = {
        "full_path_lldp_with_int_down": {"headers": None, "rows": []},
        "optics":                        {"headers": None, "rows": []},
        "fec_ber":                       {"headers": None, "rows": []},
    }

    for path in input_paths:
        tag = extract_label(path)
        wb  = load_workbook(path, data_only=True)

        # ── LLDP main sheet ──────────────────────────────────────────────────
        fmt, ws = _detect_lldp_format(wb)
        if fmt is None:
            print(f"  [warn] {Path(path).name}: no LLDP sheet found — skipping.")
        else:
            print(f"  [info] {Path(path).name}: detected {fmt!r} format.")
            hdrs, rows = _load_lldp_sheet(ws, tag, fmt, lookup)
            key = "full_path_lldp_with_int_down"
            if combined[key]["headers"] is None:
                combined[key]["headers"] = hdrs
            combined[key]["rows"].extend(rows)

        # ── Optional sheets (optics / fec_ber) ──────────────────────────────
        for s in ("optics", "fec_ber"):
            if s not in wb.sheetnames:
                continue
            ws2    = wb[s]
            hdrs2  = [c.value for c in ws2[1]]
            if combined[s]["headers"] is None:
                combined[s]["headers"] = ["Source"] + list(hdrs2)
            for r in ws2.iter_rows(min_row=2, values_only=True):
                if all(v is None or v == "" for v in r):
                    continue
                combined[s]["rows"].append([tag] + list(r))

    return combined


# ────────────────────────────────────────────────────────────── Splits ───────
def split_full_path(headers, rows):
    """Separate Downlink (Act. Interface == 'interface down') from Mismatch."""
    ai = headers.index("Act. Interface")
    down, mis = [], []
    for r in rows:
        (down if r[ai] == "interface down" else mis).append(r)
    return down, mis


def split_by_pp(headers, rows):
    """
    T2-T1: PP_A is a real patch-panel string.
    T1-T0: PP_A is None or == 'PP_info_not_found' (no patch panel on that segment).
    Treating None and the sentinel string identically guards against cutsheet
    Direct entries (no PP columns) that return None rather than the sentinel.
    """
    a = headers.index("PP_A")
    t2, t1 = [], []
    for r in rows:
        pp_a = r[a]
        (t1 if (pp_a is None or pp_a == "PP_info_not_found") else t2).append(r)
    return t2, t1


# ─────────────────────────────────────────────────── Column manipulation ─────
def drop_columns(headers, rows, drop_names):
    keep       = [i for i, h in enumerate(headers) if h not in drop_names]
    new_headers = [headers[i] for i in keep]
    new_rows    = [[r[i] for i in keep] for r in rows]
    return new_headers, new_rows


def reorder_columns(headers, rows, new_order_names):
    idxs   = [headers.index(n) for n in new_order_names]
    extras = [i for i in range(len(headers)) if i not in idxs]
    final  = idxs + extras
    return [headers[i] for i in final], [[r[i] for i in final] for r in rows]


def swap_mismatch_groups(headers, rows):
    """Place Expected group before Active group in Mismatch tabs."""
    ACT = ["Active Host", "Act. Interface", "Act. Rack", "Act. Elevation"]
    EXP = ["Expected Hostname", "Exp. Interface", "Exp. Rack", "Exp. Elevation"]
    if not all(h in headers for h in ACT + EXP):
        return headers, rows
    act_idxs = [headers.index(h) for h in ACT]
    exp_idxs = [headers.index(h) for h in EXP]
    block    = set(act_idxs + exp_idxs)
    pre      = [i for i in range(len(headers)) if i not in block]
    final    = pre + exp_idxs + act_idxs
    return [headers[i] for i in final], [[r[i] for i in final] for r in rows]


# ──────────────────────────────────────────────────────────────── Dedup ──────
def dedup_bidirectional(headers, rows):
    """
    Remove rows whose Hostname/Interface <-> Expected Hostname/Exp. Interface
    already appeared in reverse order. Merges Source labels when deduplicating.
    """
    h_i   = headers.index("Hostname")
    i_i   = headers.index("Interface")
    eh_i  = headers.index("Expected Hostname")
    ei_i  = headers.index("Exp. Interface")
    src_i = headers.index("Source") if "Source" in headers else None
    seen, order = {}, []
    for r in rows:
        key = frozenset([(r[h_i], r[i_i]), (r[eh_i], r[ei_i])])
        if key in seen:
            if src_i is not None:
                existing = str(seen[key][src_i])
                parts    = [p.strip() for p in existing.split(",")]
                new_src  = str(r[src_i])
                if new_src not in parts:
                    seen[key][src_i] = f"{existing},{new_src}"
        else:
            seen[key] = list(r)
            order.append(key)
    return [seen[k] for k in order]


# ──────────────────────────────────────────────── Cutsheet enrichment ────────
def enrich(headers, rows, key_pairs, lookup):
    """
    Append CUT_COLS to headers and populate from the cutsheet lookup.
    key_pairs: [(host_col, intf_col), ...] tried in order until a hit is found.
    Lookup returns 5-tuples: (pp_a, pp_b, pp_c, other_end, other_rack).
    """
    new_headers = list(headers) + CUT_COLS
    new_rows    = []
    for r in rows:
        hit = None
        for host_col, intf_col in key_pairs:
            if host_col not in headers:
                continue
            h_idx = headers.index(host_col)
            i_idx = headers.index(intf_col)
            if r[h_idx] is None:
                continue
            key = f"{r[h_idx]} {r[i_idx]}".strip()
            h   = lookup.get(key)
            if h:
                hit = h
                break
        new_rows.append(list(r) + list(hit) if hit else list(r) + [None] * len(CUT_COLS))
    return new_headers, new_rows


def fill_empty_pp(headers, rows):
    """Replace empty Cutsheet PP_A / PP_B / PP_C with '<==>'."""
    pp_idxs = [headers.index(c)
               for c in ("Cutsheet PP_A", "Cutsheet PP_B", "Cutsheet PP_C")
               if c in headers]
    for r in rows:
        for idx in pp_idxs:
            if r[idx] in (None, ""):
                r[idx] = "<=>"
    return rows


# ──────────────────────────────────────────── Writing & styling ──────────────
def write_sheet(wb, name, headers, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    for r in rows:
        ws.append(r)
    return ws


def style_sheet(ws, pink_col_names=(), freeze_at=None):
    if ws.max_row == 0:
        return
    max_col = ws.max_column
    max_row = ws.max_row
    headers = [c.value for c in ws[1]]
    pink_idxs = {headers.index(n) + 1 for n in pink_col_names if n in headers}

    for cell in ws[1]:                    # header row
        cell.fill      = YELLOW_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = BORDER

    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.font      = BODY_FONT
            cell.alignment = CENTER
            cell.border    = BORDER
            if cell.column in pink_idxs:
                cell.fill = PINK_FILL

    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    for col_cells in ws.iter_cols(min_row=1, max_row=max_row, max_col=max_col):
        col_letter = get_column_letter(col_cells[0].column)
        max_len    = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = None

    ws.freeze_panes = freeze_at or "A2"


# ───────────────────────────────────────────── b -> B capitalisation ─────────
B_NUM = re.compile(r"\bb(\d+)\b")

def capitalize_b_numbers(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str):
                new_val, n = B_NUM.subn(lambda m: f"B{m.group(1)}", cell.value)
                if n:
                    cell.value = new_val


# ──────────────────────────────────────────────────────────── Summary ────────
def build_summary(wb, labels, cat_to_tabs):
    def count_sources(tabs):
        c = Counter()
        for tn in tabs:
            if tn not in wb.sheetnames:
                continue
            ws      = wb[tn]
            headers = [cell.value for cell in ws[1]]
            if "Source" not in headers:
                continue
            s_idx = headers.index("Source")
            for r in ws.iter_rows(min_row=2, values_only=True):
                src = r[s_idx]
                if src is None:
                    continue
                for part in str(src).split(","):
                    p = part.strip()
                    if p:
                        c[p] += 1
        return c

    if "summary" in wb.sheetnames:
        del wb["summary"]
    ws = wb.create_sheet("summary", 0)
    ws.append(["Category"] + list(labels) + ["Total"])
    for cat, tabs in cat_to_tabs:
        c         = count_sources(tabs)
        row_idx   = ws.max_row + 1
        last_col  = get_column_letter(1 + len(labels))
        ws.append([cat] + [c.get(l, 0) for l in labels]
                  + [f"=SUM(B{row_idx}:{last_col}{row_idx})"])
    return ws


# ─────────────────────────────────────────────────────── Main pipeline ───────
def main():
    cutsheet_path, input_paths = select_files()
    print(f"Cutsheet : {cutsheet_path}")
    print(f"Inputs   : {len(input_paths)} file(s)")

    labels = [extract_label(p) for p in input_paths]
    print(f"Labels   : {labels}")

    lookup = build_cutsheet_lookup(cutsheet_path)
    print(f"Cutsheet : {len(lookup)} unique endpoints loaded")

    # load_combined now needs the lookup so it can inject PP_A/PP_B into new-format rows
    combined = load_combined(input_paths, lookup)

    # ── LLDP: 4 sub-tabs ────────────────────────────────────────────────────
    fp = combined["full_path_lldp_with_int_down"]
    if fp["headers"] is None:
        sys.exit("ERROR: no LLDP sheet found in any input file.")

    fp_headers = fp["headers"]
    fp_rows    = fp["rows"]

    dl_rows,  mis_rows  = split_full_path(fp_headers, fp_rows)
    t2_dl_rows,  t1_dl_rows  = split_by_pp(fp_headers, dl_rows)
    t2_mis_rows, t1_mis_rows = split_by_pp(fp_headers, mis_rows)

    t2_dl_hdr,  t2_dl_rows  = drop_columns(fp_headers, t2_dl_rows,  DROP_SPECS["T2-T1 Downlink"])
    t1_dl_hdr,  t1_dl_rows  = drop_columns(fp_headers, t1_dl_rows,  DROP_SPECS["T1-T0 Downlink"])
    t2_mis_hdr, t2_mis_rows = drop_columns(fp_headers, t2_mis_rows, DROP_SPECS["T2-T1 Mismatch"])
    t1_mis_hdr, t1_mis_rows = drop_columns(fp_headers, t1_mis_rows, DROP_SPECS["T1-T0 Mismatch"])

    t2_mis_hdr, t2_mis_rows = swap_mismatch_groups(t2_mis_hdr, t2_mis_rows)
    t1_mis_hdr, t1_mis_rows = swap_mismatch_groups(t1_mis_hdr, t1_mis_rows)

    t1_dl_rows  = dedup_bidirectional(t1_dl_hdr,  t1_dl_rows)
    t1_mis_rows = dedup_bidirectional(t1_mis_hdr, t1_mis_rows)

    # ── optics / fec_ber (optional) ─────────────────────────────────────────
    def process_optional(sheet_key, drop_spec):
        data = combined[sheet_key]
        if data["headers"] is None:
            print(f"  [info] '{sheet_key}' not present in any input — tab will be empty.")
            return [], []
        return drop_columns(data["headers"], data["rows"], drop_spec)

    op_hdr, op_rows = process_optional("optics",  DROP_SPECS["optics"])
    if op_hdr:
        ORDER = ["Source", "Input Power", "Output Power", "Hostname", "Interface"]
        op_hdr, op_rows = reorder_columns(op_hdr, op_rows,
                                          [c for c in ORDER if c in op_hdr])

    fb_hdr, fb_rows = process_optional("fec_ber", DROP_SPECS["fec_ber"])

    # ── Cutsheet enrichment ──────────────────────────────────────────────────
    # Mismatch tabs and optional sheets get the full cutsheet columns appended.
    # (Downlink tabs already have PP_A/PP_B from the normalised LLDP load.)
    KEY_PAIRS_MIS = [("Active Host", "Act. Interface"), ("Hostname", "Interface")]
    KEY_PAIRS_OPT = [("Hostname", "Interface")]

    t2_mis_hdr, t2_mis_rows = enrich(t2_mis_hdr, t2_mis_rows, KEY_PAIRS_MIS, lookup)
    t1_mis_hdr, t1_mis_rows = enrich(t1_mis_hdr, t1_mis_rows, KEY_PAIRS_MIS, lookup)
    if op_hdr:
        op_hdr, op_rows = enrich(op_hdr, op_rows, KEY_PAIRS_OPT, lookup)
    if fb_hdr:
        fb_hdr, fb_rows = enrich(fb_hdr, fb_rows, KEY_PAIRS_OPT, lookup)

    for hdr, rows in [(t2_mis_hdr, t2_mis_rows), (t1_mis_hdr, t1_mis_rows),
                      (op_hdr, op_rows), (fb_hdr, fb_rows)]:
        if hdr:
            fill_empty_pp(hdr, rows)

    # ── Assemble workbook ────────────────────────────────────────────────────
    sheets_to_write = [
        ("T2-T1 Downlink", t2_dl_hdr  + ["Note"], [r + [None] for r in t2_dl_rows]),
        ("T1-T0 Downlink", t1_dl_hdr  + ["Note"], [r + [None] for r in t1_dl_rows]),
        ("T2-T1 Mismatch", t2_mis_hdr + ["Note"], [r + [None] for r in t2_mis_rows]),
        ("T1-T0 Mismatch", t1_mis_hdr + ["Note"], [r + [None] for r in t1_mis_rows]),
        ("optics",  (op_hdr or []) + ["Note"], [r + [None] for r in op_rows]),
        ("fec_ber", (fb_hdr or []) + ["Note"], [r + [None] for r in fb_rows]),
    ]

    wb = Workbook()
    wb.remove(wb.active)
    for name, hdr, rows in sheets_to_write:
        write_sheet(wb, name, hdr, rows)

    build_summary(wb, labels, [
        ("Downlink", ["T2-T1 Downlink", "T1-T0 Downlink"]),
        ("Mismatch", ["T2-T1 Mismatch", "T1-T0 Mismatch"]),
        ("optics",   ["optics"]),
        ("fec_ber",  ["fec_ber"]),
    ])

    canonical = ["summary", "T2-T1 Downlink", "T1-T0 Downlink",
                 "T2-T1 Mismatch", "T1-T0 Mismatch", "optics", "fec_ber"]
    wb._sheets = [wb[n] for n in canonical if n in wb.sheetnames]

    for name in wb.sheetnames:
        capitalize_b_numbers(wb[name])

    for name in wb.sheetnames:
        pink    = PINK_COLS if name.endswith("Mismatch") else ()
        freeze  = "D2" if name == "optics" else "A2"
        style_sheet(wb[name], pink_col_names=pink, freeze_at=freeze)

    out_name = "-".join(labels) + ".xlsx"
    out_path = Path(input_paths[0]).parent / out_name
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
