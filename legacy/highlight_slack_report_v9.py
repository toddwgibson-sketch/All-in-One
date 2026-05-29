"""
Slack Validation Report Auto-Highlighter v9
============================================
v9 fix: Old-format reports (with_pp) can still have PP_info_not_found in
    every connection column when the report generator could not resolve the
    patch-panel path. Previously the script would pass those sentinel strings
    straight through to the output. Now, whenever a cell contains a value
    starting with 'PP_info', the script falls back to the cutsheet lookup
    keyed by (Hostname, Interface) — or, if that is absent, the partner lane
    — before writing the value.  This applies to all three output surfaces:
    read_lldp_rows (LLDP tab), the optics loop, and the FEC loop.

v8 fix: the cutsheet loader was using hardcoded positional column indices
    that assumed a legacy "combined Device A/B" layout. The actual cutsheet
    has separate named columns (Hostname, Interface, L/R, Source_port,
    Z Hostname, Z Interface, Z L/R, Z Rack, Z Elevation, etc.). The loader
    now reads the header row first and maps all data by column name, making
    it work with both the current and any future cutsheet layouts.

Handles both report formats automatically:

OLD format (with_pp):  lldp_sp already contains Source_port, DMARC1, DMARC2,
    Destination_port, Z Hostname, Z Interface, Z Rack, Z Elevation.
    If any of those cells hold PP_info_not_found the cutsheet is used as
    a fallback (v9 behaviour).

NEW format (no_pp):    lldp_sp only has Hostname, Interface, Rack, Elevation,
    Active/Expected columns. Source_port, DMARC1, DMARC2, Destination_port,
    Z Interface, Z Rack, Z Elevation are pulled from the cutsheet.
"""

import sys, os, re, copy, json, time
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    HAS_TK = True
except ImportError:
    HAS_TK = False

CONFIG_FILE = Path.home() / ".highlight_slack_config.json"

def load_config():
    if CONFIG_FILE.exists():
        try: return json.loads(CONFIG_FILE.read_text())
        except: pass
    return {}

def save_config(cfg):
    CONFIG_FILE.write_text(json.dumps(cfg, indent=2))

def pick_file(title, filetypes=None):
    if not HAS_TK:
        path = input(f"{title}\nEnter file path: ").strip().strip('"').strip("'")
        return path if os.path.isfile(path) else None
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    path = filedialog.askopenfilename(
        title=title,
        filetypes=filetypes or [("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    root.destroy()
    return path or None

def pick_multiple_files(title, filetypes=None):
    """Pick one or more files — returns list of paths"""
    if not HAS_TK:
        print(f"{title}")
        print("Enter file paths one per line, blank line when done:")
        paths = []
        while True:
            p = input("  Path: ").strip().strip('"').strip("'")
            if not p: break
            if os.path.isfile(p): paths.append(p)
            else: print(f"  Not found: {p}")
        return paths
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    paths = filedialog.askopenfilenames(
        title=title,
        filetypes=filetypes or [("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    root.destroy()
    return list(paths) if paths else []

def show_msg(title, msg, error=False):
    print(f"{'ERROR: ' if error else ''}{msg}")
    if HAS_TK:
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        (messagebox.showerror if error else messagebox.showinfo)(title, msg)
        root.destroy()

# ── Colours matching example file exactly ─────────────────────────────────────
WHITE      = "FFFFFF"    # white
YELLOW     = "FFFF00"    # physical row highlight
LOG_BG     = "FFFFFF"    # logical row - white
GREEN_DONE = "92D050"    # completed rows
SRC_BG     = "FCE4D6"    # source port - pink
D1_BG      = "FFF2CC"    # DMARC1 - yellow
D2_BG      = "E2F0D9"    # DMARC2 - green
DEST_BG    = "D9EAF7"    # destination - blue
Z_BG       = "DDEBF7"    # Z device - light blue
ACT_BG     = "FFC7CE"    # active - red
EXP_BG     = "C6EFCE"    # expected - green
LR_BG      = "FFFFFF"    # L&R col on physical rows - white
LR_LOG     = "FFFFFF"    # L&R col on logical rows - white
HDR_BG     = "1F4E79"    # header - navy
HDR_FG     = "FFFFFF"
PP_BG      = "FCE4D6"    # possible patch panel - pink
PD_BG      = "FFF2CC"    # possible DMARC - yellow

TAB_ALL    = "1F4E79"
TAB_MISS   = "C00000"
TAB_DOWN   = "ED7D31"
TAB_OPT    = "833C00"
TAB_FEC    = "7030A0"

def fill(h):   return PatternFill("solid", fgColor=h)
def no_fill(): return PatternFill(fill_type=None)
def font(color="000000", bold=False, sz=9):
    return Font(bold=bold, color=color, name="Arial", size=sz)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=False)
def vcenter(): return Alignment(vertical="center", wrap_text=False)

# ── Cutsheet lookup ───────────────────────────────────────────────────────────
def _load_single_cutsheet(path, t0, t1, t1_rev):
    """
    Load one cutsheet into the shared lookup dicts.

    Supports two cutsheet layouts — detected automatically from the header row:

    NEW layout (separate columns, 14 cols):
        Hostname | Interface | L/R | Rack | Elevation |
        Source_port | DMARC1 | DMARC2 | Destination_port |
        Z Hostname | Z Interface | Z L/R | Z Rack | Z Elevation

    LEGACY layout (combined Device A/B, 11+ cols):
        Label | DeviceA (host iface) | RackA |
        Source_port | DMARC1 | DMARC2 | Destination_port |
        DeviceB (host iface) | RackB | Z Elevation | T1 Label
    """
    wb = load_workbook(path, read_only=True)
    sheet = next((wb[n] for n in wb.sheetnames
                  if 'installation' in n.lower()), wb[wb.sheetnames[0]])

    # Read header row to detect layout
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = {str(h or '').strip(): i for i, h in enumerate(header_row)}

    # Detect new layout by presence of separate 'Hostname' + 'Interface' columns
    new_layout = ('Hostname' in headers and 'Interface' in headers)

    def col(name, fallback=None):
        """Get value from a row tuple by column name, with index fallback."""
        idx = headers.get(name)
        return idx if idx is not None else fallback

    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row): continue

        def v(name, legacy_idx=None):
            """Read cell value by column name or legacy index."""
            idx = headers.get(name)
            if idx is not None and idx < len(row):
                return str(row[idx] or '').strip()
            if legacy_idx is not None and legacy_idx < len(row):
                return str(row[legacy_idx] or '').strip()
            return ''

        if new_layout:
            hostname    = v('Hostname')
            interface   = v('Interface')
            lbl         = v('L/R')
            rack_a      = v('Rack')
            src         = v('Source_port')
            dmarc1      = v('DMARC1')
            dmarc2      = v('DMARC2')
            dest_p      = v('Destination_port')
            z_hostname  = v('Z Hostname')
            z_interface = v('Z Interface')
            t1l         = v('Z L/R')
            rack_b      = v('Z Rack')
            z_elev      = v('Z Elevation')
            device_a    = f"{hostname} {interface}".strip()
        else:
            # Legacy: combined Device A / Device B columns
            lbl         = v('Label',    0)
            device_a    = v('DeviceA',  1)
            rack_a      = v('RackA',    2)
            src         = v('Source_port', 3)
            dmarc1      = v('DMARC1',   4)
            dmarc2      = v('DMARC2',   5)
            dest_p      = v('Destination_port', 6)
            device_b    = v('DeviceB',  7)
            rack_b      = v('RackB',    8)
            z_elev      = v('Z Elevation', 9)
            t1l         = v('T1 Label', 10)
            dev_a_parts = device_a.split()
            hostname    = dev_a_parts[0] if dev_a_parts else ''
            interface   = dev_a_parts[1] if len(dev_a_parts) > 1 else ''
            dev_b_parts = device_b.split() if device_b else []
            z_hostname  = dev_b_parts[0] if dev_b_parts else ''
            z_interface = dev_b_parts[1] if len(dev_b_parts) > 1 else ''

        if not hostname or not interface: continue

        # T0 label lookup
        if lbl and re.match(r'\d+[LR]$', lbl):
            k = (hostname, interface)
            t0[k] = lbl
            t1[k] = t1l

        # T1 reverse lookup (for Possible columns on mismatch rows)
        if z_hostname and z_interface:
            t1_rev[(z_hostname, z_interface)] = {
                'device_a':    device_a,
                't0_lbl':      lbl,
                'rack_a':      rack_a,
                'source_port': src,
                'dmarc1':      dmarc1,
                'dmarc2':      dmarc2,
                'dest_port':   dest_p,
                'rack_b':      rack_b,
                't1_lbl':      t1l,
            }
            count += 1

    wb.close()
    return count


_cutsheet_pp = {}

def build_lookup(paths):
    """
    Build all lookup dicts from one or more cutsheet files.
    Returns (phys_t0, phys_t1, t1_rev).
    Also populates the global _cutsheet_pp keyed by (hostname, interface).
    """
    global _cutsheet_pp
    _cutsheet_pp = {}
    if isinstance(paths, str):
        paths = [paths]
    t0, t1, t1_rev = {}, {}, {}
    for path in paths:
        count = _load_single_cutsheet(path, t0, t1, t1_rev)

        # Build the per-interface PP lookup (used to fill missing report columns)
        wb2 = load_workbook(path, read_only=True)
        sheet2 = next((wb2[n] for n in wb2.sheetnames
                       if 'installation' in n.lower()), wb2[wb2.sheetnames[0]])

        header_row2 = next(sheet2.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers2 = {str(h or '').strip(): i for i, h in enumerate(header_row2)}
        new_layout2 = ('Hostname' in headers2 and 'Interface' in headers2)

        for row in sheet2.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row): continue

            def g(name, legacy_idx=None):
                idx = headers2.get(name)
                if idx is not None and idx < len(row):
                    return str(row[idx] or '').strip()
                if legacy_idx is not None and legacy_idx < len(row):
                    return str(row[legacy_idx] or '').strip()
                return ''

            if new_layout2:
                hostname    = g('Hostname')
                interface   = g('Interface')
                source_port = g('Source_port')
                dmarc1      = g('DMARC1')
                dmarc2      = g('DMARC2')
                dest_port   = g('Destination_port')
                z_hostname  = g('Z Hostname')
                z_interface = g('Z Interface')
                z_rack      = g('Z Rack')
                z_elevation = g('Z Elevation')
            else:
                dev_a       = g('DeviceA', 1)
                parts       = dev_a.split() if dev_a else []
                hostname    = parts[0] if parts else ''
                interface   = parts[1] if len(parts) > 1 else ''
                source_port = g('Source_port', 3)
                dmarc1      = g('DMARC1', 4)
                dmarc2      = g('DMARC2', 5)
                dest_port   = g('Destination_port', 6)
                dev_b       = g('DeviceB', 7)
                bparts      = dev_b.split() if dev_b else []
                z_hostname  = bparts[0] if bparts else ''
                z_interface = bparts[1] if len(bparts) > 1 else ''
                z_rack      = g('RackB', 8)
                z_elevation = g('Z Elevation', 9)

            if hostname and interface:
                _cutsheet_pp[(hostname, interface)] = {
                    'source_port': source_port,
                    'dmarc1':      dmarc1,
                    'dmarc2':      dmarc2,
                    'dest_port':   dest_port,
                    'z_hostname':  z_hostname,
                    'z_interface': z_interface,
                    'z_rack':      z_rack,
                    'z_elevation': z_elevation,
                }
        wb2.close()
        print(f"    Loaded: {os.path.basename(path)} "
              f"({len(t0)} T0 labels | {count} T1 reverse | {len(_cutsheet_pp)} PP entries)")
    return t0, t1, t1_rev

def get_prev_issues(report_path):
    """Extract all issues from a previous report for recurring detection"""
    try:
        wb = load_workbook(report_path, read_only=True)
    except Exception as e:
        print(f"  Warning: could not load previous report: {e}")
        return set(), set(), set(), {}, {}

    # LLDP issues
    ws = next((wb[n] for n in wb.sheetnames if 'lldp' in n.lower()), None)
    prev_miss = set(); prev_down = set(); prev_rack_map = {}
    if ws:
        hc = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Hostname'), None)
        ic = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Interface'), None)
        ac = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip() in ('Act. Interface','Act.Interface')), None)
        rc = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Rack'), None)
        if hc and ic and ac:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row: continue
                h=str(row[hc-1] or '').strip(); i=str(row[ic-1] or '').strip()
                ai=str(row[ac-1] or '').strip().lower()
                rack=str(row[rc-1] or '').strip() if rc else 'Unknown'
                if h and i:
                    if ai == 'interface down': prev_down.add((h,i))
                    elif ai.startswith('swp'):  prev_miss.add((h,i))
                    prev_rack_map[(h,i)] = rack or 'Unknown' 

    # Optics issues
    ws_opt = next((wb[n] for n in wb.sheetnames if 'optic' in n.lower()), None)
    prev_opt = set(); prev_opt_rack_map = {}
    if ws_opt:
        hc = next((c for c in range(1,ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip()=='Hostname'), None)
        ic = next((c for c in range(1,ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip()=='Interface'), None)
        rc = next((c for c in range(1,ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip()=='Rack'), None)
        if hc and ic:
            for row in ws_opt.iter_rows(min_row=2, values_only=True):
                if not row: continue
                h=str(row[hc-1] or '').strip(); i=str(row[ic-1] or '').strip()
                rack=str(row[rc-1] or '').strip() if rc else 'Unknown'
                if h and i:
                    prev_opt.add((h,i))
                    prev_opt_rack_map[(h,i)] = rack or 'Unknown'

    wb.close()
    print(f"  Previous report: {len(prev_miss)} mismatches, {len(prev_down)} downlinks, {len(prev_opt)} optics")
    return prev_miss, prev_down, prev_opt, prev_rack_map, prev_opt_rack_map

def get_history_flag(host, iface, current_type, prev_miss, prev_down, prev_opt):
    """Return (flag_text, flag_colour) for a row based on previous report"""
    key = (host, iface)
    if current_type == 'mismatch':
        if key in prev_miss: return "🔁 Recurring mismatch",  "FF6B6B"
        if key in prev_down: return "⬆️ Was downlink",        "FFB347"
    elif current_type == 'downlink':
        if key in prev_down: return "🔁 Recurring downlink",  "FF6B6B"
        if key in prev_opt:  return "⚡ Was optic error",      "D35400"   # orange-red — likely bad reseat
        if key in prev_miss: return "⬇️ Was mismatch",        "FFB347"
    elif current_type == 'optic':
        if key in prev_opt:  return "🔁 Recurring optic",     "FF6B6B"
        if key in prev_down: return "⬆️ Was downlink",        "FFB347"
        if key in prev_miss: return "⬇️ Was mismatch",        "FFB347"
    return "", ""

def get_labels(hostname, iface, phys_t0, phys_t1):
    key = (hostname, iface)
    if key in phys_t0:
        return phys_t0[key], phys_t1[key], True
    m = re.match(r'(swp\d+)s(\d+)', str(iface))
    if m:
        base, lane = m.group(1), int(m.group(2))
        partner_lane = {0:1, 1:0, 2:3, 3:2}.get(lane)
        if partner_lane is not None:
            p = (hostname, f"{base}s{partner_lane}")
            if p in phys_t0:
                return phys_t0[p], phys_t1[p], False
    return '', '', False

def row_type(act_iface):
    v = str(act_iface or '').strip().lower()
    if v == 'interface down': return 'downlink'
    if v.startswith('swp'):   return 'mismatch'
    return 'other'

def find_col(ws, *names):
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or '').strip() in names:
            return c
    return None

# ── Build output sheet ────────────────────────────────────────────────────────
def build_lldp_sheet(wb_out, sheet_name, rows, tab_colour, is_mismatch=False,
                     prev_miss=None, prev_down=None, prev_opt=None, is_downlinks=False):
    prev_miss = prev_miss or set()
    prev_down = prev_down or set()
    prev_opt  = prev_opt  or set()
    ws = wb_out.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_colour

    # Output column layout (matching example exactly):
    # 1=Hostname, 2=Interface, 3=L&R(T0), 4=Rack, 5=Elevation
    # 6=Source_port, 7=DMARC1, 8=DMARC2, 9=Destination_port
    # 10=Z Hostname, 11=Z Interface, 12=L&R(T1), 13=Z Rack, 14=Z Elevation
    # 15=Possible DMARC, 16=Possible patch panel  (mismatch cols)
    # 17=Active Host, 18=Act. Interface, 19=Act. Rack, 20=Act. Elevation
    # 21=Expected Hostname, 22=Exp. Interface, 23=Exp. Rack, 24=Exp. Elevation

    base_headers = [
        ("Interface",            HDR_BG),
        ("L&R",                  HDR_BG),
        ("Rack",                 HDR_BG),
        ("Elevation",            HDR_BG),
        ("Source_port",          "C0504D"),
        ("DMARC1",               "7F6000"),
        ("DMARC2",               "375623"),
        ("Destination_port",     "17375E"),
        ("Z Interface",          "17375E"),
        ("L&R",                  "17375E"),
        ("Z Rack",               "17375E"),
        ("Z Elevation",          "17375E"),
    ]
    possible_headers = [
        ("Possible Device A",    "833C00"),
        ("Possible Rack / U",    "833C00"),
        ("Possible Source Port", "833C00"),
        ("Possible DMARC1",      "7F6000"),
        ("Possible DMARC2",      "C0504D"),
        ("Possible Dest Port",   "375623"),
        ("Possible T1 Rack / U", "375623"),
        ("Possible T1 Port",     "375623"),
    ]
    tail_headers = [] if is_downlinks else [
        ("Act. Interface",       "9C0006"),
        ("Act. Rack",            "9C0006"),
        ("Act. Elevation",       "9C0006"),
        ("Exp. Interface",       "375623"),
        ("Exp. Rack",            "375623"),
        ("Exp. Elevation",       "375623"),
    ]
    tail_headers += [("History", "595959")]
    headers = base_headers + (possible_headers if is_mismatch else []) + tail_headers

    for col, (label, bg) in enumerate(headers, start=1):
        c = ws.cell(1, col)
        c.value     = label
        c.fill      = fill(bg)
        c.font      = font(HDR_FG, bold=True, sz=9)
        c.alignment = center()

    ws.row_dimensions[1].height = 20

    # Column widths - possible cols only on mismatch tab
    base_widths     = [12,6,7,6,30,28,28,30,12,6,7,6]
    possible_widths = [8,14,30,28,28,30,14,8]
    tail_widths     = ([] if is_downlinks else [12,7,6,12,7,6]) + [22]
    widths = base_widths + (possible_widths if is_mismatch else []) + tail_widths
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for out_row, rd in enumerate(rows, start=2):
        p  = rd['is_phys']
        mi = rd.get('mismatch_info', {})

        # Row background
        row_bg = YELLOW if p else LOG_BG

        # Values from the named-key cells dict
        cells = rd['cells']
        # Original: Hostname, Interface, Rack, Elevation, Source_port,
        #           DMARC1, DMARC2, Destination_port,
        #           Z Hostname, Z Interface, Z Rack, Z Elevation,
        #           Active Host, Act. Interface, Act. Rack, Act. Elevation,
        #           Expected Hostname, Exp. Interface, Exp. Rack, Exp. Elevation

        na = '#N/A' if not p else ''
        base_values = [
            cells['Interface']['value'],        # Interface
            rd['t0'],                           # L&R T0
            cells['Rack']['value'],             # Rack
            cells['Elevation']['value'],        # Elevation
            cells['Source_port']['value'],      # Source_port
            cells['DMARC1']['value'],           # DMARC1
            cells['DMARC2']['value'],           # DMARC2
            cells['Destination_port']['value'], # Destination_port
            cells['Z Interface']['value'],      # Z Interface
            rd['t1'],                           # L&R T1
            cells['Z Rack']['value'],           # Z Rack
            cells['Z Elevation']['value'],      # Z Elevation
        ]
        # mi is populated on both physical and logical rows (second pass copies from partner)
        possible_values = [
            mi.get('t0_lbl',     '') or na,  # Possible Device A
            mi.get('rack_a',     '') or na,  # Possible Rack/U
            mi.get('source_port','') or na,  # Possible Source Port
            mi.get('dmarc1',     '') or na,  # Possible DMARC1
            mi.get('dmarc2',     '') or na,  # Possible DMARC2
            mi.get('dest_port',  '') or na,  # Possible Dest Port
            mi.get('rack_b',     '') or na,  # Possible T1 Rack/U
            mi.get('t1_lbl',     '') or na,  # Possible T1 Port
        ] if is_mismatch else []
        # Get history flag
        host_val  = cells['Hostname']['value']
        iface_val = cells['Interface']['value']
        hist_flag, hist_col = get_history_flag(
            str(host_val or '').strip(), str(iface_val or '').strip(),
            rd['row_type'], prev_miss, prev_down, prev_opt
        )
        tail_values = ([] if is_downlinks else [
            cells['Act. Interface']['value'],  # Act. Interface
            cells['Act. Rack']['value'],       # Act. Rack
            cells['Act. Elevation']['value'],  # Act. Elevation
            cells['Exp. Interface']['value'],  # Exp. Interface
            cells['Exp. Rack']['value'],       # Exp. Rack
            cells['Exp. Elevation']['value'],  # Exp. Elevation
        ]) + [hist_flag]                       # History
        all_values = base_values + possible_values + tail_values
        values = {i+1: v for i, v in enumerate(all_values)}

        # Physical rows: yellow across all base+tail cols, possible cols keep own colour
        # Logical rows: LOG_BG across all base+tail cols
        if p:
            base_fills = [
                "FFFFFF", LR_BG, "FFFFFF", "FFFFFF",
                "FFFFFF", "FFFFFF", "FFFFFF", "FFFFFF",
                "FFFFFF", LR_BG, "FFFFFF", "FFFFFF",
            ]
            tail_fills = ["FFFFFF"]*8
        else:
            base_fills = [
                LOG_BG, LR_LOG, LOG_BG, LOG_BG,
                LOG_BG, LOG_BG, LOG_BG, LOG_BG,
                LOG_BG, LR_LOG, LOG_BG, LOG_BG,
            ]
            tail_fills = [LOG_BG]*8
        possible_fills = ["FDDCB5","FDDCB5","FDDCB5", D1_BG, PP_BG,
                           "D5F5E3","D5F5E3","D5F5E3"] if is_mismatch else []
        col_fills = {i+1: f for i, f in enumerate(base_fills + possible_fills + tail_fills)}
        for _c in range(1, len(all_values)+1):
            if _c not in col_fills: col_fills[_c] = row_bg

        for col in range(1, len(all_values) + 1):
            c = ws.cell(out_row, col)
            c.value     = values.get(col, '')
            c.fill      = fill(col_fills[col])
            c.font      = font(sz=8)
            c.alignment = vcenter()

        ws.row_dimensions[out_row].height = 15

    ws.freeze_panes = "A2"

    # ── Draw borders around each physical+logical pair ────────────────────────
    from openpyxl.styles import Border, Side
    thin  = Side(style="thin",   color="AAAAAA")
    thick = Side(style="medium", color="555555")

    total_data_cols = len(all_values) if rows else 0
    if total_data_cols == 0 and ws.max_column > 1:
        total_data_cols = ws.max_column

    # Group consecutive rows by their L&R value (col 3 = L&R after T0 port col)
    lr_col = 2  # L&R is col 2 in output (col1=Interface, col2=L&R)
    data_row = 2
    max_r = ws.max_row

    while data_row <= max_r:
        lr_val = ws.cell(data_row, lr_col).value
        # Find how many consecutive rows share the same L&R value
        group_end = data_row
        while group_end + 1 <= max_r and ws.cell(group_end+1, lr_col).value == lr_val and lr_val:
            group_end += 1

        # Apply border box around this group
        for row in range(data_row, group_end + 1):
            is_top    = (row == data_row)
            is_bottom = (row == group_end)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                cell.border = Border(
                    top    = thick if is_top    else thin,
                    bottom = thick if is_bottom else Side(style=None),
                    left   = thick if col == 1  else thin,
                    right  = thick if col == ws.max_column else thin,
                )

        data_row = group_end + 1

# ── Process source sheet ──────────────────────────────────────────────────────
def read_lldp_rows(ws_src, phys_t0, phys_t1, t1_rev):
    # ── Resolve column positions by name ────────────────────────────────────────
    host_col     = find_col(ws_src, 'Hostname')
    iface_col    = find_col(ws_src, 'Interface')
    rack_col     = find_col(ws_src, 'Rack')
    elev_col     = find_col(ws_src, 'Elevation')
    src_col      = find_col(ws_src, 'Source_port')
    d1_col       = find_col(ws_src, 'DMARC1')
    d2_col       = find_col(ws_src, 'DMARC2')
    dest_col     = find_col(ws_src, 'Destination_port')
    z_host_col   = find_col(ws_src, 'Z Hostname')
    z_iface_col  = find_col(ws_src, 'Z Interface')
    z_rack_col   = find_col(ws_src, 'Z Rack')
    z_elev_col   = find_col(ws_src, 'Z Elevation')
    act_h_col    = find_col(ws_src, 'Active Host')
    act_if_col   = find_col(ws_src, 'Act. Interface', 'Act.Interface')
    act_rack_col = find_col(ws_src, 'Act. Rack')
    act_elev_col = find_col(ws_src, 'Act. Elevation')
    exp_h_col    = find_col(ws_src, 'Expected Hostname')
    exp_if_col   = find_col(ws_src, 'Exp. Interface')
    exp_rack_col = find_col(ws_src, 'Exp. Rack')
    exp_elev_col = find_col(ws_src, 'Exp. Elevation')

    # Detect format: new format has no Source_port / Z Interface columns
    new_format = (src_col is None)
    if new_format:
        print("  Detected new report format — PP and Z data will be sourced from cutsheet.")

    def cell_val(row, col):
        return ws_src.cell(row, col).value if col else None

    def cell_item(row, col):
        if not col:
            return {'value': '', 'fill': no_fill()}
        return {'value': ws_src.cell(row, col).value,
                'fill':  copy.copy(ws_src.cell(row, col).fill)}

    def blank():
        return {'value': '', 'fill': no_fill()}

    def cs_item(val):
        return {'value': val or '', 'fill': no_fill()}

    # First pass — index row numbers by (host, iface) for partner-copy logic
    raw_rows = {}
    for _r in range(2, ws_src.max_row + 1):
        _h = str(cell_val(_r, host_col) or '').strip()
        _i = str(cell_val(_r, iface_col) or '').strip()
        if _h and _i: raw_rows[(_h, _i)] = _r

    rows = []
    for row in range(2, ws_src.max_row + 1):
        host  = str(cell_val(row, host_col)  or '').strip()
        iface = str(cell_val(row, iface_col) or '').strip()
        if not host or not iface: continue

        t0, t1, is_p = get_labels(host, iface, phys_t0, phys_t1)
        act_if = cell_val(row, act_if_col) if act_if_col else None
        rtype  = row_type(act_if)

        # ── Mismatch reverse-lookup (physical rows only) ─────────────────────
        mi = {}
        if is_p and act_h_col and act_if_col:
            ah = str(cell_val(row, act_h_col) or '').strip()
            ai = str(cell_val(row, act_if_col) or '').strip()
            if ai.lower().startswith('swp'):
                mi = t1_rev.get((ah, ai), {})
                if not mi:
                    m2 = re.match(r'(swp\d+)s(\d+)', ai)
                    if m2:
                        base2, lane2 = m2.group(1), int(m2.group(2))
                        partner = {0:1,1:0,2:3,3:2}.get(lane2)
                        if partner is not None:
                            mi = t1_rev.get((ah, f"{base2}s{partner}"), {})

        # ── Build named cells dict ───────────────────────────────────────────
        if new_format:
            # NEW FORMAT: Source_port/DMARC/Z cols not in report — pull from cutsheet
            cs = _cutsheet_pp.get((host, iface), {})

            # If not found directly, try the partner lane (shares same PP connector)
            if not cs:
                m_p = re.match(r'(swp\d+)s(\d+)', iface)
                if m_p:
                    base_p = m_p.group(1)
                    partner_lane = {0:1,1:0,2:3,3:2}.get(int(m_p.group(2)))
                    if partner_lane is not None:
                        cs = _cutsheet_pp.get((host, f"{base_p}s{partner_lane}"), {})

            # Z info: always prefer Expected columns from report — they are per-interface.
            # Cutsheet partner data reflects the PARTNER lane's Z connection (wrong elevation).
            z_iface_val = str(cell_val(row, exp_if_col)   or '').strip() or cs.get('z_interface', '')
            z_rack_val  = str(cell_val(row, exp_rack_col) or '').strip() or cs.get('z_rack',      '')
            z_elev_val  = str(cell_val(row, exp_elev_col) or '').strip() or cs.get('z_elevation', '')
            z_host_val  = str(cell_val(row, exp_h_col)    or '').strip() or cs.get('z_hostname',  '')

            cells = {
                'Hostname':         cell_item(row, host_col),
                'Interface':        cell_item(row, iface_col),
                'Rack':             cell_item(row, rack_col),
                'Elevation':        cell_item(row, elev_col),
                'Source_port':      cs_item(cs.get('source_port', '')),
                'DMARC1':           cs_item(cs.get('dmarc1',      '')),
                'DMARC2':           cs_item(cs.get('dmarc2',      '')),
                'Destination_port': cs_item(cs.get('dest_port',   '')),
                'Z Hostname':       cs_item(z_host_val),
                'Z Interface':      cs_item(z_iface_val),
                'Z Rack':           cs_item(z_rack_val),
                'Z Elevation':      cs_item(z_elev_val),
                'Active Host':      cell_item(row, act_h_col),
                'Act. Interface':   cell_item(row, act_if_col),
                'Act. Rack':        cell_item(row, act_rack_col),
                'Act. Elevation':   cell_item(row, act_elev_col),
                'Exp. Interface':   cell_item(row, exp_if_col),
                'Exp. Rack':        cell_item(row, exp_rack_col),
                'Exp. Elevation':   cell_item(row, exp_elev_col),
            }
        else:
            # OLD FORMAT: all columns present in the report
            cells = {
                'Hostname':         cell_item(row, host_col),
                'Interface':        cell_item(row, iface_col),
                'Rack':             cell_item(row, rack_col),
                'Elevation':        cell_item(row, elev_col),
                'Source_port':      cell_item(row, src_col),
                'DMARC1':           cell_item(row, d1_col),
                'DMARC2':           cell_item(row, d2_col),
                'Destination_port': cell_item(row, dest_col),
                'Z Hostname':       cell_item(row, z_host_col),
                'Z Interface':      cell_item(row, z_iface_col),
                'Z Rack':           cell_item(row, z_rack_col),
                'Z Elevation':      cell_item(row, z_elev_col),
                'Active Host':      cell_item(row, act_h_col),
                'Act. Interface':   cell_item(row, act_if_col),
                'Act. Rack':        cell_item(row, act_rack_col),
                'Act. Elevation':   cell_item(row, act_elev_col),
                'Exp. Interface':   cell_item(row, exp_if_col),
                'Exp. Rack':        cell_item(row, exp_rack_col),
                'Exp. Elevation':   cell_item(row, exp_elev_col),
            }

            # ── v9: PP_info_not_found fallback ───────────────────────────────
            # Old-format reports sometimes contain PP_info_not_found when the
            # report generator could not resolve the patch-panel path.
            # Replace those sentinel values with cutsheet data.
            _sp_val = str(cells['Source_port']['value'] or '')
            if _sp_val.startswith('PP_info'):
                # Try direct lookup first, then partner lane
                _cs = _cutsheet_pp.get((host, iface), {})
                if not _cs:
                    _m = re.match(r'(swp\d+)s(\d+)', iface)
                    if _m:
                        _pl = {0:1, 1:0, 2:3, 3:2}.get(int(_m.group(2)))
                        if _pl is not None:
                            _cs = _cutsheet_pp.get((host, f"{_m.group(1)}s{_pl}"), {})
                if _cs:
                    _pp_map = [
                        ('Source_port',      'source_port'),
                        ('DMARC1',           'dmarc1'),
                        ('DMARC2',           'dmarc2'),
                        ('Destination_port', 'dest_port'),
                        ('Z Hostname',       'z_hostname'),
                        ('Z Interface',      'z_interface'),
                        ('Z Rack',           'z_rack'),
                        ('Z Elevation',      'z_elevation'),
                    ]
                    for _ck, _vk in _pp_map:
                        _v = _cs.get(_vk, '')
                        if _v:
                            cells[_ck] = cs_item(_v)

            # For logical rows in old format — fill PP from partner if still blank
            if not is_p and not cells['Source_port']['value']:
                m = re.match(r'(swp\d+)s(\d+)', iface)
                if m:
                    base, lane = m.group(1), int(m.group(2))
                    partner_lane  = {0:1, 1:0, 2:3, 3:2}.get(lane)
                    partner_iface = f"{base}s{partner_lane}"
                    partner_row   = raw_rows.get((host, partner_iface))
                    if partner_row:
                        for k, col in [('Source_port', src_col), ('DMARC1', d1_col),
                                       ('DMARC2', d2_col), ('Destination_port', dest_col)]:
                            if col:
                                _pv = cell_item(partner_row, col)
                                # Prefer partner row value only if it isn't also a sentinel
                                if not str(_pv['value'] or '').startswith('PP_info'):
                                    cells[k] = _pv
                    if str(cells['Source_port']['value'] or '').startswith('PP_info') or \
                            not cells['Source_port']['value']:
                        pp_data = _cutsheet_pp.get((host, partner_iface), {})
                        if pp_data:
                            for k, cs_key in [('Source_port', 'source_port'),
                                              ('DMARC1', 'dmarc1'),
                                              ('DMARC2', 'dmarc2'),
                                              ('Destination_port', 'dest_port')]:
                                cells[k] = cs_item(pp_data.get(cs_key, ''))

        rows.append({
            't0': t0, 't1': t1, 'is_phys': is_p,
            'row_type': rtype, 'cells': cells,
            'mismatch_info': mi,
            '_host': host, '_iface': iface
        })

    # Second pass — logical rows inherit mismatch_info from physical partner
    mi_lookup = {}
    for rd in rows:
        if rd['is_phys'] and rd['mismatch_info']:
            mi_lookup[(rd['_host'], rd['_iface'])] = rd['mismatch_info']
    for rd in rows:
        if not rd['is_phys'] and not rd['mismatch_info']:
            m3 = re.match(r'(swp\d+)s(\d+)', rd['_iface'])
            if m3:
                base3, lane3 = m3.group(1), int(m3.group(2))
                partner3 = {0:1,1:0,2:3,3:2}.get(lane3)
                if partner3 is not None:
                    partner_key = (rd['_host'], f"{base3}s{partner3}")
                    if partner_key in mi_lookup:
                        rd['mismatch_info'] = mi_lookup[partner_key]

    return rows


# ── Summary Tab ───────────────────────────────────────────────────────────────
def build_summary_tab(wb_out, lldp_rows, miss_rows, down_rows,
                      prev_miss, prev_down, prev_opt,
                      report_name, prev_report_name,
                      prev_rack_map=None, prev_opt_rack_map=None, curr_opt_rack=None,
                      optics_count=0, fec_count=0):
    prev_rack_map     = prev_rack_map     or {}
    prev_opt_rack_map = prev_opt_rack_map or {}
    curr_opt_rack     = curr_opt_rack     or {}  # rack -> set of (host,iface) current optics
    import re as _re
    from datetime import datetime

    ws = wb_out.create_sheet("Summary", 0)  # insert as first tab
    ws.sheet_properties.tabColor = "1F4E79"

    NAVY  = "1F4E79"; WHITE = "FFFFFF"; RED   = "C00000"
    GREEN = "1E8449"; AMBER = "B7770D"; TEAL  = "0D7377"
    LRED  = "FADBD8"; LGRN  = "D5F5E3"; LYEL  = "FEF9E7"
    LGRY  = "F2F2F2"; DGRY  = "595959"; ORNG  = "E67E22"

    def fill(h):  return PatternFill("solid", fgColor=h)
    def font(color="000000", bold=False, sz=10, italic=False):
        return Font(bold=bold, italic=italic, color=color, name="Arial", size=sz)
    def center(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    def left():             return Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # ── Classify each row ──────────────────────────────────────────────────────
    def classify(rows, current_type, pm, pd, po):
        total = new = recurring = type_change = 0
        for rd in rows:
            h = str(rd['cells']['Hostname']['value']  or '').strip()
            i = str(rd['cells']['Interface']['value'] or '').strip()
            key = (h, i)
            flag, _ = get_history_flag(h, i, current_type, pm, pd, po)
            total += 1
            if not flag:                    new += 1
            elif '🔁' in flag:             recurring += 1
            else:                           type_change += 1
        return total, new, recurring, type_change

    miss_total,  miss_new,  miss_rec,  miss_tc  = classify(miss_rows,  'mismatch', prev_miss, prev_down, prev_opt)
    down_total,  down_new,  down_rec,  down_tc  = classify(down_rows,  'downlink', prev_miss, prev_down, prev_opt)

    # Optics from lldp_rows context not available directly — derive from tab
    # Use lldp_rows all for total counts
    total_issues = len(lldp_rows)
    has_prev = bool(prev_miss or prev_down or prev_opt)

    # Per-rack breakdown
    # Build current rack data with actual key sets for accurate Fixed calculation
    rack_data = {}
    curr_miss_by_rack = {}  # rack -> set of (host,iface) currently mismatch
    curr_down_by_rack = {}  # rack -> set of (host,iface) currently downlink

    curr_opt_by_rack = curr_opt_rack  # populated from optics sheet in main

    for rd in lldp_rows:
        h    = str(rd['cells']['Hostname']['value']  or '').strip()
        rack = str(rd['cells']['Rack']['value']      or '').strip() or 'Unknown'
        if rack not in rack_data:
            rack_data[rack] = {'miss':0,'down':0,'opt':0,'miss_rec':0,'down_rec':0,'miss_new':0,'down_new':0}
            curr_miss_by_rack[rack] = set()
            curr_down_by_rack[rack] = set()
            if rack not in curr_opt_by_rack: curr_opt_by_rack[rack] = set()
        key = (str(rd['cells']['Hostname']['value']  or '').strip(),
               str(rd['cells']['Interface']['value'] or '').strip())
        rtype = rd['row_type']
        flag, _ = get_history_flag(key[0], key[1], rtype, prev_miss, prev_down, prev_opt)
        if rtype == 'mismatch':
            rack_data[rack]['miss'] += 1
            curr_miss_by_rack[rack].add(key)
            if '🔁' in flag: rack_data[rack]['miss_rec'] += 1
            elif not flag:   rack_data[rack]['miss_new'] += 1
        elif rtype == 'downlink':
            rack_data[rack]['down'] += 1
            curr_down_by_rack[rack].add(key)
            if '🔁' in flag: rack_data[rack]['down_rec'] += 1
            elif not flag:   rack_data[rack]['down_new'] += 1

    # Build previous rack data key sets directly from prev_rack_map
    # prev_rack_map is passed in so we can look up rack for prev report keys
    # even if those links no longer exist in the current report
    prev_miss_by_rack = {}
    prev_down_by_rack = {}
    prev_opt_by_rack  = {}

    for (h, i), rack in prev_rack_map.items():
        key = (h, i)
        if key in prev_miss:
            if rack not in prev_miss_by_rack: prev_miss_by_rack[rack] = set()
            prev_miss_by_rack[rack].add(key)
        if key in prev_down:
            if rack not in prev_down_by_rack: prev_down_by_rack[rack] = set()
            prev_down_by_rack[rack].add(key)

    for (h, i), rack in prev_opt_rack_map.items():
        key = (h, i)
        if key in prev_opt:
            if rack not in prev_opt_by_rack: prev_opt_by_rack[rack] = set()
            prev_opt_by_rack[rack].add(key)

    # Also build current optics by rack from optics tab data
    # We need to pass current optics rows into summary - use lldp_rows for rack lookup
    # and cross-ref against all current optic keys
    # For now build from lldp_rows host_to_rack for any optic key we know about
    host_to_rack_curr = {(str(rd['cells']['Hostname']['value']  or '').strip(),
                          str(rd['cells']['Interface']['value'] or '').strip()):
                         str(rd['cells']['Rack']['value'] or '').strip() or 'Unknown'
                         for rd in lldp_rows}

    # Store key sets in rack_data for fixed calculation
    all_racks_set = (set(rack_data.keys()) | set(prev_miss_by_rack.keys()) |
                     set(prev_down_by_rack.keys()) | set(prev_opt_by_rack.keys()) |
                     set(curr_opt_by_rack.keys()))
    for rack in all_racks_set:
        if rack not in rack_data:
            rack_data[rack] = {'miss':0,'down':0,'opt':0,'miss_rec':0,'down_rec':0,'miss_new':0,'down_new':0}
        rack_data[rack]['curr_miss_keys'] = curr_miss_by_rack.get(rack, set())
        rack_data[rack]['curr_down_keys'] = curr_down_by_rack.get(rack, set())
        rack_data[rack]['curr_opt_keys']  = curr_opt_by_rack.get(rack, set())  # from curr_opt_rack passed in
        rack_data[rack]['prev_miss_keys'] = prev_miss_by_rack.get(rack, set())
        rack_data[rack]['prev_down_keys'] = prev_down_by_rack.get(rack, set())
        rack_data[rack]['prev_opt_keys']  = prev_opt_by_rack.get(rack, set())

    # ── Layout ─────────────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 3   # left margin
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 3   # gap

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("B1:G1")
    c = ws["B1"]; c.value = "VALIDATION REPORT — SUMMARY"
    c.fill = fill(NAVY); c.font = Font(bold=True, color=WHITE, name="Arial", size=14)
    c.alignment = center(); ws.row_dimensions[1].height = 32

    ws.merge_cells("B2:G2")
    c = ws["B2"]; c.value = f"Report: {report_name}   |   Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.fill = fill(TEAL); c.font = Font(italic=True, color=WHITE, name="Arial", size=9)
    c.alignment = center(); ws.row_dimensions[2].height = 16

    if prev_report_name:
        ws.merge_cells("B3:G3")
        c = ws["B3"]; c.value = f"Compared against: {prev_report_name}"
        c.fill = fill("2E4057"); c.font = Font(italic=True, color=WHITE, name="Arial", size=9)
        c.alignment = center(); ws.row_dimensions[3].height = 14
    else:
        ws.merge_cells("B3:G3")
        c = ws["B3"]; c.value = "No previous report selected — recurring analysis not available"
        c.fill = fill(DGRY); c.font = Font(italic=True, color=WHITE, name="Arial", size=9)
        c.alignment = center(); ws.row_dimensions[3].height = 14

    # ── KPI Banner ─────────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 8  # spacer

    grand_total = miss_total + down_total + optics_count + fec_count
    kpi_labels = ["TOTAL ISSUES", "MISMATCHES", "DOWNLINKS", "OPTICS", "FEC ERRORS"]
    kpi_values = [grand_total, miss_total, down_total, optics_count, fec_count]
    kpi_bgs    = [NAVY, RED, AMBER, "833C00", "7030A0"]

    for i, (lbl, val, bg) in enumerate(zip(kpi_labels, kpi_values, kpi_bgs)):
        col = i + 2  # B=2 through F=6
        ws.row_dimensions[5].height = 16
        ws.row_dimensions[6].height = 30
        c = ws.cell(5, col); c.value = lbl
        c.fill = fill(bg); c.font = Font(bold=True, color=WHITE, name="Arial", size=8)
        c.alignment = center(wrap=True)
        c = ws.cell(6, col); c.value = val
        c.fill = fill(bg); c.font = Font(bold=True, color=WHITE, name="Arial", size=20)
        c.alignment = center()

    # ── Error Type Breakdown ───────────────────────────────────────────────────
    ws.row_dimensions[7].height = 10  # spacer
    ws.merge_cells("B8:G8")
    c = ws["B8"]; c.value = "ERROR TYPE BREAKDOWN"
    c.fill = fill(NAVY); c.font = Font(bold=True, color=WHITE, name="Arial", size=10)
    c.alignment = center(); ws.row_dimensions[8].height = 20

    # Headers
    hdrs = ["Type", "Total", "🆕 New", "🔁 Recurring", "↔️ Type Change", "% Recurring"]
    bgs  = [NAVY, NAVY, GREEN, RED, ORNG, NAVY]
    for i, (h, bg) in enumerate(zip(hdrs, bgs)):
        c = ws.cell(9, i+2); c.value = h
        c.fill = fill(bg); c.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        c.alignment = center(); ws.row_dimensions[9].height = 18

    # Optics recurring/new from curr_opt_rack vs prev_opt set
    curr_opt_set = set()
    for s in curr_opt_rack.values():
        curr_opt_set |= s
    opt_rec_count = len(curr_opt_set & prev_opt)
    opt_new_count = len(curr_opt_set - prev_opt)

    rows_data = [
        ("Mispatches",  miss_total,    miss_new,      miss_rec,      miss_tc),
        ("Downlinks",   down_total,    down_new,      down_rec,      down_tc),
        ("Optics",      optics_count,  opt_new_count, opt_rec_count, 0),
        ("FEC Errors",  fec_count,     fec_count,     0,             0),
    ]
    for row_i, (lbl, tot, new_, rec, tc) in enumerate(rows_data):
        row = 10 + row_i
        ws.row_dimensions[row].height = 20
        pct = f"{round(rec/tot*100)}%" if tot > 0 else "—"
        row_bg = LRED if rec > 0 else LGRN
        vals = [lbl, tot, new_, rec, tc, pct]
        bgs2  = [LGRY, LGRY, LGRN, LRED, LYEL, LGRY]
        for col_i, (v, bg) in enumerate(zip(vals, bgs2)):
            c = ws.cell(row, col_i+2)
            c.value = v; c.fill = fill(bg)
            bold = col_i in (0, 1)
            c.font = Font(bold=bold, name="Arial", size=10,
                          color=RED if bg==LRED and v else (GREEN if bg==LGRN and v else "000000"))
            c.alignment = center() if col_i > 0 else left()

    # ── Per-Rack Breakdown ─────────────────────────────────────────────────────
    # Build previous report rack data
    prev_rack = {}
    for key in prev_miss:
        for rd in lldp_rows:
            h2 = str(rd['cells']['Hostname']['value']  or '').strip()
            i2 = str(rd['cells']['Interface']['value'] or '').strip()
            if (h2, i2) == key:
                r2 = str(rd['cells']['Rack']['value'] or '').strip() or 'Unknown'
                if r2 not in prev_rack: prev_rack[r2] = {'miss':0,'down':0,'opt':0}
                prev_rack[r2]['miss'] += 1
                break
    for key in prev_down:
        for rd in lldp_rows:
            h2 = str(rd['cells']['Hostname']['value']  or '').strip()
            i2 = str(rd['cells']['Interface']['value'] or '').strip()
            if (h2, i2) == key:
                r2 = str(rd['cells']['Rack']['value'] or '').strip() or 'Unknown'
                if r2 not in prev_rack: prev_rack[r2] = {'miss':0,'down':0,'opt':0}
                prev_rack[r2]['down'] += 1
                break

    # All racks from either prev or current
    all_racks = sorted(set(list(rack_data.keys()) + list(prev_rack.keys())))

    ws.row_dimensions[13].height = 10  # spacer
    ws.merge_cells("B14:N14")
    c = ws["B14"]; c.value = "PER-RACK BREAKDOWN  —  Previous vs Now"
    c.fill = fill(NAVY); c.font = Font(bold=True, color=WHITE, name="Arial", size=10)
    c.alignment = center(); ws.row_dimensions[14].height = 20

    # Column layout:
    # B=Rack | C=Miss Prev | D=Miss Now | E=Miss Fixed | F=Miss New
    #        | G=Down Prev | H=Down Now | I=Down Fixed | J=Down New
    #        | K=Opt Prev  | L=Opt Now  | M=Opt Fixed  | N=Opt New
    ws.column_dimensions['B'].width = 10
    for ltr, w in zip('CDEFGHIJKLMN', [9,9,9,9, 9,9,9,9, 9,9,9,9]):
        ws.column_dimensions[ltr].width = w

    # Group headers row 15
    for col, label, bg in [
        (3,  "MISMATCHES", "C00000"),
        (7,  "DOWNLINKS",  AMBER),
        (11, "OPTICS",     "7D3C98"),
    ]:
        ws.merge_cells(start_row=15, start_column=col, end_row=15, end_column=col+3)
        c = ws.cell(15, col); c.value = label; c.fill = fill(bg)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        c.alignment = center()
    ws.cell(15, 2).value = "Rack"; ws.cell(15, 2).fill = fill(NAVY)
    ws.cell(15, 2).font = Font(bold=True, color=WHITE, name="Arial", size=9)
    ws.cell(15, 2).alignment = center()
    ws.row_dimensions[15].height = 16

    # Sub-headers row 16
    sub_hdrs = ["Rack", "Prev","Now","Fixed","New", "Prev","Now","Fixed","New", "Prev","Now","Fixed","New"]
    sub_bgs  = [NAVY, LGRY,LGRY,LGRN,LRED, LGRY,LGRY,LGRN,LRED, LGRY,LGRY,LGRN,LRED]
    for i, (h2, bg) in enumerate(zip(sub_hdrs, sub_bgs)):
        c = ws.cell(16, i+2); c.value = h2; c.fill = fill(bg)
        c.font = Font(bold=True, color="000000" if i>0 else WHITE, name="Arial", size=8)
        c.alignment = center()
    ws.row_dimensions[16].height = 14

    # Data rows
    for row_i, rack in enumerate(all_racks):
        row = 17 + row_i
        ws.row_dimensions[row].height = 18
        curr = rack_data.get(rack, {'miss':0,'down':0,'miss_new':0,'down_new':0})
        prev = prev_rack.get(rack, {'miss':0,'down':0,'opt':0})

        curr_miss = curr.get('miss', 0)
        curr_down = curr.get('down', 0)
        curr_new_miss = curr.get('miss_new', 0)
        curr_new_down = curr.get('down_new', 0)
        prev_miss_r = len(curr.get('prev_miss_keys', set()))
        prev_down_r = len(curr.get('prev_down_keys', set()))

        # Fixed = links in prev report for this rack that are NOT in current report
        # Uses actual set of keys stored per rack rather than arithmetic
        miss_fixed = len(curr.get('prev_miss_keys', set()) - curr.get('curr_miss_keys', set()))
        down_fixed = len(curr.get('prev_down_keys', set()) - curr.get('curr_down_keys', set()))
        opt_fixed  = len(curr.get('prev_opt_keys',  set()) - curr.get('curr_opt_keys',  set()))

        curr_opt   = len(curr.get('curr_opt_keys', set()))
        prev_opt_r = len(curr.get('prev_opt_keys', set()))
        curr_new_opt = len(curr.get('curr_opt_keys', set()) - curr.get('prev_opt_keys', set()))

        vals = [
            rack,
            prev_miss_r, curr_miss,  miss_fixed, curr_new_miss,
            prev_down_r, curr_down,  down_fixed, curr_new_down,
            prev_opt_r,  curr_opt,   opt_fixed,  curr_new_opt,
        ]
        for col_i, v in enumerate(vals):
            c = ws.cell(row, col_i+2)
            c.value = v
            # Colour: Fixed=green, New=red, others plain
            if col_i in (3, 7, 11) and v > 0:   c.fill = fill(LGRN)  # fixed
            elif col_i in (4, 8, 12) and v > 0:  c.fill = fill(LRED)  # new
            else:                                  c.fill = fill("FFFFFF" if col_i>0 else LGRY)
            c.font = Font(bold=(col_i==0), name="Arial", size=9,
                          color=GREEN if col_i in (3,7,11) and v>0
                          else RED if col_i in (4,8,12) and v>0 else "000000")
            c.alignment = center() if col_i > 0 else left()

    print(f"  Summary tab built — {len(all_racks)} racks")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    cfg = load_config()

    # Step 1: cutsheets — support multiple for full mismatch coverage
    saved_paths = cfg.get('cutsheet_paths', [])
    # Back-compat: handle old single path config
    if not saved_paths and cfg.get('cutsheet_path'):
        saved_paths = [cfg['cutsheet_path']]
    saved_paths = [p for p in saved_paths if os.path.isfile(p)]

    cutsheet_paths = []
    if saved_paths:
        names = '\n'.join(f"  • {os.path.basename(p)}" for p in saved_paths)
        if HAS_TK:
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            use_saved = messagebox.askyesno("Physical Cutsheets",
                f"Use saved cutsheets?\n\n{names}\n\nClick No to pick different ones.")
            root.destroy()
        else:
            print(f"Saved cutsheets:\n{names}")
            use_saved = input("Use these? (y/n): ").strip().lower() != 'n'
        if use_saved:
            cutsheet_paths = saved_paths

    if not cutsheet_paths:
        show_msg("Select Cutsheets",
            "Select all physical cutsheets for this job.\n\nHold Ctrl to select multiple files.")
        cutsheet_paths = pick_multiple_files(
            "Select Physical Cutsheet(s) — hold Ctrl for multiple")
        if not cutsheet_paths:
            show_msg("Cancelled", "No cutsheets selected.", error=True); sys.exit(0)
        cfg['cutsheet_paths'] = cutsheet_paths
        cfg['cutsheet_path']  = cutsheet_paths[0]  # back-compat
        save_config(cfg)

    # Step 2: previous report (optional)
    time.sleep(0.3)
    prev_report_path = None
    if HAS_TK:
        root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True)
        want_prev = messagebox.askyesno("Previous Report",
            "Do you have a previous report to compare against?\n\n"
            "This will flag recurring issues across all tabs.")
        root.destroy()
    else:
        want_prev = input("Compare against a previous report? (y/n): ").strip().lower() == 'y'

    if want_prev:
        time.sleep(0.3)
        prev_report_path = pick_file("Select PREVIOUS Slack Validation Report")
        if prev_report_path:
            print(f"Previous report: {os.path.basename(prev_report_path)}")

    # Step 3: current report
    time.sleep(0.4)
    report_path = pick_file("Select Slack Validation Report")
    if not report_path:
        show_msg("Cancelled", "No report selected.", error=True); sys.exit(0)

    # Check for local hall cutsheet in same folder - append to list
    local = os.path.join(os.path.dirname(report_path), "cutsheet.xlsx")
    if os.path.isfile(local) and local not in cutsheet_paths:
        cutsheet_paths.append(local)
        print(f"Also loading local hall cutsheet: {local}")

    print(f"Loading cutsheet...")
    phys_t0, phys_t1, t1_rev = build_lookup(cutsheet_paths)
    print(f"  {len(phys_t0)} T0 entries | {len(t1_rev)} T1 reverse entries")

    # Load previous report issues if provided
    prev_miss = set(); prev_down = set(); prev_opt = set(); prev_rack_map = {}; prev_opt_rack_map = {}
    if prev_report_path:
        print("Loading previous report for comparison...")
        prev_miss, prev_down, prev_opt, prev_rack_map, prev_opt_rack_map = get_prev_issues(prev_report_path)

    print(f"Processing: {os.path.basename(report_path)}")
    wb_src = load_workbook(report_path)

    def find_sheet(wb, *patterns):
        for name in wb.sheetnames:
            for p in patterns:
                if p.lower() in name.lower():
                    return wb[name]
        return None

    ws_lldp        = find_sheet(wb_src, 'lldp')
    # Split optics: rx/tx threshold sheet and optional temperature sheet
    ws_optics      = find_sheet(wb_src, 'optics_rx_tx', 'rx_tx_threshold')
    ws_optics_temp = find_sheet(wb_src, 'optics_tmp', 'optics_temp')
    # Fallback: if no specific threshold sheet matched, use any optics sheet
    if ws_optics is None:
        for _n in wb_src.sheetnames:
            if 'optic' in _n.lower() and 'tmp' not in _n.lower() and 'temp' not in _n.lower():
                ws_optics = wb_src[_n]; break
    ws_fec         = find_sheet(wb_src, 'fec')

    wb_out = Workbook(); wb_out.remove(wb_out.active)

    lldp_rows = miss_rows = down_rows = []

    if ws_lldp:
        lldp_rows = read_lldp_rows(ws_lldp, phys_t0, phys_t1, t1_rev)
        miss_rows = [r for r in lldp_rows if r['row_type'] == 'mismatch']
        down_rows = [r for r in lldp_rows if r['row_type'] == 'downlink']

        build_lldp_sheet(wb_out, "Mispatches", miss_rows, TAB_MISS, is_mismatch=True, prev_miss=prev_miss, prev_down=prev_down, prev_opt=prev_opt)
        build_lldp_sheet(wb_out, "Downlinks",  down_rows, TAB_DOWN, prev_miss=prev_miss, prev_down=prev_down, prev_opt=prev_opt, is_downlinks=True)
        print(f"  LLDP — Mismatches:{len(miss_rows)} Downlinks:{len(down_rows)}")
        # Store for summary tab
        _miss_rows = miss_rows; _down_rows = down_rows; _lldp_rows = lldp_rows

    # Pass prev_opt to optics loop
    prev_opt_for_loop = prev_opt

    # Build downlinks set for cross-reference (hostname+interface)
    downlink_set = set()
    if ws_lldp:
        act_if_col_lldp = find_col(ws_lldp, 'Act. Interface', 'Act.Interface')
        host_col_lldp   = find_col(ws_lldp, 'Hostname')
        iface_col_lldp  = find_col(ws_lldp, 'Interface')
        if act_if_col_lldp and host_col_lldp and iface_col_lldp:
            for row in range(2, ws_lldp.max_row + 1):
                act_if = str(ws_lldp.cell(row, act_if_col_lldp).value or '').strip().lower()
                if act_if == 'interface down':
                    h = str(ws_lldp.cell(row, host_col_lldp).value or '').strip()
                    i = str(ws_lldp.cell(row, iface_col_lldp).value or '').strip()
                    if h and i:
                        downlink_set.add((h, i))

    # Optics and FEC — physical highlighting + T0 L&R + T1 L&R + blank PP_info
    for ws_extra, tab_name, tab_col in [
        (ws_optics,      "Optics",      TAB_OPT),
        (ws_optics_temp, "Optics Temp", "C65911"),
        (ws_fec,         "FEC Errors",  TAB_FEC),
    ]:
        current_type_for_tab = 'optic' 
        if not ws_extra: continue
        host_col    = find_col(ws_extra, 'Hostname')
        iface_col   = find_col(ws_extra, 'Interface')
        z_iface_col = find_col(ws_extra, 'Z Interface')
        if not host_col or not iface_col: continue

        # FEC-specific: extra PP/Z columns injected after Elevation
        is_fec = (tab_name == "FEC Errors")
        elev_col_src          = find_col(ws_extra, 'Elevation')     if is_fec else None
        remote_iface_col_src  = find_col(ws_extra, 'Remote Interface') if is_fec else None

        # PP/Z block definition for FEC (8 extra cols after Elevation)
        PP_Z_HEADERS = [
            ("Source_port",      "C0504D", 30),
            ("DMARC1",           "7F6000", 28),
            ("DMARC2",           "375623", 28),
            ("Destination_port", "17375E", 30),
            ("Z Interface",      "17375E", 12),
            ("Z L&R",            "17375E",  6),
            ("Z Rack",           "17375E",  7),
            ("Z Elevation",      "17375E",  6),
        ]
        pp_z_start_col = None   # set once we've passed Elevation in mapping loop

        ws_out = wb_out.create_sheet(tab_name)
        ws_out.sheet_properties.tabColor = tab_col
        ncols = ws_extra.max_column

        # Cols to skip in output (by name) — covers both old and new format columns
        skip_col_names = {'Hostname', 'Z Hostname', 'Remote Host', 'Transceiver',
                          'Min Threshold (dBm)', 'Max Threshold (dBm)',
                          'Source Sheet', 'Placement Group', 'Building', '_idx', 'Index'}
        skip_src_set = {c for c in range(1, ncols+1)
                        if str(ws_extra.cell(1,c).value or '').strip() in skip_col_names}

        # Build src->out mapping; for FEC inject 8 PP/Z cols immediately after Elevation
        src_to_out_map = {}; out_c = 1; t0_lr_col = None; t1_lr_col = None
        for sc in range(1, ncols + 1):
            if sc in skip_src_set: continue
            if sc == iface_col:
                src_to_out_map[sc] = out_c; out_c += 1
                t0_lr_col = out_c; out_c += 1
            elif z_iface_col and sc == z_iface_col:
                src_to_out_map[sc] = out_c; out_c += 1
                t1_lr_col = out_c; out_c += 1
            else:
                src_to_out_map[sc] = out_c; out_c += 1
            # After Elevation: inject PP/Z block (FEC only)
            if is_fec and sc == elev_col_src and pp_z_start_col is None:
                pp_z_start_col = out_c
                out_c += len(PP_Z_HEADERS)

        total_out_cols = out_c - 1
        flag_col     = total_out_cols + 1
        hist_col_num = flag_col + 1

        # Write source-column headers
        for sc, oc in src_to_out_map.items():
            c = ws_out.cell(1, oc); c.value = ws_extra.cell(1, sc).value
            c.fill = fill(HDR_BG); c.font = font(HDR_FG, bold=True, sz=9)
            c.alignment = center()

        # L&R header(s)
        for lr_c, lr_lbl in ([(t0_lr_col, "L&R")] + ([(t1_lr_col, "Z L&R")] if t1_lr_col else [])):
            c = ws_out.cell(1, lr_c); c.value = lr_lbl
            c.fill = fill(HDR_BG); c.font = font(HDR_FG, bold=True, sz=9)
            c.alignment = center()
            ws_out.column_dimensions[get_column_letter(lr_c)].width = 6

        # FEC PP/Z block headers
        if pp_z_start_col is not None:
            for i, (name, bg, w) in enumerate(PP_Z_HEADERS):
                col_idx = pp_z_start_col + i
                c = ws_out.cell(1, col_idx)
                c.value = name; c.fill = fill(bg)
                c.font = font(HDR_FG, bold=True, sz=9); c.alignment = center()
                ws_out.column_dimensions[get_column_letter(col_idx)].width = w

        # DL Flag + History headers
        cf = ws_out.cell(1, flag_col)
        cf.value = "DL Flag"; cf.fill = fill("595959")
        cf.font = Font(bold=True, color=WHITE, name="Arial", size=8)
        cf.alignment = Alignment(horizontal="center", vertical="center")
        ws_out.column_dimensions[get_column_letter(flag_col)].width = 24

        ch = ws_out.cell(1, hist_col_num)
        ch.value = "History"; ch.fill = fill("595959")
        ch.font = Font(bold=True, color=WHITE, name="Arial", size=8)
        ch.alignment = Alignment(horizontal="center", vertical="center")
        ws_out.column_dimensions[get_column_letter(hist_col_num)].width = 22
        ws_out.row_dimensions[1].height = 20

        dl_overlap = 0
        # First pass — build (host, iface) -> row number lookup for partner alignment
        raw_opt_rows = {}
        for _r in range(2, ws_extra.max_row + 1):
            _h = str(ws_extra.cell(_r, host_col).value or '').strip()
            _i = str(ws_extra.cell(_r, iface_col).value or '').strip()
            if _h and _i: raw_opt_rows[(_h, _i)] = _r

        # Find patch panel cols in this sheet (Source_port, DMARC1, DMARC2, Destination_port)
        pp_col_names = ['Source_port', 'DMARC1', 'DMARC2', 'Destination_port']
        pp_cols = []
        for name in pp_col_names:
            c = find_col(ws_extra, name)
            if c: pp_cols.append(c)

        out_row = 2
        for row in range(2, ws_extra.max_row + 1):
            host  = str(ws_extra.cell(row, host_col).value or '').strip()
            iface = str(ws_extra.cell(row, iface_col).value or '').strip()
            if not host or not iface: continue
            t0_lbl, t1_lbl, is_p = get_labels(host, iface, phys_t0, phys_t1)

            # For logical rows — find physical partner and copy its patch panel cols
            partner_pp_override = {}
            if not is_p:
                m_pp = re.match(r'(swp\d+)s(\d+)', iface)
                if m_pp:
                    base_pp, lane_pp = m_pp.group(1), int(m_pp.group(2))
                    partner_lane_pp = {0:1,1:0,2:3,3:2}.get(lane_pp)
                    if partner_lane_pp is not None:
                        partner_iface_pp = f"{base_pp}s{partner_lane_pp}"
                        partner_row_pp = raw_opt_rows.get((host, partner_iface_pp))
                        if partner_row_pp:
                            for pp_c in pp_cols:
                                partner_pp_override[pp_c] = ws_extra.cell(partner_row_pp, pp_c).value
                        else:
                            pp_fb = _cutsheet_pp.get((host, partner_iface_pp), {})
                            if pp_fb:
                                pp_key_map = {
                                    'source_port': 'Source_port',
                                    'dmarc1':      'DMARC1',
                                    'dmarc2':      'DMARC2',
                                    'dest_port':   'Destination_port',
                                }
                                for cs_key, col_name in pp_key_map.items():
                                    c_num = find_col(ws_extra, col_name)
                                    if c_num:
                                        partner_pp_override[c_num] = pp_fb.get(cs_key, "")

            is_also_downlink = (host, iface) in downlink_set
            if is_also_downlink: dl_overlap += 1

            row_bg = "C8C8C8" if is_also_downlink else ("FFFFFF" if is_p else LOG_BG)
            lr_bg  = "A8A8A8" if is_also_downlink else (LR_BG if is_p else LR_LOG)
            txt_fg = "888888" if is_also_downlink else "000000"

            # Build cutsheet lookup for this row (used for PP_info fallback below)
            _cs_row = _cutsheet_pp.get((host, iface), {})
            if not _cs_row:
                _mf = re.match(r'(swp\d+)s(\d+)', iface)
                if _mf:
                    _pf = {0:1,1:0,2:3,3:2}.get(int(_mf.group(2)))
                    if _pf is not None:
                        _cs_row = _cutsheet_pp.get((host, f'{_mf.group(1)}s{_pf}'), {})
            _col_to_cs = {
                'Source_port': 'source_port', 'DMARC1': 'dmarc1',
                'DMARC2': 'dmarc2', 'Destination_port': 'dest_port',
                'Z Hostname': 'z_hostname', 'Z Interface': 'z_interface',
                'Z Rack': 'z_rack', 'Z Elevation': 'z_elevation',
            }

            # Write data using src_to_out_map
            # v9: replace PP_info_not_found with cutsheet value instead of blanking
            for sc, oc in src_to_out_map.items():
                c = ws_out.cell(out_row, oc)
                raw = partner_pp_override.get(sc, ws_extra.cell(row, sc).value)
                if str(raw or '').startswith('PP_info'):
                    col_name = str(ws_extra.cell(1, sc).value or '').strip()
                    cs_key   = _col_to_cs.get(col_name)
                    raw = _cs_row.get(cs_key, '') if cs_key else ''
                c.value = raw; c.fill = fill(row_bg); c.font = font(sz=8, color=txt_fg)
                c.alignment = vcenter()

            # T0 L&R
            c = ws_out.cell(out_row, t0_lr_col)
            c.value = t0_lbl; c.fill = fill(lr_bg)
            c.font = font(sz=8, bold=True, color=txt_fg); c.alignment = center()
            if t1_lr_col:
                c = ws_out.cell(out_row, t1_lr_col)
                c.value = t1_lbl; c.fill = fill(lr_bg)
                c.font = font(sz=8, bold=True, color=txt_fg); c.alignment = center()

            # ── FEC PP/Z block — filled from cutsheet ──────────────────────────
            if pp_z_start_col is not None:
                # Direct cutsheet lookup; fall back to partner lane
                cs_fec = _cutsheet_pp.get((host, iface), {})
                if not cs_fec:
                    m_f = re.match(r'(swp\d+)s(\d+)', iface)
                    if m_f:
                        pl_f = {0:1,1:0,2:3,3:2}.get(int(m_f.group(2)))
                        if pl_f is not None:
                            cs_fec = _cutsheet_pp.get((host, f'{m_f.group(1)}s{pl_f}'), {})

                # Z Interface: prefer actual Remote Interface from source sheet
                remote_if_val = (str(ws_extra.cell(row, remote_iface_col_src).value or '').strip()
                                 if remote_iface_col_src else '')
                z_iface_fec = remote_if_val or cs_fec.get('z_interface', '')

                pp_z_vals = [
                    cs_fec.get('source_port', ''),   # Source_port
                    cs_fec.get('dmarc1',      ''),   # DMARC1
                    cs_fec.get('dmarc2',      ''),   # DMARC2
                    cs_fec.get('dest_port',   ''),   # Destination_port
                    z_iface_fec,                     # Z Interface
                    t1_lbl,                          # Z L&R (from get_labels)
                    cs_fec.get('z_rack',      ''),   # Z Rack
                    cs_fec.get('z_elevation', ''),   # Z Elevation
                ]
                for i, val in enumerate(pp_z_vals):
                    c = ws_out.cell(out_row, pp_z_start_col + i)
                    c.value = val
                    c.fill = fill(row_bg)
                    c.font = font(sz=8, color=txt_fg)
                    c.alignment = vcenter()

            # DL Flag
            cf = ws_out.cell(out_row, flag_col)
            if is_also_downlink:
                cf.value = "⬇️ Also Downlink — skip"
                cf.fill  = fill("C8C8C8")
                cf.font  = Font(bold=True, color="666666", name="Arial", size=8)
            else:
                cf.fill = fill(row_bg); cf.font = font(sz=8)
            cf.alignment = Alignment(horizontal="center", vertical="center")

            # History flag
            hist_text, hist_colour = get_history_flag(
                host, iface, current_type_for_tab, prev_miss, prev_down, prev_opt_for_loop
            )
            ch = ws_out.cell(out_row, hist_col_num)
            if hist_text:
                ch.value = hist_text
                ch.fill  = fill(hist_colour)
                ch.font  = Font(bold=True, color="FFFFFF", name="Arial", size=8)
            else:
                ch.fill = fill(row_bg); ch.font = font(sz=8)
            ch.alignment = Alignment(horizontal="center", vertical="center")

            ws_out.row_dimensions[out_row].height = 15
            out_row += 1

        ws_out.freeze_panes = "A2"

        # ── Borders around each physical+logical pair ─────────────────────────
        from openpyxl.styles import Border, Side
        thin2  = Side(style="thin",   color="AAAAAA")
        thick2 = Side(style="medium", color="555555")
        lr_col2 = t0_lr_col  # L&R col position
        dr = 2
        while dr <= ws_out.max_row:
            lr_val2 = ws_out.cell(dr, lr_col2).value
            grp_end = dr
            while (grp_end + 1 <= ws_out.max_row and
                   ws_out.cell(grp_end+1, lr_col2).value == lr_val2 and lr_val2):
                grp_end += 1
            for rr in range(dr, grp_end + 1):
                is_top2    = (rr == dr)
                is_bottom2 = (rr == grp_end)
                for cc in range(1, ws_out.max_column + 1):
                    ws_out.cell(rr, cc).border = Border(
                        top    = thick2 if is_top2    else thin2,
                        bottom = thick2 if is_bottom2 else Side(style=None),
                        left   = thick2 if cc == 1    else thin2,
                        right  = thick2 if cc == ws_out.max_column else thin2,
                    )
            dr = grp_end + 1

        print(f"  {tab_name} — {out_row-2} rows | {dl_overlap} flagged as also-downlink")

    # ── Build Summary tab ────────────────────────────────────────────────────
    if ws_lldp:
        # Build current optics by rack for summary
        curr_opt_rack = {}
        if ws_optics:
            _ohc = find_col(ws_optics, 'Hostname')
            _oic = find_col(ws_optics, 'Interface')
            _orc = find_col(ws_optics, 'Rack')
            if _ohc and _oic:
                for _r in range(2, ws_optics.max_row+1):
                    _h = str(ws_optics.cell(_r, _ohc).value or '').strip()
                    _i = str(ws_optics.cell(_r, _oic).value or '').strip()
                    _rack = str(ws_optics.cell(_r, _orc).value or '').strip() if _orc else 'Unknown'
                    if _h and _i:
                        if _rack not in curr_opt_rack: curr_opt_rack[_rack] = set()
                        curr_opt_rack[_rack].add((_h, _i))

        _opt_count = max(0, ws_optics.max_row - 1) if ws_optics else 0
        if ws_optics_temp: _opt_count += max(0, ws_optics_temp.max_row - 1)
        _fec_count = max(0, ws_fec.max_row - 1) if ws_fec else 0

        build_summary_tab(wb_out, _lldp_rows, _miss_rows, _down_rows,
                          prev_miss, prev_down, prev_opt,
                          os.path.basename(report_path),
                          os.path.basename(prev_report_path) if prev_report_path else None,
                          prev_rack_map=prev_rack_map, prev_opt_rack_map=prev_opt_rack_map,
                          curr_opt_rack=curr_opt_rack,
                          optics_count=_opt_count, fec_count=_fec_count)

    base, ext = os.path.splitext(report_path)
    out_path  = base + "_highlighted" + ext
    wb_out.save(out_path)

    msg = (f"Done!\n\n"
           f"All: {len(lldp_rows)}  |  Mispatches: {len(miss_rows)}  |  Downlinks: {len(down_rows)}\n\n"
           f"Saved to:\n{out_path}")
    print(f"\n{msg}")
    show_msg("Complete ✅", msg)

    try:
        import subprocess
        if sys.platform   == "win32":  os.startfile(out_path)
        elif sys.platform == "darwin": subprocess.run(["open", out_path])
        else:                          subprocess.run(["xdg-open", out_path])
    except: pass

if __name__ == "__main__":
    main()