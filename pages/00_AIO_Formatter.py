#!/usr/bin/env python3
"""
AIO Validation Formatter (All-In-One)

The single tool that can handle ANY hall, ANY fabric type, ANY source.

It reads the cutsheet first (the cutsheet knows what this is),
then automatically chooses and runs the correct rich formatting logic
(GFAB, HOPS, CFAB, IPR, future LV Portal, etc.).

This is the recommended starting point for most users.
"""

import streamlit as st
import sys
from pathlib import Path
import tempfile
import os

# Robust import for both local run and Streamlit Cloud
root_dir = Path(__file__).parent.parent
if str(root_dir) not in sys.path:
    sys.path.insert(0, str(root_dir))

try:
    from logic import aio_logic as aio
except ImportError as e:
    st.error(f"Failed to import logic.aio_logic: {e}")
    st.stop()

st.set_page_config(
    page_title="AIO Validation Formatter",
    page_icon="🧠",
    layout="wide"
)

st.title("🧠 All In One Validation Formatter")
st.caption("Any Hall • Any Fabric (GFAB/CFAB/HOPS/LV-Portal/etc.) • Slack or LVV Portal • One Button")

st.markdown("""
**This is the main tool you should use.**

Upload your raw validation exports + the relevant **cutsheet**.  
The AIO reads the cutsheet to understand what kind of fabric/pathway this is 
(GFAB, CFAB/HOPS, T3-T2 vs T2-T1-T0, LV Portal style, Slack export, etc.), 
then automatically dispatches to the correct rich processing logic we have built.

No more guessing which script to run.
""")

# =============================================================================
# Session State for multi-step flow (prevents looping back to start)
# =============================================================================
if "aio_stage" not in st.session_state:
    st.session_state.aio_stage = "upload"
if "aio_detection" not in st.session_state:
    st.session_state.aio_detection = None
if "aio_input_bytes" not in st.session_state:
    st.session_state.aio_input_bytes = []   # list of (filename, bytes)
if "aio_cutsheet_bytes" not in st.session_state:
    st.session_state.aio_cutsheet_bytes = None
if "aio_override" not in st.session_state:
    st.session_state.aio_override = None
if "aio_result" not in st.session_state:
    st.session_state.aio_result = None  # (bytes, filename)

# =============================================================================
# STAGE 1: Upload + Detect
# =============================================================================
if st.session_state.aio_stage == "upload":
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("1. Raw Validation Files")
        input_files = st.file_uploader(
            "Upload one or more validation export files (.xlsx)",
            type=["xlsx"],
            accept_multiple_files=True,
            key="aio_inputs",
            help="These are the messy files coming out of the hall testing / Slack / LVV Portal"
        )

    with col2:
        st.subheader("2. Cutsheet (Required)")
        cutsheet_file = st.file_uploader(
            "The cutsheet that describes this fabric (this file tells the AIO what to do)",
            type=["xlsx"],
            key="aio_cutsheet",
            help="The cutsheet contains the intelligence about hall, fabric type, and source system"
        )

    if not (input_files and cutsheet_file):
        st.info("Please upload at least one validation file and a cutsheet to enable the button below.")

    if st.button("🚀 Detect & Process (AIO)", type="primary", disabled=not (input_files and cutsheet_file)):
        # Save files to session state
        st.session_state.aio_input_bytes = [(f.name, f.getbuffer().tobytes()) for f in input_files]
        st.session_state.aio_cutsheet_bytes = (cutsheet_file.name, cutsheet_file.getbuffer().tobytes())

        with st.spinner("Inspecting cutsheet and detecting configuration..."):
            with tempfile.TemporaryDirectory() as tmpdir:
                # Write to temp for detection
                cutsheet_path = os.path.join(tmpdir, "cutsheet.xlsx")
                with open(cutsheet_path, "wb") as out:
                    out.write(st.session_state.aio_cutsheet_bytes[1])

                detection = aio.inspect_cutsheet(cutsheet_path)
                st.session_state.aio_detection = detection
                st.session_state.aio_stage = "detected"
                st.rerun()

# =============================================================================
# STAGE 2: Detection Results + Override + Confirm
# =============================================================================
elif st.session_state.aio_stage == "detected":
    detection = st.session_state.aio_detection

    st.subheader("🔍 Cutsheet Detection Results")
    st.info(detection.summary())

    if detection.reasons:
        with st.expander("Why the AIO thinks this:", expanded=False):
            for r in detection.reasons:
                st.write(f"• {r}")

    if detection.cutsheet_metadata:
        with st.expander("Cutsheet metadata (first sheet + sheets)", expanded=False):
            st.json(detection.cutsheet_metadata)

    # Override
    st.divider()
    st.subheader("Optional Override")

    processor_options = ["auto (use detection)", "gfab", "hops", "cfab", "ipr", "lv_portal"]
    default_index = 0
    if st.session_state.aio_override:
        try:
            default_index = processor_options.index(st.session_state.aio_override)
        except ValueError:
            default_index = 0

    chosen = st.selectbox(
        "Force a specific processor (only if detection is wrong)",
        processor_options,
        index=default_index,
        key="aio_override_select"
    )

    if st.button("✅ Confirm and Generate Report", type="primary"):
        override = None if chosen == "auto (use detection)" else chosen
        st.session_state.aio_override = override

        processor_to_use = override or detection.recommended_processor

        # Guard for types that don't have full support yet
        supported = ["gfab", "hops", "cfab", "ipr", "lv_portal"]
        if processor_to_use not in supported:
            st.error(f"Processor '{processor_to_use}' is not yet fully wired in the AIO.")
            st.info("For JBP15 T0-to-Host / LV Portal reports, please use the dedicated tool in your JBP15 T0_to_Host folder for now.")
            if st.button("← Back to Upload"):
                for key in list(st.session_state.keys()):
                    if key.startswith("aio_"):
                        del st.session_state[key]
                st.rerun()
            st.stop()

        with st.spinner(f"Running {processor_to_use} logic..."):
            try:
                with tempfile.TemporaryDirectory() as tmpdir:
                    input_paths = []
                    for i, (name, data) in enumerate(st.session_state.aio_input_bytes):
                        p = os.path.join(tmpdir, f"input_{i}.xlsx")
                        with open(p, "wb") as out:
                            out.write(data)
                        input_paths.append(p)

                    cutsheet_path = os.path.join(tmpdir, "cutsheet.xlsx")
                    with open(cutsheet_path, "wb") as out:
                        out.write(st.session_state.aio_cutsheet_bytes[1])

                    if len(input_paths) == 1:
                        xlsx_bytes, fname, info = aio.process_aio_validation(
                            input_paths, cutsheet_path, override_processor=override
                        )
                        st.session_state.aio_result = (xlsx_bytes, fname, info, False)
                    else:
                        zip_bytes, info = aio.process_multiple_aio_validation(
                            input_paths, cutsheet_path, override_processor=override
                        )
                        st.session_state.aio_result = (zip_bytes, "AIO_Formatted_Reports.zip", info, True)

                    st.session_state.aio_stage = "processed"
                    st.rerun()

            except NotImplementedError as nie:
                st.warning(str(nie))
                st.info("Detection is now correct for JBP15 T0-to-Host / LV Portal reports. Full end-to-end generation is the next piece being integrated.")
            except Exception as proc_err:
                st.error(f"Processing failed for this file type.")
                st.exception(proc_err)
                st.info("This combination may not be fully supported yet. Try forcing a different processor in the Override dropdown.")
                if st.button("← Back to Upload"):
                    for key in list(st.session_state.keys()):
                        if key.startswith("aio_"):
                            del st.session_state[key]
                    st.rerun()

    if st.button("← Start Over"):
        for key in list(st.session_state.keys()):
            if key.startswith("aio_"):
                del st.session_state[key]
        st.rerun()

# =============================================================================
# STAGE 3: Results + Analytics + Download
# =============================================================================
elif st.session_state.aio_stage == "processed":
    result = st.session_state.aio_result

    if len(result) == 4 and result[3]:  # multi-file zip
        zip_bytes, fname, info, _ = result
        st.success(f"Generated ZIP with multiple reports")
        st.download_button(
            "📥 Download All Formatted Reports (ZIP)",
            zip_bytes,
            fname,
            mime="application/zip",
            use_container_width=True
        )
    else:
        xlsx_bytes, fname, info = result
        st.success(f"✅ Report generated: **{fname}**")

        # === ANALYTICS DASHBOARD (pre-download) ===
        try:
            from io import BytesIO
            from openpyxl import load_workbook

            wb_preview = load_workbook(BytesIO(xlsx_bytes), data_only=False)

            st.subheader("📊 Quick Analytics — See the damage before you download")

            total_sheets = len(wb_preview.sheetnames)
            data_rows = 0
            tab_counts = {}

            for sheet_name in wb_preview.sheetnames:
                ws = wb_preview[sheet_name]
                row_count = max(0, ws.max_row - 1)
                tab_counts[sheet_name] = row_count
                if sheet_name.lower() != "summary":
                    data_rows += row_count

            col1, col2, col3 = st.columns(3)
            col1.metric("Total Data Rows", f"{data_rows:,}")
            col2.metric("Tabs in Report", total_sheets)
            col3.metric("Biggest Tab", max(tab_counts, key=tab_counts.get) if tab_counts else "N/A")

            st.markdown("**Issues by Category**")
            display_counts = {k: v for k, v in tab_counts.items() if v > 0 and k.lower() != "summary"}

            if display_counts:
                import pandas as pd
                df_counts = pd.DataFrame(list(display_counts.items()), columns=["Tab", "Rows"]).sort_values("Rows", ascending=False)
                st.dataframe(df_counts, use_container_width=True, hide_index=True)
                st.bar_chart(df_counts.set_index("Tab")["Rows"])
            else:
                st.info("No significant data tabs found.")

            if "Summary" in wb_preview.sheetnames:
                ws_sum = wb_preview["Summary"]
                st.markdown("**Summary Tab Preview**")
                summary_data = []
                for row in ws_sum.iter_rows(min_row=1, max_row=min(12, ws_sum.max_row), values_only=True):
                    summary_data.append([str(c) if c is not None else "" for c in row])
                if summary_data:
                    st.table(summary_data)

        except Exception as preview_err:
            st.warning(f"Could not generate full preview analytics: {preview_err}")

        st.download_button(
            "📥 Download Formatted Report",
            xlsx_bytes,
            fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    if st.button("← Start New Analysis"):
        for key in list(st.session_state.keys()):
            if key.startswith("aio_"):
                del st.session_state[key]
        st.rerun()

# =============================================================================
# Sidebar help
# =============================================================================
with st.sidebar:
    st.header("How the AIO Works")
    st.markdown("""
    1. You upload raw validation files + the cutsheet.
    2. The AIO **inspects the cutsheet** (not just the raw files).
    3. It extracts:
       - Which hall (JPB19, JBP15, SYD20...)
       - Fabric type (GFAB / CFAB / HOPS / LV Portal...)
       - Source (Slack export vs LVV Portal)
    4. It automatically picks the correct rich formatter we built.
    5. You get the properly formatted, actionable report.

    The cutsheet is the single source of truth.  
    This is why we can support **any room**.
    """)

    st.divider()
    st.caption("Currently supports: GFAB, HOPS, CFAB, IPR (more coming as we add halls)")

st.caption("All In One • Master logic from all cleaned formatters • Old scripts in legacy/")

st.markdown("---")
st.caption("Have an idea to make this even better? Go back to the main **app.py** page and use the **💡 Suggest a Feature** section at the bottom. It creates a ready-to-submit GitHub issue for you.")
