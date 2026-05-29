"""
jpb19_cfab_logic.py

Full clean port of CFAB_CODE19NEW.py for JPB19.

Implements the complete CFAB rack validation formatter:
  1. Split LLDP into Downlink + Mismatch
  2. Column cleanup on data sheets
  3. Enrich with cutsheet (Hop 1-6 / PP columns + peer info)
  4. Split every tab into T3-T2 (long-path) vs T2-T1-T0 (short-path)
  5. Trim unused Hop columns from short tabs
  6. Add "Current ..." B-side columns on Mismatch tabs + pink fill
  7. Grey-out Optic rows that match Downlink
  8. Add NOTE column everywhere
  9. Final styling (yellow headers, borders, autofilter, freezes)
 10. Rebuild Summary with formulas
 11. Desired tab ordering
 12. Reciprocal mismatch pair highlighting (orange/yellow + pink preservation)

Public API:
    process_cfab_validation(input_path, cutsheet_path) -> (bytes, suggested_filename)
    process_multiple_cfab_files(input_paths, cutsheet_path) -> zip_bytes
"""

from __future__ import annotations

import io
import os
import zipfile
from copy import copy
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Any

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter


# =============================================================================
# STYLES
# =============================================================================

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
PINK   = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
ORANGE = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)
LIGHT_GREY = "A6A6A6"

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)

HOP_HEADERS = ["Device Rack U", "Hop 1", "Hop 2", "Hop 3", "Hop 4", "Hop 5", "Hop 6"]


# =============================================================================
# CUTSHEET LOOKUP (long vs short path)
# =============================================================================

def build_cutsheet_lookup(path: str) -> Dict[str, Dict[str, Any]]:
    """
    Build lookup keyed by 'device port'.
    Registers both ends of every cable (long-path through 4 PPs or short/direct).
    """
    df = pd.read_excel(path, sheet_name=0, header=None)
    lookup: Dict[str, Dict[str, Any]] = {}

    def add(key, device_rack, pp_list, peer_device, peer_rack):
        if not key or key in lookup:
            return
        lookup[key] = {
            "device_rack": device_rack,
            "pp1": pp_list[0] if len(pp_list) > 0 else None,
            "pp2": pp_list[1] if len(pp_list) > 1 else None,
            "pp3": pp_list[2] if len(pp_list) > 2 else None,
            "pp4": pp_list[3] if len(pp_list) > 3 else None,
            "peer_device": peer_device,
            "peer_rack": peer_rack,
        }

    for _, row in df.iterrows():
        is_long = pd.notna(row[4])
        if is_long:
            add(str(row[0]) if pd.notna(row[0]) else None,
                row[1], [row[2], row[3], row[4], row[5]], row[6], row[7])
            add(str(row[6]) if pd.notna(row[6]) else None,
                row[7], [row[5], row[4], row[3], row[2]], row[0], row[1])
        else:
            add(str(row[0]) if pd.notna(row[0]) else None,
                row[1], [], row[2], row[3])
            add(str(row[2]) if pd.notna(row[2]) else None,
                row[3], [], row[0], row[1])
    return lookup


# =============================================================================
# HELPERS
# =============================================================================

def style_cell(cell, ref):
    cell.font = copy(ref.font)
    cell.fill = copy(ref.fill)
    cell.border = copy(ref.border)
    cell.alignment = copy(ref.alignment)
    cell.number_format = ref.number_format


def header_indices(ws):
    return {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value}


def delete_columns_by_name(ws, names: list[str]):
    for name in names:
        hdrs = [c.value for c in ws[1]]
        if name in hdrs:
            ws.delete_cols(hdrs.index(name) + 1)


def rack_number_from(ws) -> str:
    hdrs = header_indices(ws)
    if "Device A Rack" not in hdrs:
        return "output"
    col = hdrs["Device A Rack"]
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v:
            return str(v).split(":")[0].strip()
    return "output"


def autofit_columns(ws, min_w=12, max_w=40):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, min_w), max_w)


# =============================================================================
# STEP FUNCTIONS (exact behaviour from original)
# =============================================================================

def split_lldp_sheet(wb):
    src_name = "LLDP Mismatch + Link Down"
    if src_name not in wb.sheetnames:
        return
    src = wb[src_name]
    hdrs = [c.value for c in src[1]]
    status_idx = hdrs.index("LLDP Status")

    header_styles = [{
        "font": copy(c.font), "fill": copy(c.fill),
        "border": copy(c.border), "alignment": copy(c.alignment),
        "number_format": c.number_format,
    } for c in src[1]]
    col_widths = {k: v.width for k, v in src.column_dimensions.items()}

    down, mismatch = [], []
    for row in src.iter_rows(min_row=2, values_only=True):
        if row[status_idx] == "DOWN":
            down.append(row)
        elif row[status_idx] == "MISMATCH":
            mismatch.append(row)

    def build(name, rows):
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        ws.append(hdrs)
        for i, c in enumerate(ws[1], start=1):
            s = header_styles[i - 1]
            c.font = s["font"]; c.fill = s["fill"]; c.border = s["border"]
            c.alignment = s["alignment"]; c.number_format = s["number_format"]
        for r in rows:
            ws.append(r)
        for k, w in col_widths.items():
            if w:
                ws.column_dimensions[k].width = w
        ws.freeze_panes = "A2"

    build("Downlink", down)
    build("Mismatch", mismatch)
    del wb[src_name]


def clean_columns(wb):
    if "Downlink" in wb.sheetnames:
        delete_columns_by_name(wb["Downlink"], [
            "Device A Rack", "Expected Device B Rack", "Expected Device B Name",
            "Expected Device B Port", "Device B Rack", "Device B Name",
            "Device B Port", "LLDP Status", "Patch Panel Matrix",
        ])

    if "Mismatch" in wb.sheetnames:
        delete_columns_by_name(wb["Mismatch"], [
            "Device A Rack", "Expected Device B Rack", "Expected Device B Name",
            "Expected Device B Port", "Device B Rack",
            "LLDP Status", "Patch Panel Matrix",
        ])

    if "Optic Errors" in wb.sheetnames:
        ws = wb["Optic Errors"]
        delete_columns_by_name(ws, [
            "Remote Device Name", "Remote Device Port", "Patch Panel Matrix",
        ])
        hdrs = [c.value for c in ws[1]]
        if "Rx Power" in hdrs and hdrs.index("Rx Power") != 0:
            rx_idx = hdrs.index("Rx Power")
            new_order = [rx_idx] + [i for i in range(len(hdrs)) if i != rx_idx]
            data = [[c.value for c in row] for row in ws.iter_rows()]
            styles = [[{
                "font": copy(c.font), "fill": copy(c.fill),
                "border": copy(c.border), "alignment": copy(c.alignment),
                "number_format": c.number_format,
            } for c in row] for row in ws.iter_rows()]
            old_widths = [ws.column_dimensions[get_column_letter(i + 1)].width
                          for i in range(len(hdrs))]
            ws.delete_rows(1, ws.max_row)
            for r_i, (row_vals, row_styles) in enumerate(zip(data, styles), start=1):
                for new_c, old_c in enumerate(new_order, start=1):
                    cell = ws.cell(row=r_i, column=new_c, value=row_vals[old_c])
                    st = row_styles[old_c]
                    cell.font = st["font"]; cell.fill = st["fill"]
                    cell.border = st["border"]; cell.alignment = st["alignment"]
                    cell.number_format = st["number_format"]
            for new_c, old_c in enumerate(new_order, start=1):
                w = old_widths[old_c]
                if w:
                    ws.column_dimensions[get_column_letter(new_c)].width = w

    if "Interface Down Errors" in wb.sheetnames:
        delete_columns_by_name(wb["Interface Down Errors"], [
            "Source Device Location", "Remote Device Name", "Remote Device Port",
            "Issue", "Patch Panel Matrix",
        ])


def enrich(ws, name_col_idx, port_col_idx, insert_after_idx, lookup):
    header_ref = ws.cell(row=1, column=1)
    data_ref = ws.cell(row=2, column=1) if ws.max_row >= 2 else header_ref

    new_data = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=name_col_idx).value
        port = ws.cell(row=r, column=port_col_idx).value
        key = f"{name} {port}" if name and port else None
        info = lookup.get(key) if key else None
        if info:
            if info["pp1"] is not None:
                new_data.append([
                    info["device_rack"], info["pp1"], info["pp2"], info["pp3"],
                    info["pp4"], info["peer_device"], info["peer_rack"],
                ])
            else:
                new_data.append([info["device_rack"], info["peer_device"],
                                 info["peer_rack"], None, None, None, None])
        else:
            new_data.append([None] * 7)

    for i in range(7):
        ws.insert_cols(insert_after_idx + 1 + i)
    for i, h in enumerate(HOP_HEADERS):
        c = ws.cell(row=1, column=insert_after_idx + 1 + i, value=h)
        style_cell(c, header_ref)
    for r, vals in enumerate(new_data, start=2):
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=insert_after_idx + 1 + i)
            c.value = v
            style_cell(c, data_ref)


def enrich_all(wb, lookup):
    if "Downlink" in wb.sheetnames:
        enrich(wb["Downlink"], 1, 2, 2, lookup)
    if "Optic Errors" in wb.sheetnames:
        enrich(wb["Optic Errors"], 2, 3, 3, lookup)
    if "Mismatch" in wb.sheetnames:
        enrich(wb["Mismatch"], 1, 2, 2, lookup)
    if "Interface Down Errors" in wb.sheetnames:
        enrich(wb["Interface Down Errors"], 1, 2, 2, lookup)


SPLIT_KEY_HEADERS = {
    "Downlink":              ("Device A Name",      "Device A Port"),
    "Mismatch":              ("Device A Name",      "Device A Port"),
    "Optic Errors":          ("Source Device Name", "Source Device Port"),
    "Interface Down Errors": ("Source Device Name", "Source Device Port"),
}


def split_long_short(wb, lookup):
    for src_name in ["Downlink", "Mismatch", "Optic Errors", "Interface Down Errors"]:
        if src_name not in wb.sheetnames:
            continue
        src = wb[src_name]
        headers = [c.value for c in src[1]]
        name_hdr, port_hdr = SPLIT_KEY_HEADERS[src_name]
        name_col = headers.index(name_hdr) + 1
        port_col = headers.index(port_hdr) + 1
        header_styles = [{
            "font": copy(c.font), "fill": copy(c.fill),
            "border": copy(c.border), "alignment": copy(c.alignment),
            "number_format": c.number_format,
        } for c in src[1]]
        data_ref = src.cell(row=2, column=1) if src.max_row >= 2 else src.cell(row=1, column=1)
        col_widths = {k: v.width for k, v in src.column_dimensions.items()}

        long_rows, short_rows = [], []
        for r in range(2, src.max_row + 1):
            row_vals = [src.cell(row=r, column=c).value for c in range(1, src.max_column + 1)]
            name = row_vals[name_col - 1]
            port = row_vals[port_col - 1]
            key = f"{name} {port}" if name and port else None
            info = lookup.get(key) if key else None
            if info and info.get("pp1") is not None:
                long_rows.append(row_vals)
            else:
                short_rows.append(row_vals)

        def build(name, rows):
            if name in wb.sheetnames:
                del wb[name]
            ws = wb.create_sheet(name)
            ws.append(headers)
            for i, c in enumerate(ws[1], start=1):
                s = header_styles[i - 1]
                c.font = s["font"]; c.fill = s["fill"]; c.border = s["border"]
                c.alignment = s["alignment"]; c.number_format = s["number_format"]
            for row_vals in rows:
                ws.append(row_vals)
                for c in ws[ws.max_row]:
                    style_cell(c, data_ref)
            for k, w in col_widths.items():
                if w:
                    ws.column_dimensions[k].width = w
            ws.freeze_panes = "A2"

        build(f"T3-T2 {src_name}", long_rows)
        build(f"T2-T1-T0 {src_name}", short_rows)
        del wb[src_name]


def trim_short_tabs(wb):
    for sn in ["T2-T1-T0 Downlink", "T2-T1-T0 Mismatch", "T2-T1-T0 Optic Errors",
               "T2-T1-T0 Interface Down Errors"]:
        if sn in wb.sheetnames:
            delete_columns_by_name(wb[sn], ["Hop 3", "Hop 4", "Hop 5", "Hop 6"])


def enrich_mismatch_b_side(wb, lookup):
    for sn, b_headers, expect_long in [
        ("T3-T2 Mismatch",
         ["B Rack U", "B Hop 1", "B Hop 2", "B Hop 3", "B Hop 4", "B Hop 5", "B Hop 6"],
         True),
        ("T2-T1-T0 Mismatch",
         ["B Rack U", "B Hop 1", "B Hop 2"],
         False),
    ]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        header_ref = ws.cell(row=1, column=1)
        data_ref = ws.cell(row=2, column=1) if ws.max_row >= 2 else header_ref

        rename_map = {"Device Rack U": "A Rack U"}
        for i in range(1, 7):
            rename_map[f"Hop {i}"] = f"A Hop {i}"
        for c in ws[1]:
            if c.value in rename_map:
                c.value = rename_map[c.value]

        start_col = ws.max_column + 1
        for i, h in enumerate(b_headers):
            c = ws.cell(row=1, column=start_col + i, value=h)
            style_cell(c, header_ref)

        hdrs = [c.value for c in ws[1]]
        bname_col = hdrs.index("Device B Name") + 1
        bport_col = hdrs.index("Device B Port") + 1
        slots = len(b_headers)

        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=bname_col).value
            port = ws.cell(row=r, column=bport_col).value
            key = f"{name} {port}" if name and port else None
            info = lookup.get(key) if key else None
            vals = [None] * slots
            if info:
                if info["pp1"] is not None and expect_long:
                    vals = [info["device_rack"], info["pp1"], info["pp2"], info["pp3"],
                            info["pp4"], info["peer_device"], info["peer_rack"]]
                else:
                    short_vals = [info["device_rack"], info["peer_device"], info["peer_rack"]]
                    vals = short_vals + [None] * (slots - len(short_vals))
            for i, v in enumerate(vals):
                c = ws.cell(row=r, column=start_col + i)
                c.value = v
                style_cell(c, data_ref)

        hdrs = [c.value for c in ws[1]]
        start_pink = hdrs.index("Device B Name") + 1
        for col in range(start_pink, ws.max_column + 1):
            cur_h = ws.cell(row=1, column=col).value
            if cur_h and not str(cur_h).lower().startswith("current"):
                ws.cell(row=1, column=col).value = f"Current {cur_h}"
            for r in range(1, ws.max_row + 1):
                ws.cell(row=r, column=col).fill = PINK


def grey_optics_matching_downlink(wb):
    keys = set()
    for sn in ["T3-T2 Downlink", "T2-T1-T0 Downlink"]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        hdrs = header_indices(ws)
        n_col, p_col, r_col = hdrs.get("Device A Name"), hdrs.get("Device A Port"), hdrs.get("Device Rack U")
        if not (n_col and p_col):
            continue
        for r in range(2, ws.max_row + 1):
            n = ws.cell(row=r, column=n_col).value
            p = ws.cell(row=r, column=p_col).value
            ru = ws.cell(row=r, column=r_col).value if r_col else None
            if n and p:
                keys.add((str(n).strip(), str(p).strip(),
                          str(ru).strip() if ru else None))

    for sn in ["T3-T2 Optic Errors", "T2-T1-T0 Optic Errors"]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        hdrs = header_indices(ws)
        n_col = hdrs.get("Source Device Name")
        p_col = hdrs.get("Source Device Port")
        r_col = hdrs.get("Device Rack U")
        if not (n_col and p_col):
            continue
        for r in range(2, ws.max_row + 1):
            n = ws.cell(row=r, column=n_col).value
            p = ws.cell(row=r, column=p_col).value
            ru = ws.cell(row=r, column=r_col).value if r_col else None
            key = (str(n).strip() if n else None,
                   str(p).strip() if p else None,
                   str(ru).strip() if ru else None)
            if n and p and key in keys:
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    f = copy(cell.font)
                    f.color = LIGHT_GREY
                    cell.font = f


def add_note_column(wb):
    for sn in wb.sheetnames:
        if sn == "Summary":
            continue
        ws = wb[sn]
        ref = ws.cell(row=1, column=1)
        col = ws.max_column + 1
        c = ws.cell(row=1, column=col, value="Note")
        f = copy(ref.font); f.bold = True; c.font = f
        c.alignment = copy(ref.alignment)
        c.border = copy(ref.border)
        c.fill = YELLOW
        ws.column_dimensions[get_column_letter(col)].width = 30


def finalize_styling(wb):
    for sn in wb.sheetnames:
        ws = wb[sn]
        for c in ws[1]:
            f = copy(c.font); f.bold = True; f.size = 11; c.font = f
            c.fill = YELLOW
        if sn in ("T3-T2 Optic Errors", "T2-T1-T0 Optic Errors"):
            ws.freeze_panes = "B2"
        else:
            ws.freeze_panes = "A2"
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).border = BORDER
        if sn != "Summary":
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


def rebuild_summary(wb):
    if "Summary" not in wb.sheetnames:
        s = wb.create_sheet("Summary", 0)
        s.cell(row=1, column=1, value="Error Category")
        s.cell(row=1, column=2, value="Error Count")
    s = wb["Summary"]

    for r in range(s.max_row, 1, -1):
        s.delete_rows(r)

    rows = [
        ("T3-T2 Downlink",                 "=COUNTA('T3-T2 Downlink'!A:A)-1"),
        ("T2-T1-T0 Downlink",              "=COUNTA('T2-T1-T0 Downlink'!A:A)-1"),
        ("T3-T2 Mismatch",                 "=COUNTA('T3-T2 Mismatch'!A:A)-1"),
        ("T2-T1-T0 Mismatch",              "=COUNTA('T2-T1-T0 Mismatch'!A:A)-1"),
        ("T3-T2 Optic Errors",             "=COUNTA('T3-T2 Optic Errors'!A:A)-1"),
        ("T2-T1-T0 Optic Errors",          "=COUNTA('T2-T1-T0 Optic Errors'!A:A)-1"),
        ("T3-T2 Interface Down Errors",    "=COUNTA('T3-T2 Interface Down Errors'!A:A)-1"),
        ("T2-T1-T0 Interface Down Errors", "=COUNTA('T2-T1-T0 Interface Down Errors'!A:A)-1"),
    ]
    rows = [(label, formula) for (label, formula) in rows if label in wb.sheetnames]

    data_ref_cell = s.cell(row=1, column=1)
    no_fill = PatternFill(fill_type=None)

    for i, (label, formula) in enumerate(rows, start=2):
        c1 = s.cell(row=i, column=1, value=label)
        c2 = s.cell(row=i, column=2, value=formula)
        style_cell(c1, data_ref_cell)
        style_cell(c2, data_ref_cell)
        for c in (c1, c2):
            f = copy(c.font); f.bold = False; c.font = f
            c.fill = no_fill

    total_row = len(rows) + 2
    c1 = s.cell(row=total_row, column=1, value="Total")
    c2 = s.cell(row=total_row, column=2, value=f"=SUM(B2:B{len(rows) + 1})")
    style_cell(c1, data_ref_cell); style_cell(c2, data_ref_cell)
    for c in (c1, c2):
        f = copy(c.font); f.bold = True; c.font = f
        c.fill = no_fill

    s.column_dimensions["A"].width = 24
    s.column_dimensions["B"].width = 12


def reorder_tabs(wb):
    desired = [
        "Summary",
        "T3-T2 Downlink", "T2-T1-T0 Downlink",
        "T3-T2 Mismatch", "T2-T1-T0 Mismatch",
        "T3-T2 Optic Errors", "T2-T1-T0 Optic Errors",
        "T3-T2 Interface Down Errors", "T2-T1-T0 Interface Down Errors",
    ]
    wb._sheets = [wb[n] for n in desired if n in wb.sheetnames] + \
                 [wb[n] for n in wb.sheetnames if n not in desired]


def highlight_mismatch_pairs(wb):
    no_fill = PatternFill(fill_type=None)
    for sn in ("T3-T2 Mismatch", "T2-T1-T0 Mismatch"):
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        if ws.max_row < 3:
            continue

        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

        def col(name):
            return header.index(name) + 1 if name in header else None

        a_name = col("Device A Name")
        a_port = col("Device A Port")
        b_name = col("Current Device B Name")
        b_port = col("Current Device B Port")
        if not all((a_name, a_port, b_name, b_port)):
            continue

        cur_cols = [i + 1 for i, h in enumerate(header)
                    if isinstance(h, str) and h.startswith("Current")]
        start_pink = min(cur_cols) if cur_cols else ws.max_column + 1
        aside_cols = set(range(1, start_pink))
        pink_cols = set(cur_cols)

        ncol, nrow = ws.max_column, ws.max_row
        rows = [[ws.cell(row=r, column=c).value for c in range(1, ncol + 1)]
                for r in range(2, nrow + 1)]

        def normv(v):
            return str(v).strip() if v is not None else ""

        a_keys = [(normv(rv[a_name - 1]), normv(rv[a_port - 1])) for rv in rows]
        b_keys = [(normv(rv[b_name - 1]), normv(rv[b_port - 1])) for rv in rows]

        n = len(rows)
        partner = [None] * n
        for i in range(n):
            if partner[i] is not None or not any(a_keys[i]):
                continue
            for j in range(i + 1, n):
                if partner[j] is not None:
                    continue
                if a_keys[i] == b_keys[j] and a_keys[j] == b_keys[i]:
                    partner[i] = j
                    partner[j] = i
                    break

        placed = [False] * n
        order = []
        for i in range(n):
            if placed[i] or partner[i] is None:
                continue
            j = partner[i]
            order.append(i); placed[i] = True
            if not placed[j]:
                order.append(j); placed[j] = True
        for i in range(n):
            if not placed[i]:
                order.append(i); placed[i] = True

        paired = {i for i in range(n) if partner[i] is not None}
        pair_no, counter = {}, 0
        for i in range(n):
            j = partner[i]
            if j is not None and i < j:
                pair_no[i] = counter
                pair_no[j] = counter
                counter += 1

        for out_off, src_i in enumerate(order):
            r = out_off + 2
            rv = rows[src_i]
            is_pair = src_i in paired
            pair_fill = ORANGE if pair_no.get(src_i, 0) % 2 == 0 else YELLOW
            for c in range(1, ncol + 1):
                cell = ws.cell(row=r, column=c, value=rv[c - 1])
                if c in pink_cols:
                    cell.fill = PINK
                elif is_pair and c in aside_cols:
                    cell.fill = pair_fill
                else:
                    cell.fill = no_fill


# =============================================================================
# CORE PIPELINE
# =============================================================================

def _build_formatted_workbook(input_path: str, lookup: Dict[str, Dict[str, Any]]) -> Workbook:
    wb = load_workbook(input_path)

    # Derive rack early
    rack = "output"
    if "LLDP Mismatch + Link Down" in wb.sheetnames:
        rack = rack_number_from(wb["LLDP Mismatch + Link Down"])

    split_lldp_sheet(wb)
    clean_columns(wb)
    enrich_all(wb, lookup)
    split_long_short(wb, lookup)
    trim_short_tabs(wb)
    enrich_mismatch_b_side(wb, lookup)
    grey_optics_matching_downlink(wb)
    add_note_column(wb)
    finalize_styling(wb)
    rebuild_summary(wb)
    reorder_tabs(wb)
    highlight_mismatch_pairs(wb)

    for sn in wb.sheetnames:
        if sn != "Summary":
            autofit_columns(wb[sn])

    # Store rack for filename
    wb._cfab_rack = rack
    return wb


def _compute_output_name(wb: Workbook, input_path: str) -> str:
    rack = getattr(wb, "_cfab_rack", None) or "output"
    return f"{rack}.xlsx"


# =============================================================================
# PUBLIC API
# =============================================================================

def process_cfab_validation(
    input_path: str,
    cutsheet_path: str,
    output_name: Optional[str] = None
) -> Tuple[bytes, str]:
    lookup = build_cutsheet_lookup(cutsheet_path)
    wb = _build_formatted_workbook(input_path, lookup)

    if output_name is None:
        output_name = _compute_output_name(wb, input_path)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), output_name


def process_multiple_cfab_files(
    input_paths: List[str],
    cutsheet_path: str,
) -> bytes:
    lookup = build_cutsheet_lookup(cutsheet_path)
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        for i, path in enumerate(input_paths, 1):
            wb = _build_formatted_workbook(path, lookup)
            out_name = getattr(wb, "_cfab_rack", f"cfab_{i}") + ".xlsx"
            if out_name in zipf.namelist():
                stem, ext = os.path.splitext(out_name)
                out_name = f"{stem}_{i}{ext}"
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            zipf.writestr(out_name, buf.getvalue())
    zip_buffer.seek(0)
    return zip_buffer.getvalue()
