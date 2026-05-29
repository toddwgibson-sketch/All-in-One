"""
hops_logic.py

Clean, importable core logic for HOPS/GPU Validation Formatter.

This is the full rich port of HOPS CODE19.py.

All UI, Tkinter file pickers, and side effects have been removed.

ARCHITECTURAL RULE:
    No multiprocessing allowed. Required for Streamlit Cloud.
"""
The module only contains pure transformation logic.

Public API (use from Streamlit pages or tests):
    process_hops_validation(input_path, combined_cutsheet_path) -> (bytes, suggested_filename)
    process_multiple_hops_files(input_paths, combined_cutsheet_path) -> zip_bytes
"""

from __future__ import annotations

import io
import os
import re
import warnings
import zipfile
from collections import Counter
from copy import copy
from typing import Callable, List, Optional, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# WARNINGS (match original behaviour)
# =============================================================================
warnings.filterwarnings(
    "ignore",
    message=r".*no default style.*",
    category=UserWarning,
    module=r"openpyxl\..*",
)


# =============================================================================
# CONSTANTS & STYLES
# =============================================================================

PINK   = "FFC0CB"
YELLOW = "FFFF00"
ORANGE = "FFA500"
GREY   = "A6A6A6"

THIN = Side(border_style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
YELLOW_FILL = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
PINK_FILL   = PatternFill(start_color=PINK,   end_color=PINK,   fill_type="solid")
ORANGE_FILL = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
GREY_FILL   = PatternFill(start_color=GREY,   end_color=GREY,   fill_type="solid")
NO_FILL     = PatternFill(fill_type=None)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)

DEFAULT_SKIP_HEADERS = ("DeviceA", "Device A", "PortA", "Port A")

CATEGORIES = [
    "LLDP Mismatch  (GPU)",
    "Optic Errors (GPU)",
    "FEC_BER Errors (GPU)",
    "Interface Down Errors (GPU)",
]
CATEGORY_SHORT = ["LLDP Mismatch", "Optic Errors", "FEC_BER Errors", "Interface Down"]


# =============================================================================
# GENERIC HELPERS (preserved exactly)
# =============================================================================

def norm(v) -> str:
    return str(v).strip().lower() if v is not None else ""


def canon_slot(p: str) -> str:
    """slot.../port...-1 and -2 -> ...-1 ; -3 and -4 -> ...-3"""
    if not p:
        return ""
    s = norm(p)
    m = re.match(r"^(.*-)(\d+)$", s)
    if m:
        base, num = m.group(1), int(m.group(2))
        if num in (1, 2):
            return base + "1"
        if num in (3, 4):
            return base + "3"
    return s


def canon_swp(p: str) -> str:
    """swpNs0/s1 -> swpNs0 ; swpNs2/s3 -> swpNs2"""
    if not p:
        return ""
    s = norm(p)
    m = re.match(r"^(swp\d+s)(\d+)$", s)
    if m:
        base, num = m.group(1), int(m.group(2))
        return base + str(num - num % 2)
    return s


def snapshot_col(ws, col_idx):
    """Capture a column's values + full cell formatting + width."""
    data = []
    for r in range(1, ws.max_row + 1):
        src = ws.cell(row=r, column=col_idx)
        data.append({
            "value": src.value,
            "font": copy(src.font),
            "fill": copy(src.fill),
            "alignment": copy(src.alignment),
            "border": copy(src.border),
            "number_format": src.number_format,
        })
    width = ws.column_dimensions[get_column_letter(col_idx)].width
    return data, width


def write_col(ws, col_idx, data, width):
    for r, d in enumerate(data, start=1):
        c = ws.cell(row=r, column=col_idx, value=d["value"])
        c.font = d["font"]
        c.fill = d["fill"]
        c.alignment = d["alignment"]
        c.border = d["border"]
        c.number_format = d["number_format"]
    if width:
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def find_sheet(wb_or_names, *candidates):
    """Return first matching sheet name (case/whitespace tolerant)."""
    names = wb_or_names.sheetnames if hasattr(wb_or_names, "sheetnames") else list(wb_or_names)
    lookup = {re.sub(r"\s+", " ", s.strip().lower()): s for s in names}
    for c in candidates:
        key = re.sub(r"\s+", " ", c.strip().lower())
        if key in lookup:
            return lookup[key]
    return None


def _find_header_idx(header, name_options):
    for opt in name_options:
        key = re.sub(r"\s+", " ", str(opt).strip().lower())
        for i, h in enumerate(header):
            if h is None:
                continue
            if re.sub(r"\s+", " ", str(h).strip().lower()) == key:
                return i
    return None


# =============================================================================
# COMBINED CUTSHEET LOOKUPS (core of HOPS)
# =============================================================================

def build_combined_lookups(combined_path: str):
    """
    Build the 4 lookup structures used for all enrichment:
        (combined_header, exact_A, pair_A, exact_B, pair_B)
    Keys are (norm(Device), norm(Port)).
    """
    wb = load_workbook(combined_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = list(row)
        else:
            rows.append(row)

    a_dev_i  = _find_header_idx(header, ["DeviceA", "Device A"])
    a_port_i = _find_header_idx(header, ["PortA", "Port A"])
    b_dev_i  = _find_header_idx(header, ["DeviceB", "Device B"])
    b_port_i = _find_header_idx(header, ["PortB", "Port B"])

    if a_dev_i is None or a_port_i is None:
        raise ValueError(
            "Combined cutsheet is missing required A-side columns "
            "('DeviceA' and 'PortA'). Found headers: " + repr(header)
        )

    exact_A, pair_A, exact_B, pair_B = {}, {}, {}, {}

    for r in rows:
        if a_dev_i < len(r) and r[a_dev_i]:
            k = (norm(r[a_dev_i]), norm(r[a_port_i]) if a_port_i < len(r) else "")
            exact_A[k] = r
            pair_A[(k[0], canon_slot(k[1]))] = r

        if b_dev_i is not None and b_dev_i < len(r) and r[b_dev_i]:
            k = (norm(r[b_dev_i]), norm(r[b_port_i]) if b_port_i and b_port_i < len(r) else "")
            exact_B[k] = r
            pair_B[(k[0], canon_slot(k[1]))] = r

    return header, exact_A, pair_A, exact_B, pair_B


def populate(ws, key_name_col, key_port_col,
             exact_lookup, pair_lookup, canon_fn,
             combined_header, skip_headers=DEFAULT_SKIP_HEADERS,
             header_suffix=""):
    """Enrich worksheet from combined cutsheet using exact or canon pair match."""
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    name_idx = header.index(key_name_col) + 1
    port_idx = header.index(key_port_col) + 1

    skip_norm = {re.sub(r"\s+", " ", str(s).strip().lower()) for s in skip_headers}
    add_indices = [
        i for i, h in enumerate(combined_header)
        if h is not None and re.sub(r"\s+", " ", str(h).strip().lower()) not in skip_norm
    ]
    add_headers = [f"{combined_header[i]}{header_suffix}" for i in add_indices]

    start_col = ws.max_column + 1
    hstyle = get_header_style(ws)
    for off, h in enumerate(add_headers):
        c = ws.cell(row=1, column=start_col + off, value=h)
        c.font = hstyle["font"]
        c.fill = hstyle["fill"]
        c.alignment = hstyle["alignment"]
        c.border = hstyle["border"]

    matched_exact = matched_pair = unmatched = 0
    for r in range(2, ws.max_row + 1):
        nv = ws.cell(row=r, column=name_idx).value
        pv = ws.cell(row=r, column=port_idx).value
        nkey = norm(nv)
        match = exact_lookup.get((nkey, norm(pv)))
        used_pair = False
        if not match:
            match = pair_lookup.get((nkey, canon_fn(pv)))
            used_pair = bool(match)
        if match:
            if used_pair:
                matched_pair += 1
            else:
                matched_exact += 1
            for off, idx in enumerate(add_indices):
                val = match[idx] if idx < len(match) else None
                ws.cell(row=r, column=start_col + off, value=val)
        else:
            unmatched += 1
    return matched_exact, matched_pair, unmatched


def get_header_style(ws):
    """Return a copy of the first header row's styling (or sensible defaults)."""
    src = ws.cell(row=1, column=1)
    return {
        "font": copy(src.font) if src.font else Font(bold=True, name="Calibri", size=11),
        "fill": copy(src.fill) if src.fill else YELLOW_FILL,
        "alignment": copy(src.alignment) if src.alignment else CENTER,
        "border": copy(src.border) if src.border else THIN_BORDER,
    }


# =============================================================================
# "MAYBE CT OFF" MARKING
# =============================================================================

def mark_ct_off(ws, key_col_name="Remote Device Name",
                threshold=16, note_text="maybe CT off",
                grey_color="A6A6A6"):
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if key_col_name not in header:
        return 0
    key_idx = header.index(key_col_name) + 1

    if "Note" not in header:
        new_col = ws.max_column + 1
        src = ws.cell(row=1, column=1)
        cell = ws.cell(row=1, column=new_col, value="Note")
        cell.font = copy(src.font)
        cell.fill = copy(src.fill)
        cell.alignment = copy(src.alignment)
        cell.border = copy(src.border)
        ws.column_dimensions[get_column_letter(new_col)].width = 30
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    note_idx = header.index("Note") + 1

    counts = Counter()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=key_idx).value
        if v:
            counts[str(v).strip()] += 1
    targets = {k for k, c in counts.items() if c == threshold}

    marked = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=key_idx).value
        if v and str(v).strip() in targets:
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                f = cell.font
                cell.font = Font(
                    name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                    vertAlign=f.vertAlign, underline=f.underline, strike=f.strike,
                    color=grey_color, family=f.family, scheme=f.scheme
                )
            ws.cell(row=r, column=note_idx, value=note_text)
            marked += 1
    return marked


# =============================================================================
# PER-SHEET PROCESSORS (the heart of HOPS business logic)
# =============================================================================

def process_lldp(wb, combined):
    combined_header, exact_A, pair_A, exact_B, pair_B = combined
    sheet = find_sheet(wb, "LLDP Mismatch + Link Down (GPU)", "LLDP Mismatch  (GPU)", "LLDP Mismatch (GPU)")
    if not sheet:
        return None
    ws = wb[sheet]

    # Filter out junk rows (missing or non-compute)
    cols = {h: i + 1 for i, h in enumerate(ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1))}
    cn = cols.get("Current Device B Name")
    cl = cols.get("Current B Location")
    cp = cols.get("Current Device B Port")
    if cn and cl and cp:
        rows_to_drop = []
        for r in range(2, ws.max_row + 1):
            all_missing = all(norm(ws.cell(row=r, column=c).value) == "missing" for c in (cn, cl, cp))
            cn_val = ws.cell(row=r, column=cn).value
            not_compute = cn_val is not None and str(cn_val).strip() != "" and "compute" not in str(cn_val).lower()
            if all_missing or not_compute:
                rows_to_drop.append(r)
        for r in reversed(rows_to_drop):
            ws.delete_rows(r, 1)

    new_name = "LLDP Mismatch  (GPU)"
    if sheet != new_name:
        ws.title = new_name

    populate(ws, "Expected Device B Name", "Expected Device B Port",
             exact_A, pair_A, canon_slot, combined_header)

    to_remove = ["Device A Name", "Device A Location", "Device A Port",
                 "Current B Location", "Expected B Location",
                 "Patch Panel Matrix", "Error Message"]
    drop_cols_by_name(ws, to_remove)

    populate(ws, "Current Device B Name", "Current Device B Port",
             exact_A, pair_A, canon_slot, combined_header, header_suffix=" (Current)")

    # Reverse Current columns
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    current_cols = [h for h in header if "current" in str(h).lower()]
    reversed_current = list(reversed(current_cols))
    new_order = []
    ci = 0
    for h in header:
        if "current" in str(h).lower():
            new_order.append(reversed_current[ci]); ci += 1
        else:
            new_order.append(h)
    reorder_columns(ws, new_order)

    # Move white (non-Current) to front, pink Current to back
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    white = [h for h in header if "current" not in str(h).lower()]
    pink  = [h for h in header if "current" in str(h).lower()]
    reorder_columns(ws, white + pink)

    # Pink fill on Current columns
    pink_fill = PINK_FILL
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for ci, h in enumerate(header, start=1):
        if "current" in str(h).lower():
            for r in range(1, ws.max_row + 1):
                ws.cell(row=r, column=ci).fill = pink_fill

    return ws.title


def process_optic(wb, combined):
    combined_header, exact_A, pair_A, exact_B, pair_B = combined
    sheet = find_sheet(wb, "Optic Errors (GPU)")
    if not sheet:
        return None
    ws = wb[sheet]

    # Drop rows with Rx Power == "missing"
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if "Rx Power" in header:
        rx_idx = header.index("Rx Power") + 1
        rows_to_drop = []
        for r in range(2, ws.max_row + 1):
            if norm(ws.cell(row=r, column=rx_idx).value) == "missing":
                rows_to_drop.append(r)
        for r in reversed(rows_to_drop):
            ws.delete_rows(r, 1)

    populate(ws, "Device Name", "Port", exact_A, pair_A, canon_swp, combined_header)

    drop_cols_by_name(ws, ["Patch Panel Matrix", "Issue"])
    mark_ct_off(ws, key_col_name="Remote Device Name", threshold=16, note_text="maybe CT off")

    return ws.title


def process_fec(wb, combined):
    combined_header, exact_A, pair_A, exact_B, pair_B = combined
    sheet = find_sheet(wb, "FEC_BER Errors (GPU)")
    if not sheet:
        return None
    ws = wb[sheet]

    populate(ws, "Device Name", "Port", exact_A, pair_A, canon_swp, combined_header)
    drop_cols_by_name(ws, ["Patch Panel Matrix", "Issue"])
    mark_ct_off(ws, key_col_name="Remote Device Name", threshold=16, note_text="maybe CT off")

    return ws.title


def process_interface_down(wb, combined):
    combined_header, exact_A, pair_A, exact_B, pair_B = combined
    sheet = find_sheet(wb, "Interface Down Errors (GPU)")
    if not sheet:
        return None
    ws = wb[sheet]

    populate(ws, "Device Name", "Port", exact_A, pair_A, canon_swp, combined_header)
    drop_cols_by_name(ws, [
        "Issue", "Patch Panel Matrix"
    ])
    mark_ct_off(ws, key_col_name="Remote Device Name", threshold=16, note_text="maybe CT off")
    return ws.title


# =============================================================================
# COLUMN OPERATIONS
# =============================================================================

def drop_cols_by_name(ws, names):
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    indices = sorted([header.index(n) + 1 for n in names if n in header], reverse=True)
    for idx in indices:
        ws.delete_cols(idx, 1)


def reorder_columns(ws, new_order):
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    snapshots = {h: snapshot_col(ws, header.index(h) + 1) for h in header if h in header}
    ws.delete_cols(1, ws.max_column)
    for new_idx, name in enumerate(new_order, start=1):
        if name in snapshots:
            data, width = snapshots[name]
            write_col(ws, new_idx, data, width)


# =============================================================================
# FORMATTING & SUMMARY
# =============================================================================

def add_note_column(ws):
    new_col = ws.max_column + 1
    src = ws.cell(row=1, column=1)
    cell = ws.cell(row=1, column=new_col, value="Note")
    cell.font = copy(src.font)
    cell.fill = copy(src.fill)
    cell.alignment = copy(src.alignment)
    cell.border = copy(src.border)
    ws.column_dimensions[get_column_letter(new_col)].width = 30


def update_summary(wb):
    sheet = find_sheet(wb, "Summary")
    if not sheet:
        return
    ws = wb[sheet]
    existing_sheets = set(wb.sheetnames)

    name_map = {
        "lldp mismatch + link down (gpu)": find_sheet(wb, "LLDP Mismatch  (GPU)", "LLDP Mismatch (GPU)"),
        "lldp mismatch  (gpu)":            find_sheet(wb, "LLDP Mismatch  (GPU)", "LLDP Mismatch (GPU)"),
        "lldp mismatch (gpu)":             find_sheet(wb, "LLDP Mismatch  (GPU)", "LLDP Mismatch (GPU)"),
        "optic errors (gpu)":              find_sheet(wb, "Optic Errors (GPU)"),
        "fec_ber errors (gpu)":            find_sheet(wb, "FEC_BER Errors (GPU)"),
        "interface down errors (gpu)":     find_sheet(wb, "Interface Down Errors (GPU)"),
    }

    rows_keep = [(ws.cell(row=1, column=1).value, ws.cell(row=1, column=2).value)]
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(row=r, column=1).value
        target = name_map.get(norm(cat))
        if target and target in existing_sheets:
            rows_keep.append((target, f"=COUNTA('{target}'!A:A)-1"))

    if len(rows_keep) > 1:
        first_data_row = 2
        last_data_row = len(rows_keep)
        rows_keep.append(("Total", f"=SUM(B{first_data_row}:B{last_data_row})"))

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    for i, (a, b) in enumerate(rows_keep, start=1):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)

    if len(rows_keep) > 1:
        total_row = len(rows_keep)
        for c in (1, 2):
            cell = ws.cell(row=total_row, column=c)
            f = cell.font
            cell.font = Font(
                name=f.name, size=f.size, bold=True, italic=f.italic,
                vertAlign=f.vertAlign, underline=f.underline, strike=f.strike,
                color=f.color, family=f.family, scheme=f.scheme
            )

    for r in range(ws.max_row, len(rows_keep), -1):
        ws.delete_rows(r, 1)


def apply_workbook_formatting(wb):
    center = CENTER
    border = THIN_BORDER

    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row == 0 or ws.max_column == 0:
            continue

        if sn.lower() != "summary":
            existing = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            if "Note" not in existing:
                add_note_column(ws)

        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = center
                cell.border = border

        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = YELLOW_FILL
            f = cell.font
            size = f.size if f.size else 11
            cell.font = Font(
                name=f.name, size=size + 1, bold=True, italic=f.italic,
                vertAlign=f.vertAlign, underline=f.underline, strike=f.strike,
                color=f.color, family=f.family, scheme=f.scheme
            )

        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

        for r in range(1, ws.max_row + 1):
            ws.row_dimensions[r].height = 22

        for c in range(1, ws.max_column + 1):
            col_letter = get_column_letter(c)
            max_len = max((len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, ws.max_row + 1)), default=10)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

        ws.freeze_panes = "A2"

    # Special freeze for Optic Errors
    s = find_sheet(wb, "Optic Errors (GPU)")
    if s:
        wb[s].freeze_panes = "B2"


# =============================================================================
# GPU-SPECIFIC MISMATCH PAIR HIGHLIGHTING
# =============================================================================

def highlight_mismatch_pairs(wb):
    sheet = find_sheet(wb, "LLDP Mismatch  (GPU)", "LLDP Mismatch (GPU)",
                       "LLDP Mismatch + Link Down (GPU)")
    if not sheet:
        return
    ws = wb[sheet]
    if ws.max_row < 3:
        return

    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    def col(name):
        return header.index(name) + 1 if name in header else None

    exp_name = col("Expected Device B Name")
    exp_port = col("Expected Device B Port")
    cur_name = col("Current Device B Name")
    cur_port = col("Current Device B Port")
    if not all((exp_name, exp_port, cur_name, cur_port)):
        return

    pink_cols = [i + 1 for i, h in enumerate(header) if "current" in str(h).lower()]
    start_pink = min(pink_cols) if pink_cols else ws.max_column + 1
    white_cols = set(range(1, start_pink))
    pink_set = set(pink_cols)

    ncol, nrow = ws.max_column, ws.max_row
    rows = [[ws.cell(row=r, column=c).value for c in range(1, ncol + 1)] for r in range(2, nrow + 1)]

    exp_keys = [(norm(rv[exp_name - 1]), norm(rv[exp_port - 1])) for rv in rows]
    cur_keys = [(norm(rv[cur_name - 1]), norm(rv[cur_port - 1])) for rv in rows]

    n = len(rows)
    partner = [None] * n
    for i in range(n):
        if partner[i] is not None or not any(exp_keys[i]):
            continue
        for j in range(i + 1, n):
            if partner[j] is not None:
                continue
            if cur_keys[i] == exp_keys[j] and cur_keys[j] == exp_keys[i]:
                partner[i] = j
                partner[j] = i
                break

    placed = [False] * n
    order = []
    for i in range(n):
        if placed[i] or partner[i] is None:
            continue
        j = partner[i]
        order.append(i); placed[i] = True
        if not placed[j]:
            order.append(j); placed[j] = True
    for i in range(n):
        if not placed[i]:
            order.append(i); placed[i] = True

    paired = {i for i in range(n) if partner[i] is not None}

    pair_no, counter = {}, 0
    for i in range(n):
        j = partner[i]
        if j is not None and i < j:
            pair_no[i] = counter
            pair_no[j] = counter
            counter += 1

    orange = ORANGE_FILL
    yellow = YELLOW_FILL
    pink = PINK_FILL
    nofill = NO_FILL

    for out_off, src_i in enumerate(order):
        r = out_off + 2
        rv = rows[src_i]
        is_pair = src_i in paired
        pair_fill = orange if pair_no.get(src_i, 0) % 2 == 0 else yellow
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c, value=rv[c - 1])
            if c in pink_set:
                cell.fill = pink
            elif is_pair and c in white_cols:
                cell.fill = pair_fill
            else:
                cell.fill = nofill


# =============================================================================
# RACK-BASED OUTPUT NAME
# =============================================================================

def extract_rack_number(wb):
    racks = set()
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, min(ws.max_row + 1, 50)):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(row=r, column=c).value or "")
                m = re.search(r"\b(\d{3,4})\b", val)
                if m:
                    racks.add(m.group(1))
    return sorted(racks)


# =============================================================================
# CORE PIPELINE (in-memory)
# =============================================================================

def _build_formatted_workbook(input_path: str, combined) -> Workbook:
    wb = load_workbook(input_path)

    process_lldp(wb, combined)
    process_optic(wb, combined)
    process_fec(wb, combined)
    process_interface_down(wb, combined)

    update_summary(wb)
    apply_workbook_formatting(wb)
    highlight_mismatch_pairs(wb)

    return wb


def _compute_output_name(wb: Workbook, input_path: str) -> str:
    racks = extract_rack_number(wb)
    if racks:
        base = "-".join(racks)
    else:
        base = os.path.splitext(os.path.basename(input_path))[0]
    return base + ".xlsx"


# =============================================================================
# PUBLIC API
# =============================================================================

def process_hops_validation(
    input_path: str,
    combined_cutsheet_path: str,
    output_name: Optional[str] = None
) -> Tuple[bytes, str]:
    """
    Process one JPB19 HOPS validation file against the combined cutsheet.
    Returns (xlsx_bytes, suggested_filename).
    """
    combined = build_combined_lookups(combined_cutsheet_path)
    wb = _build_formatted_workbook(input_path, combined)

    if output_name is None:
        output_name = _compute_output_name(wb, input_path)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), output_name


def process_multiple_hops_files(
    input_paths: List[str],
    combined_cutsheet_path: str,
) -> bytes:
    """Process several files and return a ZIP of the formatted outputs."""
    combined = build_combined_lookups(combined_cutsheet_path)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        for i, path in enumerate(input_paths, 1):
            wb = _build_formatted_workbook(path, combined)
            out_name = _compute_output_name(wb, path)
            if out_name in zipf.namelist():
                stem, ext = os.path.splitext(out_name)
                out_name = f"{stem}_{i}{ext}"

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            zipf.writestr(out_name, buf.getvalue())

    zip_buffer.seek(0)
    return zip_buffer.getvalue()


__all__ = [
    "process_hops_validation",
    "process_multiple_hops_files",
    "build_combined_lookups",
]
