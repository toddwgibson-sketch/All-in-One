"""
jpb19_gfab_logic.py

Clean, importable core logic for the JPB19 GFAB Validation Formatter.

This module contains ZERO UI code (no Tkinter, no Streamlit, no file dialogs).
It is designed to be called from the Streamlit page (or tests, notebooks, etc.).

All original GFAB formatting behaviour is preserved:
  - lldp split → Downlinks + Mismatches
  - Optics cleanup + column reordering
  - L/R, cutsheet fill columns (Source_port etc.)
  - Z-side enrichment for Optics/combined_fec
  - Pink Possible + Active Z columns in Mismatches
  - Global column stripping
  - Summary tab (by Rack)
  - Grey-out matched Optics rows
  - NOTE columns + autofilter
  - Reciprocal mismatch pair highlighting (orange/yellow)
  - Final autofit + centre align + borders (pink preserved on Mismatches)

Public API:
    process_gfab_validation(input_path, cutsheet_path) -> (bytes, suggested_filename)
    process_multiple_files(input_paths, cutsheet_path) -> zip_bytes
"""

from __future__ import annotations

import io
import os
import zipfile
from collections import Counter
from typing import Callable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# CONSTANTS (from original GFAB CODE19.py + extensions)
# =============================================================================

# Tab-name resolution: the source report generator periodically renames tabs.
TAB_ALIASES = {
    'lldp':         ('lldp_sp', 'full_path_lldp_with_int_down'),
    'optics':       ('optics_rx_tx_threshold', 'optics_rx_tx_threshold_with_pp'),
    'interfaces':   ('interfaces_sp', 'interfaces_sp_with_pp'),
    'combined_fec': ('combined_fec', 'combined_fec_with_pp'),
}

# Tabs that should always be dropped (case-insensitive).
TABS_TO_REMOVE = (
    'device_reporting_failure',
    'bgp_sp',
    'spectrum_health',
    'sp_power',
    'sp_fans',
    'optics_temp',
    'pre_fec_ber_threshold_with_pp',
)

# Columns to strip globally from every tab in the final output.
COLUMNS_TO_REMOVE = (
    'Building',
    'Act. Building',
    'Exp. Building',
    'PP_A',
    'PP_Z',
    'Remote Host',
    'Remote Interface',
    'Mapped Remote Host',
    'Mapped Remote Interface',
    'Mapped Remote Rack',
    'Mapped Remote Elevation',
    'Remote Host Match',
    'Remote Interface Match',
    'Remote End Match',
    'Z_end_host',
    'Z_end_intf',
    'rack_z',
    'Z_Rack',
    'Z_Elevation',
    'Index',
    'Source Sheet',
    'Placement Group',
)

# Tabs that receive Z-side columns (Z Hostname, Z Interface, etc.).
Z_FILL_TABS = ('Optics', 'combined_fec')

# Colours
PINK   = "FFB6C1"
YELLOW = "FFFF00"
ORANGE = "FFA500"

# Styles
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

PINK_FILL   = PatternFill(start_color=PINK,   end_color=PINK,   fill_type="solid")
YELLOW_FILL = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
ORANGE_FILL = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
NO_FILL     = PatternFill(fill_type=None)


# =============================================================================
# STYLE HELPERS
# =============================================================================

def thin_border() -> Border:
    return THIN_BORDER


def header_fill() -> PatternFill:
    return YELLOW_FILL


def pink_fill() -> PatternFill:
    return PINK_FILL


def header_cell(cell, value: str, fill: Optional[PatternFill] = None) -> None:
    cell.value     = value
    cell.font      = Font(bold=True, name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.border    = thin_border()
    cell.fill      = fill if fill else YELLOW_FILL


def clear_and_border(ws, pink_cols: Optional[set] = None) -> None:
    """Remove all fills except pink columns; yellow header row; black borders everywhere."""
    pink_cols = pink_cols or set()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row == 1:
                cell.fill = YELLOW_FILL
            elif cell.column in pink_cols:
                cell.fill = PINK_FILL
            else:
                cell.fill = NO_FILL
            cell.border = thin_border()
            if cell.font:
                cell.font = Font(
                    bold=cell.font.bold,
                    name=cell.font.name or "Arial",
                    size=cell.font.size or 10,
                    color="FF000000",
                )


def autofit_sheet(ws, header_row_height: int = 24, data_row_height: int = 20, max_col_width: int = 80) -> None:
    """Autofit columns (capped) and set comfortable row heights. Handles merged cells."""
    merged = set()
    for mrange in ws.merged_cells.ranges:
        for r in range(mrange.min_row, mrange.max_row + 1):
            for c in range(mrange.min_col, mrange.max_col + 1):
                merged.add((r, c))

    col_max = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue
            if (cell.row, cell.column) in merged:
                continue
            longest = max((len(line) for line in str(cell.value).splitlines()), default=0)
            letter = get_column_letter(cell.column)
            if longest > col_max.get(letter, 0):
                col_max[letter] = longest

    for letter, length in col_max.items():
        ws.column_dimensions[letter].width = min(length + 4, max_col_width)

    if ws.max_row >= 1:
        ws.row_dimensions[1].height = header_row_height
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = data_row_height


def write_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    """Create a new sheet from a DataFrame with header styling + borders."""
    ws = wb.create_sheet(name)
    for c, col in enumerate(df.columns, 1):
        header_cell(ws.cell(row=1, column=c), col)
    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, col in enumerate(df.columns, 1):
            val = row[col]
            cell = ws.cell(row=r, column=c, value=None if pd.isna(val) else val)
            cell.border = thin_border()
    for c, col in enumerate(df.columns, 1):
        mx = max([len(str(col))] + [len(str(v)) for v in df[col].dropna()])
        ws.column_dimensions[get_column_letter(c)].width = min(mx + 2, 40)
    ws.freeze_panes = "A2"


# =============================================================================
# TAB RESOLUTION
# =============================================================================

def find_tab(wb_or_sheetnames, key: str) -> Optional[str]:
    """Return the actual tab name present for a logical key, or None."""
    names = wb_or_sheetnames.sheetnames if hasattr(wb_or_sheetnames, "sheetnames") else list(wb_or_sheetnames)
    for alias in TAB_ALIASES.get(key, ()):
        if alias in names:
            return alias
    return None


# =============================================================================
# REFERENCE LOADERS
# =============================================================================

def load_cutsheet(path: str) -> pd.DataFrame:
    """Load the Installation Sheet tab from the JPB19 cutsheet."""
    return pd.read_excel(path, sheet_name="Installation Sheet")


def build_cutsheet_lookup(cut_df: pd.DataFrame) -> dict:
    """
    Build (Hostname, Interface) → {Source_port, DMARC1, DMARC2, Destination_port, ...}
    Only includes columns that actually exist in the cutsheet (handles schema evolution).
    """
    candidate_cols = ["Source_port", "DMARC1", "DMARC2", "Destination_port"]
    fill_cols = [c for c in candidate_cols if c in cut_df.columns]
    lookup = {}
    for _, row in cut_df.iterrows():
        key = (str(row["Hostname"]).strip(), str(row["Interface"]).strip())
        lookup[key] = {c: row[c] for c in fill_cols}
    lookup["__fill_cols__"] = fill_cols
    return lookup


def build_z_lookup(cut_df: pd.DataFrame) -> dict:
    """Key: (Z Hostname, Z Interface) → full row (for Possible columns in Mismatches)."""
    lookup = {}
    for _, row in cut_df.iterrows():
        key = (str(row["Z Hostname"]).strip(), str(row["Z Interface"]).strip())
        lookup[key] = row
    return lookup


def paired_subport(iface: str) -> Optional[str]:
    """Return mate for s0/s1 or s2/s3 sub-ports (used for Z-side fallback)."""
    pairs = {"s0": "s1", "s1": "s0", "s2": "s3", "s3": "s2"}
    for suffix, mate in pairs.items():
        if iface.endswith(suffix):
            return iface[: -len(suffix)] + mate
    return None


# =============================================================================
# MISMATCH PAIR HIGHLIGHTING (pure, unchanged behaviour)
# =============================================================================

def highlight_mismatch_pairs(wb: Workbook, log: Callable[[str], None] = lambda m: None) -> None:
    """Find reciprocal swap pairs in Mismatches and highlight them orange/yellow."""
    if "Mismatches" not in wb.sheetnames:
        return
    ws = wb["Mismatches"]
    if ws.max_row < 3:
        return

    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    def idxs(names):
        return [header.index(n) + 1 for n in names if n in header]

    a_cols = idxs(["Hostname", "Interface", "L/R", "Rack", "Elevation",
                   "Source_port", "Destination_port"])
    p_cols = idxs(["Possible Hostname", "Possible Interface", "Possible L/R",
                   "Possible Rack", "Possible Elevation",
                   "Possible Source_port", "Possible Destination_port"])
    orange_cols = idxs(["Hostname", "Interface", "L/R", "Rack", "Elevation",
                        "Source_port", "Destination_port", "Expected Hostname",
                        "Exp. Interface", "Exp. L/R", "Exp. Rack", "Exp. Elevation"])
    if not a_cols or not p_cols:
        return

    ncol, nrow = ws.max_column, ws.max_row
    rows = [[ws.cell(row=r, column=c).value for c in range(1, ncol + 1)]
            for r in range(2, nrow + 1)]

    def key(rv, cols):
        return tuple(str(rv[c - 1] if rv[c - 1] is not None else "").strip() for c in cols)

    a_keys = [key(rv, a_cols) for rv in rows]
    p_keys = [key(rv, p_cols) for rv in rows]

    n = len(rows)
    partner = [None] * n
    for i in range(n):
        if partner[i] is not None or not any(a_keys[i]):
            continue
        for j in range(i + 1, n):
            if partner[j] is not None:
                continue
            if a_keys[i] == p_keys[j] and a_keys[j] == p_keys[i]:
                partner[i] = j
                partner[j] = i
                break

    placed = [False] * n
    order = []
    for i in range(n):
        if placed[i] or partner[i] is None:
            continue
        j = partner[i]
        order.append(i); placed[i] = True
        if not placed[j]:
            order.append(j); placed[j] = True
    for i in range(n):
        if not placed[i]:
            order.append(i); placed[i] = True

    paired = {i for i in range(n) if partner[i] is not None}
    npairs = len(paired) // 2

    pair_no = {}
    counter = 0
    for i in range(n):
        j = partner[i]
        if j is not None and i < j:
            pair_no[i] = counter
            pair_no[j] = counter
            counter += 1

    bd = thin_border()
    pink_names = {"Possible Hostname", "Possible Interface", "Possible L/R",
                  "Possible Rack", "Possible Elevation", "Possible Source_port",
                  "Possible DMARC1", "Possible DMARC2", "Possible Destination_port",
                  "Z Hostname", "Z Interface", "Z L/R", "Z Rack", "Z Elevation"}
    pink_idx = {header.index(nm) + 1 for nm in pink_names if nm in header}
    orange_set = set(orange_cols)

    for out_off, src_i in enumerate(order):
        r = out_off + 2
        rv = rows[src_i]
        is_pair = src_i in paired
        pair_fill = (ORANGE_FILL if pair_no.get(src_i, 0) % 2 == 0 else YELLOW_FILL)
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c, value=rv[c - 1])
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = bd
            if is_pair and c in orange_set:
                cell.fill = pair_fill
            elif c in pink_idx:
                cell.fill = PINK_FILL
            else:
                cell.fill = NO_FILL

    log(f"  · Highlighted {npairs} mismatch pair(s) (alternating orange/yellow)")


# =============================================================================
# CORE PROCESSING (refactored from original process_file — in-memory only)
# =============================================================================

def _build_formatted_workbook(
    input_path: str,
    cut_df: pd.DataFrame,
    log: Callable[[str], None] = lambda m: None
) -> Workbook:
    """
    Perform the complete JPB19 GFAB transformation on an input validation file.
    Returns a fully formatted openpyxl Workbook (ready to .save()).
    All original behaviour and column ordering is preserved.
    """
    wb = load_workbook(input_path)

    cutsheet_lookup = build_cutsheet_lookup(cut_df)
    z_lookup        = build_z_lookup(cut_df)

    # ── 1. Split lldp tab → Downlinks / Mismatches ──────────────────────────
    mis_orig_df = None
    lldp_tab = find_tab(wb, "lldp")
    if lldp_tab:
        log(f"  · Splitting {lldp_tab} → Downlinks / Mismatches")
        df = pd.read_excel(input_path, sheet_name=lldp_tab)
        down_df     = df[df["Act. Interface"] == "interface down"].copy()
        mis_orig_df = df[df["Act. Interface"].str.startswith("swp", na=False)].copy()

        drop = ["Active Host", "Act. Interface", "Act. Rack", "Act. Elevation"]
        down_df.drop(columns=[c for c in drop if c in down_df.columns], inplace=True)

        exp_cols = ["Expected Hostname", "Exp. Interface", "Exp. Rack", "Exp. Elevation"]
        present_exp = [c for c in exp_cols if c in down_df.columns]
        if present_exp:
            rest = [c for c in down_df.columns if c not in present_exp]
            down_df = down_df[rest + present_exp]

        del wb[lldp_tab]
        write_sheet(wb, "Downlinks", down_df)
        write_sheet(wb, "Mismatches", mis_orig_df.drop(
            columns=[c for c in [] if c in mis_orig_df.columns]))

    # ── 2. Optics tab ───────────────────────────────────────────────────────
    optics_src = find_tab(wb, "optics")
    if optics_src:
        log(f"  · Processing Optics tab ({optics_src})")
        drop_cols = {"Transceiver", "Channel",
                     "Min Threshold (dBm)", "Max Threshold (dBm)",
                     "PP_A", "PP_Z", "Z_end_host", "Z_end_intf",
                     "rack_z", "Z_Rack", "Z_Elevation", "Index",
                     "Status", "Placement Group"}
        optics_df = pd.read_excel(input_path, sheet_name=optics_src)
        optics_df.drop(columns=[c for c in drop_cols if c in optics_df.columns], inplace=True)

        leading = [c for c in ("Metric", "Measured (dBm)") if c in optics_df.columns]
        if leading:
            rest = [c for c in optics_df.columns if c not in leading]
            optics_df = optics_df[leading + rest]

        del wb[optics_src]
        write_sheet(wb, "Optics", optics_df)
        wb["Optics"].freeze_panes = "C2" if len(leading) >= 2 else "B2"

    # ── 3. Remove interfaces tab ────────────────────────────────────────────
    interfaces_tab = find_tab(wb, "interfaces")
    if interfaces_tab:
        log(f"  · Removing {interfaces_tab}")
        del wb[interfaces_tab]

    # ── 3a. Remove other unwanted source tabs ───────────────────────────────
    drop_lower = {t.lower() for t in TABS_TO_REMOVE}
    for existing in list(wb.sheetnames):
        if existing.lower() in drop_lower:
            log(f"  · Removing {existing}")
            del wb[existing]

    # ── 3b. combined_fec: promote Lock Status + Pre-FEC BER ─────────────────
    fec_tab = find_tab(wb, "combined_fec")
    if fec_tab:
        log(f"  · Reordering {fec_tab} (Lock Status, Pre-FEC BER first)")
        fec_df = pd.read_excel(input_path, sheet_name=fec_tab)

        def _norm(s):
            return (str(s)
                    .replace("\u2011", "-")
                    .replace("\u2013", "-")
                    .replace("\u2014", "-")
                    .strip()
                    .lower())

        wanted = ["lock status", "pre-fec ber"]
        front = []
        for target in wanted:
            for col in fec_df.columns:
                if _norm(col) == target and col not in front:
                    front.append(col)
                    break

        if front:
            rest = [c for c in fec_df.columns if c not in front]
            fec_df = fec_df[front + rest]
            del wb[fec_tab]
            write_sheet(wb, "combined_fec", fec_df)
        else:
            log("    ⚠ Lock Status / Pre-FEC BER not found — leaving combined_fec as-is")

    # ── 4. Reorder primary tabs to the end ──────────────────────────────────
    desired  = ["Downlinks", "Mismatches", "Optics", "combined_fec"]
    existing = [s for s in desired if s in wb.sheetnames]
    others   = [s for s in wb.sheetnames if s not in desired]
    wb._sheets = [wb[n] for n in others + existing]

    # ── 5. Insert L/R columns (from cutsheet) ───────────────────────────────
    log("  · Adding L/R mapped columns (from cutsheet)")
    lr_from_cut = {}
    for _, row in cut_df.iterrows():
        iface = str(row.get("Interface", "") or "").strip()
        lr    = str(row.get("L/R", "") or "").strip()
        if iface and lr:
            lr_from_cut[iface] = lr
        z_iface = str(row.get("Z Interface", "") or "").strip()
        z_lr    = str(row.get("Z L/R", "") or "").strip()
        if z_iface and z_lr:
            lr_from_cut[z_iface] = z_lr

    lr_name_for = {
        "Interface":   "L/R",
        "Z Interface": "Z L/R",
        "Exp. Interface": "Exp. L/R",
    }
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        targets = [(i + 1, h) for i, h in enumerate(header) if h in lr_name_for]
        for col_idx, col_name in sorted(targets, reverse=True):
            new_name = lr_name_for[col_name]
            ws.insert_cols(col_idx + 1)
            header_cell(ws.cell(row=1, column=col_idx + 1), new_name)
            for r in range(2, ws.max_row + 1):
                val = str(ws.cell(row=r, column=col_idx).value or "").strip()
                ws.cell(row=r, column=col_idx + 1, value=lr_from_cut.get(val, ""))
                ws.cell(row=r, column=col_idx + 1).border = thin_border()
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = 10

    # ── 5b. Fill Source_port / DMARC* / Destination_port from cutsheet ──────
    fill_cols = cutsheet_lookup.get("__fill_cols__",
                                    ["Source_port", "DMARC1", "DMARC2", "Destination_port"])
    log(f"  · Filling {', '.join(fill_cols) or '(no cutsheet fill cols)'} (match on Hostname + Interface)")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if not all(c in header for c in ["Hostname", "Interface"]):
            continue

        anchor = (header.index("Elevation") + 1) if "Elevation" in header else len(header)
        insert_at = anchor + 1
        for col_name in fill_cols:
            if col_name in header:
                continue
            ws.insert_cols(insert_at)
            header_cell(ws.cell(row=1, column=insert_at), col_name)
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=insert_at).border = thin_border()
            ws.column_dimensions[get_column_letter(insert_at)].width = max(len(col_name) + 2, 14)
            insert_at += 1

        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        host_c, int_c = header.index("Hostname") + 1, header.index("Interface") + 1
        fill_idx = {c: header.index(c) + 1 for c in fill_cols}
        for r in range(2, ws.max_row + 1):
            host  = str(ws.cell(row=r, column=host_c).value or "").strip()
            iface = str(ws.cell(row=r, column=int_c).value or "").strip()
            match = cutsheet_lookup.get((host, iface))
            if match:
                for col_name, col_idx in fill_idx.items():
                    val = match.get(col_name)
                    if val is not None and not (isinstance(val, float) and pd.isna(val)):
                        ws.cell(row=r, column=col_idx, value=val)

    # ── 6c. Fill Z-side info (Z Hostname etc.) in designated tabs ───────────
    Z_COLS = ["Z Hostname", "Z Interface", "Z L/R", "Z Rack", "Z Elevation"]
    z_available = [c for c in Z_COLS if c in cut_df.columns]

    z_by_host_int = {}
    for _, row in cut_df.iterrows():
        k = (str(row["Hostname"]).strip(), str(row["Interface"]).strip())
        z_by_host_int[k] = {c: row[c] for c in z_available}

    if z_available and Z_FILL_TABS:
        for tab in Z_FILL_TABS:
            if tab not in wb.sheetnames:
                continue
            ws_z = wb[tab]
            header = [ws_z.cell(row=1, column=c).value for c in range(1, ws_z.max_column + 1)]
            if not all(c in header for c in ["Hostname", "Interface"]):
                continue
            log(f"  · Filling Z-side info in {tab}: {', '.join(z_available)}")

            if "Destination_port" in header:
                anchor = header.index("Destination_port") + 1
            elif "Elevation" in header:
                anchor = header.index("Elevation") + 1
            else:
                anchor = len(header)
            insert_at = anchor + 1
            for col_name in z_available:
                if col_name in header:
                    continue
                ws_z.insert_cols(insert_at)
                header_cell(ws_z.cell(row=1, column=insert_at), col_name)
                for r in range(2, ws_z.max_row + 1):
                    ws_z.cell(row=r, column=insert_at).border = thin_border()
                ws_z.column_dimensions[get_column_letter(insert_at)].width = max(len(col_name) + 2, 14)
                insert_at += 1

            header = [ws_z.cell(row=1, column=c).value for c in range(1, ws_z.max_column + 1)]
            host_c, int_c = header.index("Hostname") + 1, header.index("Interface") + 1
            fill_idx = {c: header.index(c) + 1 for c in z_available}
            for r in range(2, ws_z.max_row + 1):
                host  = str(ws_z.cell(row=r, column=host_c).value or "").strip()
                iface = str(ws_z.cell(row=r, column=int_c).value or "").strip()
                match = z_by_host_int.get((host, iface))
                if match:
                    for col_name, col_idx in fill_idx.items():
                        val = match.get(col_name)
                        if val is not None and not (isinstance(val, float) and pd.isna(val)):
                            ws_z.cell(row=r, column=col_idx, value=val)

    # ── 6b. Mismatches: Possible columns + Active Z columns (pink) ──────────
    if "Mismatches" in wb.sheetnames:
        log("  · Building Possible + Active Z columns in Mismatches")
        bd = thin_border()

        # Re-read original lldp to build act_lookup (tabs have been deleted from wb)
        act_lookup = {}
        src_sheets = pd.ExcelFile(input_path).sheet_names
        src_lldp = find_tab(src_sheets, "lldp")
        if src_lldp:
            orig_df  = pd.read_excel(input_path, sheet_name=src_lldp)
            mis_rows = orig_df[orig_df["Act. Interface"].str.startswith("swp", na=False)]
            for _, row in mis_rows.iterrows():
                key       = (str(row["Hostname"]).strip(), str(row["Interface"]).strip())
                act_host  = str(row["Active Host"]).strip()
                act_iface = str(row["Act. Interface"]).strip()

                cut_z_row = z_lookup.get((act_host, act_iface))
                if cut_z_row is None:
                    mate = paired_subport(act_iface)
                    if mate:
                        cut_z_row = z_lookup.get((act_host, mate))
                z_lr_val = ""
                if cut_z_row is not None:
                    raw = cut_z_row.get("Z L/R")
                    if raw is not None and not (isinstance(raw, float) and pd.isna(raw)):
                        z_lr_val = str(raw).strip()

                act_lookup[key] = {
                    "Z Hostname" : act_host,
                    "Z Interface": act_iface,
                    "Z L/R"      : z_lr_val,
                    "Z Rack"     : int(float(str(row["Act. Rack"]))),
                    "Z Elevation": int(float(str(row["Act. Elevation"]))),
                }

        ws_m   = wb["Mismatches"]
        header = [ws_m.cell(row=1, column=c).value for c in range(1, ws_m.max_column + 1)]

        act_drop = {"Active Host", "Act. Interface", "Act. Rack", "Act. Elevation"}
        for idx in sorted([i + 1 for i, h in enumerate(header) if h in act_drop], reverse=True):
            ws_m.delete_cols(idx)
        header = [ws_m.cell(row=1, column=c).value for c in range(1, ws_m.max_column + 1)]

        h_idx = header.index("Hostname") + 1
        i_idx = header.index("Interface") + 1

        act_rows = []
        for r in range(2, ws_m.max_row + 1):
            hn    = str(ws_m.cell(row=r, column=h_idx).value or "").strip()
            iface = str(ws_m.cell(row=r, column=i_idx).value or "").strip()
            act_rows.append(act_lookup.get((hn, iface), {}))

        # Possible columns (only those present in this cutsheet)
        possible_cols_all = [
            ("Possible Hostname",         "Hostname"),
            ("Possible Interface",        "Interface"),
            ("Possible L/R",             "L/R"),
            ("Possible Rack",             "Rack"),
            ("Possible Elevation",        "Elevation"),
            ("Possible Source_port",      "Source_port"),
            ("Possible DMARC1",           "DMARC1"),
            ("Possible DMARC2",           "DMARC2"),
            ("Possible Destination_port", "Destination_port"),
        ]
        cut_cols = set(cut_df.columns)
        possible_cols = [(out_col, src) for out_col, src in possible_cols_all if src in cut_cols]

        possible_data = {col: [] for col, _ in possible_cols}
        for act in act_rows:
            zh   = act.get("Z Hostname", "")
            zi   = act.get("Z Interface", "")
            match = z_lookup.get((zh, zi)) if zh else None
            if match is None and zh and zi:
                mate = paired_subport(zi)
                if mate:
                    match = z_lookup.get((zh, mate))
            for col, src in possible_cols:
                val = match.get(src, "") if match is not None else ""
                possible_data[col].append(val)

        pink_col_indices = []
        start = ws_m.max_column + 1
        for c_off, (col_name, _) in enumerate(possible_cols):
            col_idx = start + c_off
            pink_col_indices.append(col_idx)
            hdr = ws_m.cell(row=1, column=col_idx, value=col_name)
            hdr.font = Font(bold=True, name="Arial", size=10)
            hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            hdr.fill  = YELLOW_FILL
            hdr.border = bd
            ws_m.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 3, 14)
            for r_off, val in enumerate(possible_data[col_name]):
                cell = ws_m.cell(row=r_off + 2, column=col_idx,
                                 value=val if val != "" else None)
                cell.fill   = PINK_FILL
                cell.border = bd

        # Active Z columns
        act_z_cols = ["Z Hostname", "Z Interface", "Z L/R", "Z Rack", "Z Elevation"]
        start2 = ws_m.max_column + 1
        for c_off, col_name in enumerate(act_z_cols):
            col_idx = start2 + c_off
            pink_col_indices.append(col_idx)
            hdr = ws_m.cell(row=1, column=col_idx, value=col_name)
            hdr.font = Font(bold=True, name="Arial", size=10)
            hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            hdr.fill  = YELLOW_FILL
            hdr.border = bd
            ws_m.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 3, 14)
            for r_off, act in enumerate(act_rows):
                val = act.get(col_name, "")
                cell = ws_m.cell(row=r_off + 2, column=col_idx,
                                 value=val if val != "" else None)
                cell.fill   = PINK_FILL
                cell.border = bd

        # Store for later clear_and_border
        wb._pink_col_indices_for_mismatches = pink_col_indices   # type: ignore[attr-defined]

    # ── 6c. Strip unwanted global columns ───────────────────────────────────
    if COLUMNS_TO_REMOVE:
        log(f"  · Stripping columns: {', '.join(COLUMNS_TO_REMOVE)}")
        drop_set = set(COLUMNS_TO_REMOVE)
        for sheet_name in wb.sheetnames:
            ws_x = wb[sheet_name]
            header = [ws_x.cell(row=1, column=c).value for c in range(1, ws_x.max_column + 1)]
            to_drop = [i + 1 for i, h in enumerate(header) if h in drop_set]
            for idx in sorted(to_drop, reverse=True):
                ws_x.delete_cols(idx)

    # ── 7. Summary tab (first) ──────────────────────────────────────────────
    log("  · Creating Summary tab")

    tab_rack  = {}
    all_racks = set()
    center_s = Alignment(horizontal="center", vertical="center", wrap_text=False)
    bd_s     = thin_border()

    def _s(cell, value, bold=False, header=False):
        cell.value     = value
        cell.font      = Font(bold=bold, name="Arial", size=10)
        cell.alignment = center_s
        cell.border    = bd_s
        cell.fill      = YELLOW_FILL if header else NO_FILL

    for sname in wb.sheetnames:
        ws_tmp = wb[sname]
        hdr = [ws_tmp.cell(row=1, column=c).value for c in range(1, ws_tmp.max_column + 1)]
        if "Rack" not in hdr:
            tab_rack[sname] = {}
            continue
        rack_col = hdr.index("Rack") + 1
        counts = {}
        for r in range(2, ws_tmp.max_row + 1):
            val = ws_tmp.cell(row=r, column=rack_col).value
            if val is not None:
                try:
                    k = int(float(str(val)))
                    counts[k] = counts.get(k, 0) + 1
                    all_racks.add(k)
                except ValueError:
                    pass
        tab_rack[sname] = counts

    racks      = sorted(all_racks)
    tabs_order = list(wb.sheetnames)
    total_cols = 1 + len(racks) + 1

    for existing in list(wb.sheetnames):
        if existing.lower() == "summary":
            del wb[existing]
    wb.create_sheet("Summary", 0)
    ws_s = wb["Summary"]

    ws_s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_c = ws_s.cell(row=1, column=1, value="Tab Summary by Rack")
    title_c.font = Font(bold=True, name="Arial", size=13)
    title_c.alignment = center_s
    title_c.border    = bd_s
    title_c.fill      = YELLOW_FILL
    ws_s.row_dimensions[1].height = 28

    _s(ws_s.cell(row=2, column=1), "Tab Name", bold=True, header=True)
    for c, rack in enumerate(racks, start=2):
        _s(ws_s.cell(row=2, column=c), str(rack), bold=True, header=True)
    _s(ws_s.cell(row=2, column=total_cols), "Total", bold=True, header=True)

    rack_totals = {r: 0 for r in racks}
    data_tabs   = [n for n in tabs_order if n.lower() != "summary"]
    for i, tab_name in enumerate(data_tabs, start=3):
        _s(ws_s.cell(row=i, column=1), tab_name)
        row_total = 0
        for c, rack in enumerate(racks, start=2):
            count = tab_rack.get(tab_name, {}).get(rack, 0)
            _s(ws_s.cell(row=i, column=c), count if count > 0 else "")
            rack_totals[rack] += count
            row_total += count
        _s(ws_s.cell(row=i, column=total_cols), row_total, bold=True)

    tot_r = 3 + len(data_tabs)
    _s(ws_s.cell(row=tot_r, column=1), "TOTAL", bold=True)
    grand = 0
    for c, rack in enumerate(racks, start=2):
        _s(ws_s.cell(row=tot_r, column=c), rack_totals[rack], bold=True)
        grand += rack_totals[rack]
    _s(ws_s.cell(row=tot_r, column=total_cols), grand, bold=True)

    ws_s.column_dimensions["A"].width = 20
    for c in range(2, total_cols + 1):
        ws_s.column_dimensions[get_column_letter(c)].width = 14

    # ── 8. Borders + clear fills (preserve pink) ────────────────────────────
    log("  · Removing fills and applying borders")
    pink_col_indices = getattr(wb, "_pink_col_indices_for_mismatches", None)
    if pink_col_indices is None and "Mismatches" in wb.sheetnames:
        ws_m = wb["Mismatches"]
        m_header = [ws_m.cell(row=1, column=c).value for c in range(1, ws_m.max_column + 1)]
        Z_NAMES = {"Z Hostname", "Z Interface", "Z L/R", "Z Rack", "Z Elevation"}
        pink_col_indices = [
            i + 1 for i, h in enumerate(m_header)
            if (h and (str(h).startswith("Possible ") or h in Z_NAMES))
        ]

    for sheet_name in wb.sheetnames:
        pcols = pink_col_indices if sheet_name == "Mismatches" else []
        clear_and_border(wb[sheet_name], pink_cols=set(pcols))

    # ── 8c. Centre align everything ─────────────────────────────────────────
    log("  · Aligning all cells to middle-centre")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                cell.alignment = center_align

    # ── 8b. NOTE column + autofilter on every tab ───────────────────────────
    log("  · Adding NOTE column and filters to all tabs")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        col_idx = ws.max_column + 1
        hdr = ws.cell(row=1, column=col_idx, value="NOTE")
        hdr.font      = Font(bold=True, name="Arial", size=10)
        hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        hdr.fill      = YELLOW_FILL
        hdr.border    = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill   = NO_FILL
            cell.border = thin_border()
        if ws.max_row > 1 and ws.max_column > 0:
            ws.auto_filter.ref = ws.dimensions

    # ── 8d. Grey-out Optics rows that already appear in Downlinks ───────────
    if "Optics" in wb.sheetnames and "Downlinks" in wb.sheetnames:
        log("  · Greying out matched Optics rows")

        MATCH_COLS = [
            "Hostname", "Interface", "L/R", "Rack", "Elevation",
            "Source_port", "DMARC1", "DMARC2", "Destination_port",
            "Z Hostname", "Z Interface", "Z L/R", "Z Rack", "Z Elevation",
        ]
        GREY_FONT_COLOR = "FFD3D3D3"

        ws_dl = wb["Downlinks"]
        dl_header = [ws_dl.cell(row=1, column=c).value for c in range(1, ws_dl.max_column + 1)]
        dl_match_cols = [c for c in MATCH_COLS if c in dl_header]
        dl_col_idx    = {c: dl_header.index(c) + 1 for c in dl_match_cols}

        dl_keys = set()
        for r in range(2, ws_dl.max_row + 1):
            key = tuple(str(ws_dl.cell(row=r, column=dl_col_idx[c]).value or "").strip()
                        for c in dl_match_cols)
            dl_keys.add(key)

        ws_op = wb["Optics"]
        op_header = [ws_op.cell(row=1, column=c).value for c in range(1, ws_op.max_column + 1)]
        common_cols  = [c for c in dl_match_cols if c in op_header]
        op_col_idx   = {c: op_header.index(c) + 1 for c in common_cols}
        dl_col_idx_c = {c: dl_header.index(c) + 1 for c in common_cols}

        dl_keys_common = set()
        for r in range(2, ws_dl.max_row + 1):
            key = tuple(str(ws_dl.cell(row=r, column=dl_col_idx_c[c]).value or "").strip()
                        for c in common_cols)
            dl_keys_common.add(key)

        for r in range(2, ws_op.max_row + 1):
            op_key = tuple(str(ws_op.cell(row=r, column=op_col_idx[c]).value or "").strip()
                           for c in common_cols)
            if op_key in dl_keys_common:
                for c in range(1, ws_op.max_column + 1):
                    cell = ws_op.cell(row=r, column=c)
                    cell.font = Font(
                        bold=cell.font.bold if cell.font else False,
                        name=(cell.font.name if cell.font else None) or "Arial",
                        size=(cell.font.size if cell.font else None) or 10,
                        color=GREY_FONT_COLOR,
                    )

    # ── 8e. Autofit every sheet ─────────────────────────────────────────────
    log("  · Expanding all columns and rows")
    for sheet_name in wb.sheetnames:
        autofit_sheet(wb[sheet_name])

    # ── 8f. Highlight reciprocal mismatch pairs ─────────────────────────────
    log("  · Highlighting reciprocal mismatch pairs")
    highlight_mismatch_pairs(wb, log)

    # Clean up the temporary attribute
    if hasattr(wb, "_pink_col_indices_for_mismatches"):
        delattr(wb, "_pink_col_indices_for_mismatches")

    return wb


# =============================================================================
# FILENAME HELPERS (Streamlit-friendly)
# =============================================================================

def _compute_rack_based_name(wb: Workbook) -> Optional[str]:
    """Return e.g. '12+14.xlsx' from the two most common Rack numbers across all tabs."""
    all_racks: List[int] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if "Rack" not in header:
            continue
        rc = header.index("Rack") + 1
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=rc).value
            if val is not None:
                try:
                    all_racks.append(int(float(str(val))))
                except (ValueError, TypeError):
                    pass
    if not all_racks:
        return None
    top2 = [str(r) for r, _ in Counter(all_racks).most_common(2)]
    return "+".join(top2) + ".xlsx"


def _safe_base_name(path: str) -> str:
    base = os.path.splitext(os.path.basename(path))[0]
    # Remove any previous _formatted or _JPB19_GFAB suffix to avoid double suffixes
    for suffix in ("_formatted", "_JPB19_GFAB", "_GFAB"):
        if base.endswith(suffix):
            base = base[: -len(suffix)]
    return base


# =============================================================================
# PUBLIC API (what the Streamlit page calls)
# =============================================================================

def process_gfab_validation(
    input_path: str,
    cutsheet_path: str,
    output_name: Optional[str] = None
) -> Tuple[bytes, str]:
    """
    Process a single JPB19 GFAB validation export + cutsheet.

    Returns:
        (xlsx_bytes, suggested_output_filename)
    """
    cut_df = load_cutsheet(cutsheet_path)
    wb = _build_formatted_workbook(input_path, cut_df)

    if output_name is None:
        rack_name = _compute_rack_based_name(wb)
        if rack_name:
            output_name = rack_name
        else:
            base = _safe_base_name(input_path)
            output_name = f"{base}_JPB19_GFAB.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), output_name


def process_multiple_files(
    input_paths: List[str],
    cutsheet_path: str,
) -> bytes:
    """
    Process several validation files against the same cutsheet.
    Returns a ZIP file (bytes) containing one .xlsx per input.
    """
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        for i, path in enumerate(input_paths, 1):
            xlsx_bytes, filename = process_gfab_validation(path, cutsheet_path)
            # Ensure unique names inside the zip
            if filename in zipf.namelist():
                stem, ext = os.path.splitext(filename)
                filename = f"{stem}_{i}{ext}"
            zipf.writestr(filename, xlsx_bytes)
    zip_buffer.seek(0)
    return zip_buffer.getvalue()


# Convenience re-exports for anyone who wants the lower-level pieces
__all__ = [
    "process_gfab_validation",
    "process_multiple_files",
    "load_cutsheet",
    "build_cutsheet_lookup",
    "build_z_lookup",
    "_build_formatted_workbook",   # for advanced users / tests
]
