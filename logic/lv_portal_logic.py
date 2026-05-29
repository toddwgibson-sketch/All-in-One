"""
lv_portal_logic.py

Clean, importable core logic for JBP15 (and similar) T0-to-Host / LV Portal reports.

ARCHITECTURAL RULE (inherited from AIO):
    No multiprocessing allowed anywhere in this module or any processor.
    See aio_logic.py for explanation (breaks Streamlit Cloud + session model).
"""

This module provides the standard AIO interface so the universal formatter
can dispatch to it automatically when the cutsheet + report indicate a
T0-to-Host style validation (common in QFAB / GPU compute environments).

Public API (matches all other processors):
    process_lv_portal_validation(input_path, cutsheet_path, output_name=None) -> (bytes, suggested_filename)
    process_multiple_lv_portal_files(input_paths, cutsheet_path) -> zip_bytes
"""

from __future__ import annotations

import io
import os
import re
import tempfile
import zipfile
from typing import List, Optional, Tuple, Dict, Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Re-use some helpers from the existing clean style if possible
# (we keep this module self-contained for now)

# Colors & styles (matching the mature version the user had)
YELLOW = "FFFF00"
WHITE = "FFFFFF"
HDR_BG = "1F4E79"

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
HEADER_FONT = Font(bold=True, color=WHITE, name="Arial", size=9)


def _header_fill():
    return PatternFill(start_color=HDR_BG, end_color=HDR_BG, fill_type="solid")


def _yellow_fill():
    return PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")


# =============================================================================
# Core helpers (ported/adapted from the proven T0-to-Host logic)
# =============================================================================

def is_compute(name: Any) -> bool:
    return "compute" in str(name or "").lower()


def is_real_compute_host(name: Any) -> bool:
    s = str(name or "").strip().lower()
    if not s or s == "missing":
        return False
    return "compute" in s and s.startswith("jbp15-")


def build_compute_lookup(cutsheet_paths: List[str]) -> Dict:
    """
    Build T0/T1/PP lookup from the JBP15-style T0-to-Host cutsheet(s).
    This is the heart of the LV Portal T0 processing.
    """
    t0: Dict = {}
    t1: Dict = {}
    t1_rev: Dict = {}
    t0_to_pp: Dict = {}

    for path in cutsheet_paths:
        wb = load_workbook(path, read_only=True)
        sheet = next(
            (wb[n] for n in wb.sheetnames if "installation" in n.lower()),
            wb[wb.sheetnames[0]],
        )

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 9:
                continue

            lbl = str(row[0] or "").strip()
            dev_a = str(row[1] or "").strip()
            rack_a = str(row[2] or "").strip()
            src = str(row[3] or "").strip()
            dmarc1 = str(row[4] or "").strip()
            dmarc2 = str(row[5] or "").strip()
            dest = str(row[6] or "").strip()
            dev_b = str(row[7] or "").strip()
            rack_b = str(row[8] or "").strip()
            t1_lbl = str(row[10] or "").strip() if len(row) > 10 else ""

            if dev_a and lbl and re.match(r"\d+[LR]$", lbl):
                parts = dev_a.split()
                if len(parts) == 2:
                    k = (parts[0], parts[1])
                    t0[k] = lbl
                    t1[k] = t1_lbl
                    t0_to_pp[k] = {
                        "source_port": src,
                        "dmarc1": dmarc1,
                        "dmarc2": dmarc2,
                        "dest_port": dest,
                        "rack_b": rack_b,
                        "t1_lbl_pp": t1_lbl,
                    }

            if dev_b and " " in dev_b:
                parts = dev_b.split()
                if len(parts) == 2:
                    t1_rev[(parts[0], parts[1])] = {
                        "t0_lbl": lbl,
                        "rack_a": rack_a,
                        "source_port": src,
                        "dmarc1": dmarc1,
                        "dmarc2": dmarc2,
                        "dest_port": dest,
                        "rack_b": rack_b,
                        "t1_lbl": t1_lbl,
                    }

        wb.close()

    return {"t0": t0, "t1": t1, "t1_rev": t1_rev, "t0_to_pp": t0_to_pp}


# =============================================================================
# Main processing entry points (AIO contract)
# =============================================================================

def process_lv_portal_validation(
    input_path: str,
    cutsheet_path: str,
    output_name: Optional[str] = None,
) -> Tuple[bytes, str]:
    """
    Process one LV Portal / T0-to-Host validation export against its cutsheet.

    Returns (xlsx_bytes, suggested_filename)
    """
    # For now we delegate to a lightweight wrapper.
    # Full rich port of the 2000+ line logic will live here.
    # This stub ensures the AIO doesn't crash and gives useful feedback.

    # TODO (high priority): fully port the sheet builders from
    # t0 to host version 2 / lv_portal_logic.py

    # As a temporary bridge while we complete the port:
    # If the user forces "lv_portal", we currently give a clear message in the UI.
    # Once the real implementation is here, this function will do the real work.

    raise NotImplementedError(
        "Full LV Portal T0-to-Host processing is being integrated into the AIO. "
        "The detection is now correct (LV_PORTAL_T0). "
        "The actual report generation for this report type is the next piece being wired in."
    )


def process_multiple_lv_portal_files(
    input_paths: List[str],
    cutsheet_path: str,
) -> bytes:
    """Process several LV Portal files and return a ZIP."""
    # Same temporary behavior as above
    raise NotImplementedError(
        "Multi-file LV Portal T0-to-Host support is part of the ongoing integration."
    )


# Convenience for the AIO registry
__all__ = [
    "process_lv_portal_validation",
    "process_multiple_lv_portal_files",
    "build_compute_lookup",
]
