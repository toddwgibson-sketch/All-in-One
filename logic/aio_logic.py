"""
jpb19_aio_logic.py

ALL-IN-ONE (AIO) Validation Formatter - Master Logic

This is the universal entry point for any hall / any fabric type.

Core idea:
- The CUTSHEET is the source of truth for "what kind of thing this is".
- We inspect the cutsheet(s) first to determine hall, fabric type (GFAB/CFAB/HOPS/LV-Portal/etc.),
  source system (Slack export vs LVV Portal), and which rich processor to use.
- Then we dispatch to the appropriate specialized logic we already built
  (jpb19_gfab_logic, jpb19_hops_logic, jpb19_cfab_logic, etc.).
- The goal is one button for validators no matter what room or what export format they receive.

Historical raw material (old Claude scripts for many halls) lives in the `legacy/` subfolder right here.
This folder (`All in One`) is now the permanent clean home of the universal AIO Validation Formatter.

Public API:
    inspect_cutsheet(cutsheet_path) -> DetectionResult
    process_aio_validation(input_paths, cutsheet_path, override_type=None, ...) -> (bytes, filename, info)
    process_multiple_aio(...) -> zip_bytes
"""

from __future__ import annotations

import io
import os
import zipfile
from dataclasses import dataclass, field
from typing import List, Optional, Dict, Any, Callable, Tuple

import pandas as pd
from openpyxl import load_workbook

# Import all our specialized processors (the "master logic from all the sheets")
try:
    from . import gfab_logic
    from . import hops_logic
    from . import cfab_logic
    from . import ipr_logic
except ImportError as e:
    raise ImportError(f"Missing one of the specialized logic modules inside logic/: {e}")


# =============================================================================
# DETECTION RESULT
# =============================================================================

@dataclass
class DetectionResult:
    """What we learned from inspecting the cutsheet(s)."""
    hall: str = "UNKNOWN"                    # JPB19, JBP15, SYD20, etc.
    fabric_type: str = "UNKNOWN"             # GFAB, CFAB, HOPS, LV_PORTAL_T0, IPR, etc.
    source: str = "UNKNOWN"                  # slack, lvv_portal, mixed, unknown
    recommended_processor: str = "unknown"   # gfab, hops, cfab, ipr, lv_portal, etc.
    confidence: float = 0.0                  # 0.0 - 1.0
    reasons: List[str] = field(default_factory=list)
    cutsheet_metadata: Dict[str, Any] = field(default_factory=dict)
    suggested_cutsheet_role: str = "primary" # primary, combined, installation, etc.

    def is_confident(self) -> bool:
        return self.confidence >= 0.75 and self.recommended_processor != "unknown"

    def summary(self) -> str:
        return (
            f"Hall: {self.hall} | Type: {self.fabric_type} | "
            f"Source: {self.source} | Processor: {self.recommended_processor} "
            f"(confidence: {self.confidence:.0%})"
        )


# =============================================================================
# CUTSHEET INSPECTION (the key intelligence)
# =============================================================================

def inspect_cutsheet(cutsheet_path: str) -> DetectionResult:
    """
    Inspect a cutsheet and figure out what kind of validation this is for.
    This is the heart of the AIO.
    """
    result = DetectionResult()
    reasons = []

    # Load the first sheet (most cutsheets put important metadata on sheet 0 or "Installation Sheet")
    try:
        wb = load_workbook(cutsheet_path, read_only=True, data_only=True)
        first_sheet = wb[wb.sheetnames[0]]
    except Exception as e:
        result.reasons.append(f"Could not read cutsheet: {e}")
        return result

    # Collect text from the first ~30 rows for keyword scanning (very effective)
    text_blob = ""
    sheet_names = [s.lower() for s in wb.sheetnames]

    for row in first_sheet.iter_rows(min_row=1, max_row=40, values_only=True):
        for cell in row:
            if cell:
                text_blob += " " + str(cell).lower()

    text_lower = text_blob.lower()

    # --- Hall detection ---
    if "jpb19" in text_lower or "jpb-19" in text_lower:
        result.hall = "JPB19"
        reasons.append("Found 'JPB19' in cutsheet")
    elif "jbp15" in text_lower or "jbp-15" in text_lower:
        result.hall = "JBP15"
        reasons.append("Found 'JBP15' in cutsheet")
    elif "syd20" in text_lower or "syd-20" in text_lower:
        result.hall = "SYD20"
        reasons.append("Found 'SYD20' in cutsheet")
    else:
        # Try to extract from filename as fallback
        fname = os.path.basename(cutsheet_path).lower()
        if "jpb19" in fname:
            result.hall = "JPB19"
        elif "jbp15" in fname:
            result.hall = "JBP15"

    # --- Fabric / Pathway type detection (most important) ---
    # Priority 1: Explicit signals (future-proof — user said the cutsheet will carry this)
    explicit_type = None
    for row in first_sheet.iter_rows(min_row=1, max_row=20, values_only=True):
        row_text = " ".join(str(c).lower() for c in row if c)
        if "aio_type" in row_text or "processor:" in row_text or "fabric_type:" in row_text:
            # Look for explicit markers like "AIO_Type: hops" or "Processor = gfab"
            for cell in row:
                if cell and isinstance(cell, str):
                    low = cell.lower().strip()
                    if low.startswith(("aio_type:", "processor:", "fabric_type:", "type:")):
                        val = low.split(":", 1)[1].strip()
                        explicit_type = val
                        break
        if explicit_type:
            break

    if explicit_type:
        result.fabric_type = explicit_type.upper()
        result.recommended_processor = explicit_type.lower().replace(" ", "_")
        reasons.append(f"Explicit type found in cutsheet: {explicit_type}")

    # Priority 2: Strong heuristics from known patterns across halls
    if not explicit_type:
        if any(k in text_lower for k in ["hops", "gpu", "ll dp mismatch (gpu)", "optic errors (gpu)"]):
            result.fabric_type = "HOPS"
            result.recommended_processor = "hops"
            reasons.append("HOPS/GPU indicators found (LLDP Mismatch (GPU), Optic Errors (GPU), etc.)")

        elif any(k in text_lower for k in ["t3-t2", "t2-t1-t0", "t3 t2", "cfab"]):
            result.fabric_type = "CFAB"
            result.recommended_processor = "cfab"
            reasons.append("CFAB style detected (T3-T2 / T2-T1-T0 language or structure)")

        elif any(k in text_lower for k in ["gfab", "full_path_lldp_with_int_down", "installation sheet"]):
            if result.fabric_type == "UNKNOWN":
                result.fabric_type = "GFAB"
                result.recommended_processor = "gfab"
                reasons.append("GFAB / standard fabric indicators")

        elif any(k in text_lower for k in ["lv portal", "t0-to-host", "t0 to host", "qfabt0"]):
            result.fabric_type = "LV_PORTAL"
            result.recommended_processor = "lv_portal"
            reasons.append("LV Portal / T0-to-Host / QFAB T0 style detected")

        # Additional patterns from the full collection of old scripts
        elif "optic errors" in text_lower and "fec_ber" in text_lower and "interface down" in text_lower:
            # Classic LV Portal non-GPU pattern (no "(GPU)" suffix)
            if result.fabric_type == "UNKNOWN":
                result.fabric_type = "LV_PORTAL"
                result.recommended_processor = "lv_portal"
                reasons.append("LV Portal classic sheet set (Optic Errors + FEC_BER + Interface Down)")

    # Sheet-name based fallback (very reliable for HOPS vs others)
    if result.fabric_type == "UNKNOWN":
        if any("gpu" in s for s in sheet_names):
            result.fabric_type = "HOPS"
            result.recommended_processor = "hops"
            reasons.append("GPU-specific sheets present")

    # --- Source system detection (Slack vs LVV Portal) ---
    if "slack" in text_lower:
        result.source = "slack"
        reasons.append("Explicit 'Slack' reference in cutsheet")
    elif any(k in text_lower for k in ["lvv", "lv portal", "validation portal", "full path report"]):
        result.source = "lvv_portal"
        reasons.append("LVV Portal indicators found")

    # Sheet name heuristics (very reliable)
    if any("gpu" in s for s in sheet_names):
        if result.fabric_type == "UNKNOWN":
            result.fabric_type = "HOPS"
            result.recommended_processor = "hops"
            reasons.append("GPU-specific sheets present in cutsheet")

    if "combined" in " ".join(sheet_names) and result.fabric_type == "UNKNOWN":
        # Many HOPS and some CFAB use "combined" cutsheets
        result.fabric_type = "HOPS_OR_CFAB"
        reasons.append("Combined cutsheet naming convention detected")

    # Confidence scoring
    confidence = 0.3  # base
    if result.hall != "UNKNOWN":
        confidence += 0.25
    if result.fabric_type != "UNKNOWN":
        confidence += 0.35
    if result.source != "UNKNOWN":
        confidence += 0.1
    result.confidence = min(confidence, 1.0)

    result.reasons = reasons
    result.cutsheet_metadata = {
        "sheets": wb.sheetnames[:8],
        "first_sheet_name": wb.sheetnames[0],
    }

    return result


# =============================================================================
# PROCESSOR DISPATCH
# =============================================================================

_PROCESSORS: Dict[str, Callable] = {
    "gfab": gfab_logic.process_gfab_validation,
    "hops": hops_logic.process_hops_validation,
    "cfab": cfab_logic.process_cfab_validation,
    "ipr": ipr_logic.process_ipr_validation,
    # "lv_portal": ... (we can wire the JBP15 one later)
}


def get_processor(detection: DetectionResult) -> Optional[Callable]:
    """Return the actual processing function for this detection."""
    proc_name = detection.recommended_processor
    return _PROCESSORS.get(proc_name)


# =============================================================================
# PUBLIC AIO API
# =============================================================================

def process_aio_validation(
    input_paths: List[str],
    cutsheet_path: str,
    override_processor: Optional[str] = None,
    output_name: Optional[str] = None,
) -> Tuple[bytes, str, Dict[str, Any]]:
    """
    The main AIO entry point.

    1. Inspect the cutsheet to understand what we're dealing with.
    2. (Optionally) let the caller override the processor.
    3. Dispatch to the correct rich logic.
    4. Return (bytes, suggested_filename, info_dict)
    """
    detection = inspect_cutsheet(cutsheet_path)

    processor_name = override_processor or detection.recommended_processor
    processor = _PROCESSORS.get(processor_name)

    info = {
        "detection": detection,
        "used_processor": processor_name,
        "overridden": bool(override_processor),
    }

    if processor is None:
        # Fallback: try GFAB as the most common "standard" path
        processor = gfab_logic.process_gfab_validation
        info["used_processor"] = "gfab (fallback)"
        info["warning"] = f"No processor for '{processor_name}'. Fell back to GFAB."

    # Most of our processors take (input_path, cutsheet_path)
    # For multiple inputs the caller should use process_multiple_aio
    if len(input_paths) == 1:
        xlsx_bytes, fname = processor(input_paths[0], cutsheet_path)
        if output_name:
            fname = output_name
        return xlsx_bytes, fname, info
    else:
        # Multi-file case - use the multi version if available
        multi_fn = getattr(
            processor.__module__ if hasattr(processor, "__module__") else None,
            "process_multiple_" + processor_name + "_files",
            None
        )
        # Simpler: just call the single one in a loop and zip (the pages usually do this)
        # For now return a clear message
        raise NotImplementedError(
            "For multiple input files please use process_multiple_aio_validation() "
            "or let the Streamlit page handle zipping."
        )


def process_multiple_aio_validation(
    input_paths: List[str],
    cutsheet_path: str,
    override_processor: Optional[str] = None,
) -> Tuple[bytes, Dict[str, Any]]:
    """
    Process multiple files and return a single ZIP.
    """
    detection = inspect_cutsheet(cutsheet_path)
    processor_name = override_processor or detection.recommended_processor

    # Try to find the multi-file function on the right module
    multi_fn = None
    if processor_name == "gfab":
        multi_fn = gfab_logic.process_multiple_files
    elif processor_name == "hops":
        multi_fn = hops_logic.process_multiple_hops_files
    elif processor_name == "cfab":
        multi_fn = cfab_logic.process_multiple_cfab_files
    elif processor_name == "ipr":
        multi_fn = ipr_logic.process_multiple_ipr_files

    info = {
        "detection": detection,
        "used_processor": processor_name,
        "overridden": bool(override_processor),
    }

    if multi_fn is None:
        # Fallback: process one by one using GFAB multi (or implement generic zip)
        multi_fn = gfab_logic.process_multiple_files
        info["warning"] = f"No multi-file processor for '{processor_name}'. Used GFAB multi as fallback."

    zip_bytes = multi_fn(input_paths, cutsheet_path)
    return zip_bytes, info


# Convenience re-export
__all__ = [
    "inspect_cutsheet",
    "DetectionResult",
    "process_aio_validation",
    "process_multiple_aio_validation",
    "get_processor",
]
